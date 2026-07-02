"""考纲百套卷 10 列（A–J）规划总表：解析（D1）、卷号编排（D3）、渲染。

列：A知识模块 B专题名称 C考点名称 D考点内容 E考点卷数量 F考点卷卷号
    G专题卷数量 H专题卷卷号 I综合卷数量 J综合卷卷号。
A/B/G/H/I/J 在同组内合并（仅首行有值），解析时按“最近非空值”向下承载。
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

HEADERS = ["知识模块", "专题名称", "考点名称", "掌握/理解/了解考点内容",
           "考点训练卷数量", "考点训练卷卷号", "专题训练卷数量", "专题训练卷卷号",
           "课程综合卷数量", "课程综合卷卷号"]
_COL_WIDTHS = [16, 18, 20, 46, 10, 12, 10, 12, 10, 12]
_FONT = "微软雅黑"

_VOL_ONE = re.compile(r"第\s*(\d+)\s*卷")
_VOL_RANGE = re.compile(r"第\s*(\d+)\s*[-—~]\s*(\d+)\s*卷")


def parse_volume_no(s: Any) -> int | None:
    """'第1卷' → 1；无法解析返回 None。"""
    m = _VOL_ONE.search(str(s or ""))
    return int(m.group(1)) if m else None


def parse_volume_range(s: Any) -> tuple[int, int] | None:
    """'第34-36卷' → (34,36)；'第31卷' → (31,31)；无法解析返回 None。"""
    txt = str(s or "")
    m = _VOL_RANGE.search(txt)
    if m:
        return int(m.group(1)), int(m.group(2))
    one = parse_volume_no(txt)
    return (one, one) if one else None


def _txt(v: Any) -> str:
    return str(v).strip() if v is not None else ""


# ---------------- 窄考点判定 D2（规则位）----------------
NARROW_THRESHOLD = 80


def parse_kpoint_counts(md_text: str) -> dict[str, int]:
    """解析《{课程}_知识点题目数量.md》，返回 {叶子知识点名: 合计题量}。

    叶子表格行形如：| 87637 | 人体触电常识 | 48 | 13 | ... | 79 |
    取行首非空的“知识点名”与行末“合计”。
    """
    counts: dict[str, int] = {}
    for line in md_text.splitlines():
        if "|" not in line:
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) < 3:
            continue
        # 跳过表头/分隔行
        if cells[0] in ("节点ID", "序号") or set(cells[0]) <= {"-", ":"}:
            continue
        name = cells[1] if cells[0].isdigit() else cells[0]
        last = cells[-1]
        if name and last.lstrip("-").isdigit():
            counts[name] = int(last)
    return counts


def flag_narrow_points(rows: list[dict[str, Any]], counts: dict[str, int],
                       threshold: int = NARROW_THRESHOLD) -> list[dict[str, Any]]:
    """标注窄考点（合计 < 阈值）。就地写入 is_narrow / total_q，返回 rows。

    “内容相近才可合并”属语义判断，交由 LLM/人工在闸门确认；此处只做确定性的题量判定，
    并给出「同专题相邻」候选提示（same_theme_neighbor）。
    """
    def _lookup(name: str) -> int | None:
        if name in counts:
            return counts[name]
        for k, v in counts.items():  # 名称模糊匹配
            if k in name or name in k:
                return v
        return None

    for i, r in enumerate(rows):
        total = _lookup(r.get("point_name") or "")
        r["total_q"] = total
        r["is_narrow"] = (total is not None and total < threshold)
    # 相邻同专题窄考点候选（供人工/LLM 判断是否内容相近后合并，≤3）
    for i, r in enumerate(rows):
        if not r.get("is_narrow"):
            r["merge_candidate"] = False
            continue
        prev_ok = i > 0 and rows[i - 1].get("is_narrow") and rows[i - 1].get("theme") == r.get("theme")
        next_ok = i + 1 < len(rows) and rows[i + 1].get("is_narrow") and rows[i + 1].get("theme") == r.get("theme")
        r["merge_candidate"] = bool(prev_ok or next_ok)
    return rows


def _find_header_row(ws) -> int:
    for r in range(1, min(ws.max_row, 12) + 1):
        if _txt(ws.cell(r, 1).value).replace("\n", "").startswith("知识模块"):
            return r
    return 6


# ---------------- 解析 D1 ----------------
def parse_10col(path: str | Path) -> list[dict[str, Any]]:
    """解析 10 列规划总表为考点训练卷行（合并单元格按最近非空承载）。"""
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    start = _find_header_row(ws) + 1

    cur = {"course": "", "theme": "", "theme_vol": "", "course_vol": ""}
    rows: list[dict[str, Any]] = []
    for r in range(start, ws.max_row + 1):
        a, b = _txt(ws.cell(r, 1).value), _txt(ws.cell(r, 2).value)
        c, d = _txt(ws.cell(r, 3).value), _txt(ws.cell(r, 4).value)
        f = _txt(ws.cell(r, 6).value)
        h, j = _txt(ws.cell(r, 8).value), _txt(ws.cell(r, 10).value)
        if a:
            cur["course"] = a
        if b:
            cur["theme"] = b
        if h:
            cur["theme_vol"] = h
        if j:
            cur["course_vol"] = j
        # 合计行/题型汇总行：A 含“合计”或无考点名 → 跳过
        if not c and not d:
            if a and ("合计" in a or "题型" in a):
                break
            continue
        rows.append({
            "course": cur["course"], "theme": cur["theme"],
            "point_name": c, "knowledge": d,
            "kp_vol": f, "paper_no": parse_volume_no(f),
            "theme_vol": cur["theme_vol"], "theme_vol_no": parse_volume_no(cur["theme_vol"]),
            "course_vol": cur["course_vol"], "course_vol_range": parse_volume_range(cur["course_vol"]),
        })
    return rows


# ---------------- 卷号编排 D3 ----------------
def arrange_volume_numbers(rows: list[dict[str, Any]], *, comprehensive_per_course: int = 3) -> dict[str, Any]:
    """全局连续编排卷号：考点卷 F（逐行+1）→ 专题卷 H（每专题+1）→ 综合卷 J（每课程连续区间）。

    就地写回 rows 的 paper_no/kp_vol/theme_vol/course_vol，返回编排摘要。
    保持课程/专题在 rows 中的出现顺序。
    """
    n = 0
    # 1) 考点训练卷
    for r in rows:
        n += 1
        r["paper_no"] = n
        r["kp_vol"] = f"第{n}卷"
    # 2) 专题训练卷（按专题出现顺序，每专题一卷）
    theme_vol: dict[tuple, int] = {}
    for r in rows:
        key = (r["course"], r["theme"])
        if key not in theme_vol:
            n += 1
            theme_vol[key] = n
        r["theme_vol_no"] = theme_vol[key]
        r["theme_vol"] = f"第{theme_vol[key]}卷"
    # 3) 课程综合卷（每课程 comprehensive_per_course 连续卷）
    course_range: dict[str, tuple[int, int]] = {}
    for r in rows:
        c = r["course"]
        if c not in course_range:
            start = n + 1
            n += comprehensive_per_course
            course_range[c] = (start, n)
        lo, hi = course_range[c]
        r["course_vol_range"] = (lo, hi)
        r["course_vol"] = f"第{lo}-{hi}卷" if hi > lo else f"第{lo}卷"
    return {"total_volumes": n, "kpoint_count": len(rows),
            "theme_count": len(theme_vol), "course_count": len(course_range)}


# ---------------- 渲染 ----------------
def _cell(ws, r, c, value, *, size=10, bold=False, align="left", wrap=True):
    cell = ws.cell(r, c, value)
    cell.font = Font(name=_FONT, size=size, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return cell


def render_10col(rows: list[dict[str, Any]], *, title: str, subtitle: str,
                 out_path: str | Path) -> Path:
    """渲染 10 列规划总表（标题/副标题/表头 + A/B/G/H/I/J 合并）。

    首部不含"考试科目表名称"行与"每卷考试信息"行（用户 2026-07-02 决定移除）。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "考点规划总表"
    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _cell(ws, 1, 1, title, size=14, bold=True, align="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    _cell(ws, 2, 1, subtitle, size=12, bold=True, align="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    # 第 3 行留空分隔，表头从第 4 行起
    for c, head in enumerate(HEADERS, 1):
        _cell(ws, 4, c, head, size=10, bold=True, align="center")

    r = 5
    course_span: dict[str, list[int]] = {}
    theme_span: dict[tuple, list[int]] = {}
    for row in rows:
        _cell(ws, r, 1, row.get("course", ""))
        _cell(ws, r, 2, row.get("theme", ""))
        _cell(ws, r, 3, row.get("point_name", ""))
        _cell(ws, r, 4, row.get("knowledge", ""))
        _cell(ws, r, 5, 1, align="center")
        _cell(ws, r, 6, row.get("kp_vol", ""), align="center")
        _cell(ws, r, 7, 1, align="center")
        _cell(ws, r, 8, row.get("theme_vol", ""), align="center")
        crange = row.get("course_vol_range")
        _cell(ws, r, 9, (crange[1] - crange[0] + 1) if crange else 1, align="center")
        _cell(ws, r, 10, row.get("course_vol", ""), align="center")
        course_span.setdefault(row.get("course", ""), []).append(r)
        theme_span.setdefault((row.get("course", ""), row.get("theme", "")), []).append(r)
        r += 1

    # 合并 A(课程)/I/J（同课程）、B/G/H（同专题）
    def _merge_col(rowspan: list[int], col: int):
        if len(rowspan) > 1:
            ws.merge_cells(start_row=rowspan[0], start_column=col, end_row=rowspan[-1], end_column=col)

    for rs in course_span.values():
        for col in (1, 9, 10):
            _merge_col(rs, col)
    for rs in theme_span.values():
        for col in (2, 7, 8):
            _merge_col(rs, col)

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out
