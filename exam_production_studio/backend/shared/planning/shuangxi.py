"""考点双析卷规划表：扁平解析与渲染（含卷别，一主题展开教师/学生两行）。

双析卷模型：**一个主题 = 两份卷**（教师讲解卷 + 学生练习卷，同一套题奇偶拆，见 steps/split.py）。
- 内部生成单元是"主题 seq=k"（装配时 教师卷号=2k-1、学生卷号=2k）；
- 规划表 xlsx **展开为 2N 行**：序号 1..2N，成对的教师/学生行内容相同、仅"卷别"不同。

列（10）：课程 | 序号 | 试卷主题 | 卷别 | 考纲知识点 | 级别 | 题型 | 难度 | 套数 | 考纲标号
顶部标题/配置 2 行合并居中，数据区无跨列合并。
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ["课程", "序号", "试卷主题", "卷别", "考纲知识点", "级别", "题型", "难度", "套数", "考纲标号"]
_COL_WIDTHS = [18, 6, 20, 14, 52, 8, 24, 10, 6, 14]
_NCOL = len(HEADERS)
LEVELS = ("极重要", "重要", "标准")
ROLE_TEACHER = "教师讲解卷"
ROLE_STUDENT = "学生练习卷"
ROLES = (ROLE_TEACHER, ROLE_STUDENT)

_FONT = "微软雅黑"
_FILL_HEADER = PatternFill("solid", fgColor="4472C4")
_FILL_CRIT = PatternFill("solid", fgColor="FFD7D7")
_FILL_STD = PatternFill("solid", fgColor="FFF2CC")
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIAS = {
    "课程": "course", "知识模块": "course",
    "序号": "seq_disp",
    "试卷主题": "topic", "主题": "topic",
    "卷别": "role", "卷型": "role",
    "考纲知识点": "point_name", "知识点": "point_name",
    "级别": "level", "题型": "qtype", "难度": "difficulty_str",
    "套数": "sets", "考纲标号": "syllabus_no",
}


def _as_int(v: Any) -> int | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str) and v.strip().isdigit():
        return int(v.strip())
    return None


def _txt(v: Any) -> str:
    return str(v).strip() if v is not None else ""


# ---------------- 解析（2N 行 → 收回 N 主题）----------------
def parse_9col(path: str | Path) -> list[dict[str, Any]]:
    """解析双析卷规划表，返回**主题行**（N 个，paper_no=1..N）。

    规划表展开为教师/学生 2N 行；收回时以"教师讲解卷"行（或无卷别列时的每一行）为主题基准，
    学生行视为同主题的另一份，忽略（内容一致）。
    """
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    header_row, cols = 0, {}
    for r in range(1, min(ws.max_row, 12) + 1):
        names = [_txt(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 12) + 1)]
        m = {_ALIAS[n]: i + 1 for i, n in enumerate(names) if n in _ALIAS}
        if "topic" in m and "point_name" in m:
            header_row, cols = r, m
            break
    if not header_row:
        return []
    has_role = "role" in cols
    topics: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        def col(f):
            idx = cols.get(f)
            return ws.cell(r, idx).value if idx else None
        topic, point = _txt(col("topic")), _txt(col("point_name"))
        if not topic and not point:
            continue
        role = _txt(col("role"))
        # 有卷别列：只取教师讲解卷行作为主题基准（学生行内容相同，跳过）
        if has_role and role and role != ROLE_TEACHER:
            continue
        topics.append({
            "course": _txt(col("course")),
            "paper_no": len(topics) + 1,
            "topic": topic, "point_name": point,
            "level": _txt(col("level")) or "标准",
            "qtype": _txt(col("qtype")), "difficulty_str": _txt(col("difficulty_str")),
            "sets": _as_int(col("sets")) or 1, "syllabus_no": _txt(col("syllabus_no")),
        })
    return topics


def expand_to_volumes(topic_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """把 N 个主题行展开为 2N 份卷行（序号 1..2N，交替 教师讲解卷/学生练习卷）。"""
    out: list[dict[str, Any]] = []
    for k, row in enumerate(topic_rows, 1):
        for j, role in enumerate(ROLES):
            e = dict(row)
            e["seq_disp"] = 2 * (k - 1) + j + 1  # 教师=2k-1、学生=2k
            e["role"] = role
            out.append(e)
    return out


# ---------------- 渲染 ----------------
def _set(cell, value, *, size=9, bold=False, color=None, fill=None, align="left",
         wrap=False, border=False):
    cell.value = value
    cell.font = Font(name=_FONT, size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill
    if border:
        cell.border = _BORDER


def render_9col(topic_rows: list[dict[str, Any]], *, title: str, config_line: str,
                out_path: str | Path) -> Path:
    """渲染双析卷规划表：把 N 个主题行**展开为 2N 份卷行**（含卷别列）。

    顶部标题/配置 2 行合并居中；数据区每行一份卷、自带全列、无跨列合并。
    """
    rows = expand_to_volumes(topic_rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "考点双析卷规划表"
    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last = get_column_letter(_NCOL)

    _set(ws.cell(1, 1), title, size=12, bold=True, align="center")
    ws.merge_cells(f"A1:{last}1")
    _set(ws.cell(2, 1), config_line, size=9, align="center")
    ws.merge_cells(f"A2:{last}2")
    for c, head in enumerate(HEADERS, 1):
        _set(ws.cell(4, c), head, size=10, bold=True, color="FFFFFF", fill=_FILL_HEADER,
             align="center", border=True)

    r = 5
    for row in rows:
        fill = _FILL_CRIT if row.get("level") == "极重要" else _FILL_STD
        vals = [row.get("course", ""), row.get("seq_disp"), row.get("topic", ""),
                row.get("role", ""), row.get("point_name", ""), row.get("level", ""),
                row.get("qtype", ""), row.get("difficulty_str", ""),
                row.get("sets", 1), row.get("syllabus_no", "")]
        for c, v in enumerate(vals, 1):
            align = "center" if c in (2, 4, 6, 9) else "left"
            _set(ws.cell(r, c), v, size=9, align=align, wrap=(c == 5), fill=fill, border=True)
        r += 1

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out
