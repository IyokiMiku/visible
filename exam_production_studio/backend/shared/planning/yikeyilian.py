"""一课一练规划表：解析（C1，双格式）、带样式渲染（C3，扁平列式）、三行标题（C4）。

**扁平列式（当前标准，11 列 A–K）**——每行自带全列、数据区无跨列合并，便于变量读取：
  课程 | 单元 | 章 | 序号 | 考纲知识点 | 试卷主题 | 级别 | 题型 | 难度 | 套数 | 考纲标号
仅顶部标题/配置/教材 3 行做 A:K 合并居中（给人看）。

解析兼容两种来源：
- 新扁平表（表头含“课程/单元/章”）：按列直读；
- 旧合并表（一级行合并 A:H、二级行在 B 列、考点行 A=序号）：状态机解析。
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from shared.planning.numbers import number_to_cn, strip_leading_order

# 扁平列式表头（当前标准）
HEADERS = ["课程", "单元", "章", "序号", "考纲知识点", "试卷主题", "级别", "题型", "难度", "套数", "考纲标号"]
_COL_WIDTHS = [16, 22, 18, 6, 56, 18, 8, 26, 10, 6, 16]
_NCOL = len(HEADERS)
LEVELS = ("极重要", "重要", "标准")

_FONT = "微软雅黑"
_FILL_HEADER = PatternFill("solid", fgColor="4472C4")
_FILL_CRIT = PatternFill("solid", fgColor="FFD7D7")   # 极重要
_FILL_STD = PatternFill("solid", fgColor="FFF2CC")     # 重要/标准
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 表头名 → 规范字段（解析用）
_HEADER_ALIAS = {
    "课程": "course", "知识模块": "course",
    "单元": "unit_name", "一级标题": "unit_name",
    "章": "chapter_name", "二级标题": "chapter_name",
    "序号": "paper_no",
    "考纲知识点": "point_name", "知识点": "point_name",
    "试卷主题": "topic", "主题": "topic",
    "级别": "level",
    "题型": "qtype",
    "难度": "difficulty_str",
    "套数": "sets",
    "考纲标号": "syllabus_no",
}


def _as_int(v: Any) -> int | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str) and v.strip().isdigit():
        return int(v.strip())
    return None


def _txt(v: Any) -> str:
    return str(v).strip() if v is not None else ""


# ---------------- 层级号 / 序号推导 ----------------
def compute_hier_numbers(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """由 unit_name/chapter_name 变化推导 unit_no/chapter_no/section_no（**按课程分组重置**）。

    不改动 paper_no（扁平表序号显式给出；如需按课程重排序号见 renumber_by_course）。
    """
    last_course = last_unit = last_chapter = None
    unit_no = chapter_no = section_no = 0
    for r in rows:
        course = r.get("course") or ""
        if course != last_course:  # 进入新课程：层级号全部重置
            last_course = course
            unit_no = chapter_no = section_no = 0
            last_unit = last_chapter = None
        u = r.get("unit_name") or ""
        c = r.get("chapter_name") or ""
        if u != last_unit:
            unit_no += 1
            last_unit = u
            chapter_no = 0
            last_chapter = None
        if c != last_chapter:
            chapter_no += 1
            last_chapter = c
            section_no = 0
        section_no += 1
        r["unit_no"] = unit_no
        r["chapter_no"] = chapter_no or 1
        r["section_no"] = section_no
    return rows


def renumber_by_course(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """按课程重排 A 列序号：**进入新课程时序号重新从 1 开始**（命名"第X练"随之从头）。

    同时按课程重置层级号。极重要占两个连续号由数据本身的两行体现。
    """
    compute_hier_numbers(rows)
    last_course = None
    seq = 0
    for r in rows:
        course = r.get("course") or ""
        if course != last_course:
            last_course = course
            seq = 0
        seq += 1
        r["paper_no"] = seq
    return rows


# ---------------- 解析 C1（双格式）----------------
def _detect_header(ws) -> tuple[int, dict[str, int] | None]:
    """返回 (表头行号, 扁平字段→列号)。非扁平（旧合并格式）时第二项为 None。"""
    for r in range(1, min(ws.max_row, 15) + 1):
        names = [_txt(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 15) + 1)]
        mapping = {_HEADER_ALIAS[n]: i + 1 for i, n in enumerate(names) if n in _HEADER_ALIAS}
        # 扁平表：含“课程”列，且含序号/知识点/主题
        if "course" in mapping and "paper_no" in mapping and "point_name" in mapping:
            return r, mapping
        # 旧格式：首列“序号”、次列“考纲知识点”
        if names[:2] == ["序号", "考纲知识点"] or (names and names[0] == "序号"):
            return r, None
    return 5, None


def _parse_flat(ws, header_row: int, cols: dict[str, int]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        def col(field: str) -> Any:
            idx = cols.get(field)
            return ws.cell(r, idx).value if idx else None

        pno = _as_int(col("paper_no"))
        point = _txt(col("point_name"))
        topic = _txt(col("topic"))
        if pno is None and not point and not topic:
            continue  # 跳过空行
        rows.append({
            "course": _txt(col("course")),
            "unit_name": _txt(col("unit_name")),
            "chapter_name": _txt(col("chapter_name")),
            "paper_no": pno if pno is not None else len(rows) + 1,
            "point_name": point,
            "topic": topic,
            "level": _txt(col("level")) or "标准",
            "qtype": _txt(col("qtype")),
            "difficulty_str": _txt(col("difficulty_str")),
            "sets": _as_int(col("sets")) or 1,
            "syllabus_no": _txt(col("syllabus_no")),
        })
    compute_hier_numbers(rows)
    return rows


def _parse_old_merged(ws, header_row: int) -> list[dict[str, Any]]:
    """旧合并格式：一级行(A列文本,合并)/二级行(B列文本)/考点行(A=序号)。"""
    rows: list[dict[str, Any]] = []
    unit_name = chapter_name = ""
    unit_no = chapter_no = section_no = 0
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b, c, d = _txt(ws.cell(r, 2).value), _txt(ws.cell(r, 3).value), _txt(ws.cell(r, 4).value)
        e, f = _txt(ws.cell(r, 5).value), _txt(ws.cell(r, 6).value)
        g, h = ws.cell(r, 7).value, _txt(ws.cell(r, 8).value)
        a_int, a_txt = _as_int(a), _txt(a)
        if a_int is not None:
            section_no += 1
            if chapter_no == 0:
                chapter_no = 1
            rows.append({
                "course": "", "paper_no": a_int, "point_name": b, "topic": c, "level": d or "标准",
                "qtype": e, "difficulty_str": f, "sets": _as_int(g) or 1, "syllabus_no": h,
                "unit_name": unit_name, "unit_no": unit_no,
                "chapter_name": chapter_name, "chapter_no": chapter_no, "section_no": section_no,
            })
        elif a_txt:
            unit_no += 1
            unit_name = a_txt
            chapter_name, chapter_no, section_no = "", 0, 0
        elif b:
            chapter_no += 1
            chapter_name = b
            section_no = 0
    return rows


def parse_8col(path: str | Path) -> list[dict[str, Any]]:
    """解析一课一练规划表为考点行列表（自动识别扁平新格式 / 旧合并格式）。"""
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    header_row, cols = _detect_header(ws)
    if cols is not None:
        return _parse_flat(ws, header_row, cols)
    return _parse_old_merged(ws, header_row)


# ---------------- 渲染 C3（扁平；顶部 3 行合并）----------------
def _set(cell, value, *, size=9, bold=False, color=None, fill=None, align="left",
         wrap=False, border=False):
    cell.value = value
    cell.font = Font(name=_FONT, size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill
    if border:
        cell.border = _BORDER


def render_8col(rows: list[dict[str, Any]], *, title: str, config_line: str,
                textbook_line: str, out_path: str | Path) -> Path:
    """扁平 11 列渲染：顶部标题/配置/教材 3 行 A:K 合并居中，数据区每行自带全列、无合并。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "考点规划表"
    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last_col = get_column_letter(_NCOL)

    def merge_top(r):
        ws.merge_cells(f"A{r}:{last_col}{r}")

    _set(ws.cell(1, 1), title, size=12, bold=True, align="center")
    merge_top(1)
    _set(ws.cell(2, 1), config_line, size=9, align="center")
    merge_top(2)
    _set(ws.cell(3, 1), textbook_line, size=9, align="center")
    merge_top(3)
    for c, head in enumerate(HEADERS, 1):
        _set(ws.cell(5, c), head, size=10, bold=True, color="FFFFFF", fill=_FILL_HEADER,
             align="center", border=True)

    r = 6
    for row in rows:
        fill = _FILL_CRIT if row.get("level") == "极重要" else _FILL_STD
        vals = [
            row.get("course", ""), row.get("unit_name", ""), row.get("chapter_name", ""),
            row.get("paper_no"), row.get("point_name", ""), row.get("topic", ""),
            row.get("level", ""), row.get("qtype", ""), row.get("difficulty_str", ""),
            row.get("sets", 1), row.get("syllabus_no", ""),
        ]
        for c, v in enumerate(vals, 1):
            wrap = c == 5  # 考纲知识点自动换行
            align = "center" if c in (4, 7, 10) else "left"
            _set(ws.cell(r, c), v, size=9, align=align, wrap=wrap, fill=fill, border=True)
        r += 1

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out


# ---------------- 三行标题 C4 ----------------
def build_title_lines(ctx, paper_no: int, row: dict[str, Any] | None = None) -> list[str]:
    """一课一练正文三行标题（单元/章/节汉字、第y练阿拉伯）；含无 unit 的两级降级。"""
    row = row or {}
    province = (getattr(ctx, "province", "") or "").strip()
    exam = (getattr(ctx, "exam_type_name", "") or "").strip()
    textbook = (getattr(ctx, "textbook", "") or "").strip()
    edition = (getattr(ctx, "edition", "") or "").strip()

    line1 = f"{province}（{exam}）一课一练" if exam else f"{province}一课一练"
    edition = edition.strip("（）() ")
    edtxt = f"（{edition}）" if edition else ""
    line2 = f"《{textbook}》{edtxt} 第{paper_no}练".strip()

    topic = (row.get("topic") or "").strip()
    unit_name = strip_leading_order(row.get("unit_name") or "")
    chapter_name = strip_leading_order(row.get("chapter_name") or "")
    unit_no = int(row.get("unit_no") or 0)
    chapter_no = int(row.get("chapter_no") or 0) or 1
    section_no = int(row.get("section_no") or 0) or 1

    parts: list[str] = []
    if unit_name:
        parts.append(f"第{number_to_cn(unit_no or 1)}单元 {unit_name}")
    if chapter_name:
        parts.append(f"第{number_to_cn(chapter_no)}章 {chapter_name}")
    elif unit_name:
        parts.append(f"第{number_to_cn(chapter_no)}章")
    if topic:
        parts.append(f"第{number_to_cn(section_no)}节 {topic}")
    line3 = " ".join(parts).strip() or topic
    return [line1, line2, line3]
