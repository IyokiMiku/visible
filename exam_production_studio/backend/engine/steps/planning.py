"""规划表步骤（阶段五 steps/planning，阶段 C/D 增强）。

按 ctx.paper_type 分流：
- yikeyilian：解析上传的 8 列规划表；无上传则由结构化教材目录驱动合成（取代占位）。
- kaogang_100 / 其它：沿用通用逻辑（Phase D 再细化 kaogang 的 10 列）。

产出 生产规划/{规划表}.xlsx，并把卷落库 papers。
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from engine import registry, repo
from shared.ocr import load_structured_tocs
from shared.planning import kaogang as kg
from shared.planning import shuangxi as sx
from shared.planning import yikeyilian as yy

_PLAN_HEADERS = ["序号", "卷号", "试卷主题", "考纲知识点", "卷型", "难度", "套数"]


def _find_uploaded_plan(ctx) -> Path | None:
    in_dir = ctx.input_dir()
    if not in_dir.exists():
        return None
    xlsxs = [p for p in in_dir.rglob("*.xlsx") if "映射" not in p.name]
    return xlsxs[0] if xlsxs else None


# ================= 规划表产物缓存（供创建向导「切换来源」复用，不重跑 OCR）=================
def _ocr_cache_xlsx(ctx) -> Path:
    """本地生成(OCR)结果的独立 xlsx 备份。

    最终产物（如 {课程}_规划表.xlsx）是「本地生成 / 上传」两来源共用、会被覆盖的，
    不能当缓存；这里单独存一份，切回本地生成时用于展示产物名，不受上传源覆盖影响。
    """
    return ctx.dir("生产规划") / f".本地生成缓存_{ctx.course or '课程'}.xlsx"


def _plan_cache_json(ctx, source: str) -> Path:
    """按来源缓存 papers 的 JSON（卷量 = len(papers)，对所有卷类都准确）。

    考纲百套卷的卷量是「考点+专题+综合」三类聚合，无法靠重新解析渲染表得到，
    故这里缓存生成时真实产出的 papers 列表，切换来源时据此秒还原卷量与落库。
    """
    tag = "上传" if source == "upload" else "本地生成"
    return ctx.dir("生产规划") / f".规划缓存_{tag}_{ctx.course or '课程'}.json"


def _write_plan_cache(ctx, source: str, plan_path: Path, rows: list[dict[str, Any]]) -> None:
    """生成成功后写入该来源的 papers 缓存；失败静默（下次照常重跑，不影响主流程）。"""
    try:
        data = {"plan_file": Path(plan_path).name, "papers": rows}
        _plan_cache_json(ctx, source).write_text(
            json.dumps(data, ensure_ascii=False), encoding="utf-8")
    except Exception:  # noqa: BLE001
        pass


def plan_status(ctx, source: str) -> tuple[bool, int]:
    """只读探测某来源是否已有可复用产物及其卷量（不生成、不落库）。

    供创建向导「切换规划表来源」时判断该来源能否直接复用：
    - 有 papers 缓存 → 直接返回其卷量；
    - upload 来源即使无缓存，只要已上传 xlsx 也视为「存在」（卷量待生成时确定，返回 0）。
    """
    cache = _plan_cache_json(ctx, source)
    if cache.exists():
        try:
            data = json.loads(cache.read_text(encoding="utf-8"))
            return True, len(data.get("papers") or [])
        except Exception:  # noqa: BLE001
            pass
    if source == "upload" and _find_uploaded_plan(ctx) is not None:
        return True, 0
    return False, 0


def _subtype(ctx) -> str:
    return {"yikeyilian": "一课一练", "shuangxi": "考点训练卷"}.get(ctx.paper_type, "考点训练卷")


def _difficulty_str(diff: dict[str, Any]) -> str:
    e = diff.get("easy", 80)
    m = diff.get("medium", 10)
    h = diff.get("hard", 10)
    return f"{e}:{m}:{h}"


# ================= 一课一练（8 列，目录驱动 / 解析上传）=================
def _synthesize_from_toc(ctx) -> list[dict[str, Any]]:
    """由结构化教材目录合成 8 列考点行骨架（供 LLM/人工后续富化）。

    level1→单元，level2→章，更细节点→考点主题；无更细节点时二级本身作主题。
    知识点(B) 缺 LLM/考纲匹配时留占位「待补充」，级别默认「标准」。
    """
    tocs = load_structured_tocs(ctx)
    rows: list[dict[str, Any]] = []
    seq = 0
    for st in tocs:
        unit_name, unit_no = "", 0
        chapter_name, chapter_no, section_no = "", 0, 0
        nodes = st.nodes
        for i, n in enumerate(nodes):
            nxt_level = nodes[i + 1].level if i + 1 < len(nodes) else 0
            if n.level <= 1:
                unit_no += 1
                unit_name = n.title
                chapter_name, chapter_no, section_no = "", 0, 0
            elif n.level == 2:
                # 若下一节点更深（level3），则本节点为章；否则本节点直接作考点主题
                if nxt_level >= 3:
                    chapter_no += 1
                    chapter_name = n.title
                    section_no = 0
                else:
                    if chapter_no == 0:
                        chapter_no = 1
                    seq += 1
                    section_no += 1
                    rows.append(_toc_row(seq, n.title, unit_name, unit_no, chapter_name, chapter_no, section_no, st.textbook))
            else:  # level3+ 作考点主题
                if chapter_no == 0:
                    chapter_no = 1
                seq += 1
                section_no += 1
                rows.append(_toc_row(seq, n.title, unit_name, unit_no, chapter_name, chapter_no, section_no, st.textbook))
    return rows


def _toc_row(seq, topic, unit_name, unit_no, chapter_name, chapter_no, section_no, course="") -> dict[str, Any]:
    from shared.planning.numbers import strip_leading_order
    return {
        "paper_no": seq,
        "course": course,
        "topic": strip_leading_order(topic),
        "point_name": "待补充：教材目录主题“%s”未匹配考纲" % topic,
        "level": "标准",
        "qtype": "", "difficulty_str": "", "sets": 1, "syllabus_no": "待人工确认",
        "unit_name": unit_name, "unit_no": unit_no,
        "chapter_name": chapter_name, "chapter_no": chapter_no, "section_no": section_no,
    }


def _gen_yikeyilian(ctx, source: str, diff: dict[str, Any]) -> tuple[Path, list[dict[str, Any]]]:
    uploaded = _find_uploaded_plan(ctx) if source == "upload" else None
    if not uploaded:
        # 无论 source 如何，若输入目录里有规划表 xlsx，优先解析它
        uploaded = _find_uploaded_plan(ctx)

    if uploaded:
        base_rows = yy.parse_8col(uploaded)
    else:
        base_rows = _synthesize_from_toc(ctx)

    total = max((r["paper_no"] for r in base_rows), default=0)
    selected = set(ctx.selected_papers(total) or range(1, total + 1))
    rows = [r for r in base_rows if r["paper_no"] in selected] if selected else base_rows

    dstr = _difficulty_str(diff)
    course = ctx.course or ctx.textbook or "课程"
    for r in rows:  # 先补课程列（renumber 依赖 course 分组）
        if not r.get("course"):  # 旧合并样本无课程列 → 用项目课程/教材名
            r["course"] = course
    # 进入新课程时序号重新从 1 开始（命名"第X练"随之从头），并重置层级号
    yy.renumber_by_course(rows)
    for r in rows:
        r.setdefault("paper_subtype", "一课一练")
        r["original_paper_no"] = r["paper_no"]  # 课程内练号
        if not r.get("difficulty_str"):
            r["difficulty_str"] = dstr
        r["difficulty"] = diff
        r["status"] = "planned"

    out_dir = ctx.dir("生产规划")
    out_dir.mkdir(parents=True, exist_ok=True)
    title = f"{ctx.province}{ctx.exam_category}《一课一练》考点规划表 v1"
    config_line = f"题型：{ctx.volume_config.get('qtype_line', '见各行')} | 难度：{dstr}"
    textbook_line = f"参考教材：《{ctx.textbook}》{ctx.edition}"
    out_path = out_dir / f"{ctx.province}{ctx.exam_category}_{ctx.textbook or ctx.course}_一课一练考点规划表.xlsx"
    yy.render_8col(rows, title=title, config_line=config_line,
                   textbook_line=textbook_line, out_path=out_path)

    repo.replace_papers(ctx.project_id, rows)
    return out_path, rows


# ================= 通用（kaogang_100 / shuangxi / 兜底）=================
def _parse_uploaded_generic(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter, [])]
    out: list[dict[str, Any]] = []
    for i, row in enumerate(rows_iter, 1):
        cells = {headers[j] if j < len(headers) else f"c{j}": row[j] for j in range(len(row))}
        topic = str(cells.get("试卷主题") or cells.get("主题") or cells.get("试卷名称") or f"主题{i}").strip()
        point = str(cells.get("考纲知识点") or cells.get("知识点") or topic).strip()
        out.append({"paper_no": i, "topic": topic, "point_name": point})
    return out


def _synthesize_generic(ctx, total_hint: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for i in range(1, total_hint + 1):
        rows.append({
            "paper_no": i,
            "topic": f"{ctx.course or '课程'} 第{i}{'练' if ctx.paper_type == 'yikeyilian' else '卷'}主题",
            "point_name": f"{ctx.course or '课程'}知识点{i}",
        })
    return rows


def _gen_generic(ctx, source: str, diff: dict[str, Any]) -> tuple[Path, list[dict[str, Any]]]:
    uploaded = _find_uploaded_plan(ctx) if source == "upload" else None
    base_rows = _parse_uploaded_generic(uploaded) if uploaded else None

    total = len(base_rows) if base_rows else 3
    selected = ctx.selected_papers(total)
    if not selected:
        selected = list(range(1, total + 1))

    if base_rows:
        by_no = {r["paper_no"]: r for r in base_rows}
        rows = [by_no.get(n, {"paper_no": n, "topic": f"主题{n}", "point_name": ""}) for n in selected]
        for idx, r in enumerate(rows, 1):
            r["paper_no"] = idx
    else:
        synth = _synthesize_generic(ctx, max(selected))
        rows = [synth[n - 1] for n in selected]
        for idx, r in enumerate(rows, 1):
            r["paper_no"] = idx

    for r in rows:
        r.setdefault("paper_subtype", _subtype(ctx))
        r["difficulty"] = diff
        r["status"] = "planned"

    out_dir = ctx.dir("生产规划")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{ctx.course or '课程'}_规划表.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "规划表"
    ws.append(_PLAN_HEADERS)
    for r in rows:
        ws.append([r["paper_no"], r["paper_no"], r["topic"], r["point_name"],
                   r["paper_subtype"], _difficulty_str(diff), 1])
    wb.save(str(out_path))

    repo.replace_papers(ctx.project_id, rows)
    return out_path, rows


# ================= 考纲百套卷（10 列 A–J）=================
def _agg_texts(members: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """聚合卷题源自带其成员考点的文本（course/point_name/knowledge），
    使卷号范围筛选丢掉底层考点卷后，聚合卷仍能在映射阶段独立解析 kpointId。"""
    return [{"course": m.get("course", ""), "point_name": m.get("point_name", ""),
             "knowledge": m.get("knowledge", "")} for m in members]


def build_kaogang_papers(kp_rows: list[dict[str, Any]], diff: dict[str, Any]) -> list[dict[str, Any]]:
    """由已编排卷号的考点行（arrange_volume_numbers 后）构造三类卷 papers 行：
    考点训练卷（逐行）、专题训练卷（每专题聚合其考点）、课程综合卷（每课程区间共享该课程全部考点聚合）。

    卷号沿用全局编号（考点 1..N → 专题 → 综合），不重编号。
    """
    papers: list[dict[str, Any]] = []
    # 1) 考点训练卷（每行一卷）
    for r in kp_rows:
        papers.append({
            "paper_no": r.get("paper_no"),
            "topic": r.get("point_name"), "point_name": r.get("knowledge"),
            "paper_subtype": "考点训练卷", "difficulty": diff, "status": "planned",
            "module": r.get("course"), "theme": r.get("theme"),
            "original_paper_no": r.get("paper_no"),
            "meta": {
                "course": r.get("course"), "theme": r.get("theme"),
                "point_name": r.get("point_name"), "knowledge": r.get("knowledge"),
                "kp_vol": r.get("kp_vol"), "theme_vol_no": r.get("theme_vol_no"),
                "course_vol_range": r.get("course_vol_range"),
                "paper_subtype": "考点训练卷",
            },
        })

    # 2) 专题训练卷（每 (课程, 专题) 一卷；题源＝其下考点聚合）
    themes: dict[tuple, list[dict[str, Any]]] = {}
    for r in kp_rows:
        themes.setdefault((r.get("course"), r.get("theme"), r.get("theme_vol_no")), []).append(r)
    for (course, theme, tvol), members in themes.items():
        if not tvol:
            continue
        papers.append({
            "paper_no": tvol,
            "topic": theme, "point_name": f"{theme}（专题综合）",
            "paper_subtype": "专题训练卷", "difficulty": diff, "status": "planned",
            "module": course, "theme": theme,
            "original_paper_no": tvol,
            "meta": {
                "course": course, "theme": theme,
                "paper_subtype": "专题训练卷", "is_aggregate": True, "agg_kind": "theme",
                "theme_vol_no": tvol,
                "agg_texts": _agg_texts(members),
                "source_paper_nos": [m.get("paper_no") for m in members],
            },
        })

    # 3) 课程综合卷（每课程 course_vol_range 连续多卷，均聚合该课程全部考点）
    courses: dict[str, dict[str, Any]] = {}
    for r in kp_rows:
        rng = r.get("course_vol_range")
        if not rng:
            continue
        c = courses.setdefault(r.get("course"), {"range": tuple(rng), "members": []})
        c["members"].append(r)
    for course, info in courses.items():
        lo, hi = info["range"]
        members = info["members"]
        span = hi - lo + 1
        agg = _agg_texts(members)
        source_nos = [m.get("paper_no") for m in members]
        for k, pno in enumerate(range(lo, hi + 1), 1):
            papers.append({
                "paper_no": pno,
                "topic": f"{course}综合", "point_name": f"{course}课程综合（{k}/{span}）",
                "paper_subtype": "课程综合卷", "difficulty": diff, "status": "planned",
                "module": course, "theme": "课程综合",
                "original_paper_no": pno,
                "meta": {
                    "course": course, "theme": "课程综合",
                    "paper_subtype": "课程综合卷", "is_aggregate": True, "agg_kind": "course",
                    "course_vol_range": [lo, hi],
                    "agg_texts": agg, "source_paper_nos": source_nos,
                },
            })

    papers.sort(key=lambda p: (p.get("paper_no") or 0))
    return papers


def _gen_kaogang(ctx, source: str, diff: dict[str, Any]) -> tuple[Path, list[dict[str, Any]]]:
    uploaded = _find_uploaded_plan(ctx)
    if not uploaded:
        # 无上传的 10 列总表：退回通用合成（真正的目录/考纲驱动内容由 CD 的 LLM 生成）
        return _gen_generic(ctx, source, diff)

    kp_rows = kg.parse_10col(uploaded)
    summary = kg.arrange_volume_numbers(kp_rows)
    total = summary.get("total_volumes") or len(kp_rows)

    # 卷数口径＝三类卷全局编号总数（考点＋专题＋综合），卷号范围按此全局编号筛选
    selected = set(ctx.selected_papers(total) or range(1, total + 1))
    papers = [p for p in build_kaogang_papers(kp_rows, diff) if (p.get("paper_no") or 0) in selected]
    # 规划总表仅渲染选中的考点卷行（与生产口径一致；若选区不含考点则回退全量）
    sel_kp = [r for r in kp_rows if (r.get("paper_no") or 0) in selected] or kp_rows

    out_dir = ctx.dir("生产规划")
    out_dir.mkdir(parents=True, exist_ok=True)
    try:
        series = registry.get(ctx.paper_type).display_name
    except Exception:
        series = "考纲百套卷"
    out_path = out_dir / f"{ctx.province}_{ctx.exam_category}_考点规划总表.xlsx"
    kg.render_10col(sel_kp, title=f"《{ctx.province}{ctx.exam_category}》系列卷考点规划总表（{series}）",
                    subtitle=f"《{ctx.province}{ctx.exam_category}》系列卷考点规划总表",
                    out_path=out_path)
    repo.replace_papers(ctx.project_id, papers)
    return out_path, papers


# ================= 考点双析卷（扁平 9 列，每行一练，装配时奇偶拆教师/学生）=================
def _gen_shuangxi(ctx, source: str, diff: dict[str, Any]) -> tuple[Path, list[dict[str, Any]]]:
    uploaded = _find_uploaded_plan(ctx)
    if uploaded:
        base_rows = sx.parse_9col(uploaded) or _parse_uploaded_generic(uploaded)
    else:
        # 无上传：由考纲/通用合成占位（真正内容由 CD 的 LLM 生成）
        base_rows = _synthesize_generic(ctx, 3)

    total = max((r.get("paper_no") or 0 for r in base_rows), default=0)
    selected = set(ctx.selected_papers(total) or range(1, total + 1))
    rows = [r for r in base_rows if (r.get("paper_no") or 0) in selected] if selected else base_rows

    dstr = _difficulty_str(diff)
    course = ctx.course or "课程"
    for idx, r in enumerate(rows, 1):
        r.setdefault("course", course)
        r.setdefault("topic", r.get("point_name") or f"内容{idx}")
        r.setdefault("point_name", r.get("topic"))
        r.setdefault("level", "标准")
        r["paper_no"] = idx  # 内容序号 seq（装配时教师=2seq-1、学生=2seq）
        r["original_paper_no"] = idx
        if not r.get("difficulty_str"):
            r["difficulty_str"] = dstr
        r["paper_subtype"] = "考点双析卷"
        r["difficulty"] = diff
        r["status"] = "planned"

    out_dir = ctx.dir("生产规划")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{ctx.province}_{ctx.exam_category}_{ctx.course or '课程'}_考点双析卷规划表.xlsx"
    sx.render_9col(rows, title=f"{ctx.province}{ctx.exam_category}《考点双析卷》考点规划表 v1",
                   config_line=f"题型：{ctx.volume_config.get('qtype_line', '见各行')} | 难度：{dstr}",
                   out_path=out_path)
    repo.replace_papers(ctx.project_id, rows)
    return out_path, rows


def gen_planning(ctx, source: str = "ocr", force: bool = False) -> tuple[Path, list[dict[str, Any]]]:
    mode = registry.get(ctx.paper_type)
    diff = (ctx.volume_config or mode.default_volume_config).get(
        "difficulty", {"easy": 80, "medium": 10, "hard": 10})

    # 本地生成：非强制且命中缓存 → 复用缓存的 papers，跳过慢速 OCR 合成。
    # force=True（「重新生成」按钮）时无视缓存重跑。上传来源始终按已传 xlsx 现算，不走此缓存。
    if source == "ocr" and not force:
        cj, cx = _plan_cache_json(ctx, "ocr"), _ocr_cache_xlsx(ctx)
        if cj.exists() and cx.exists():
            try:
                rows = json.loads(cj.read_text(encoding="utf-8")).get("papers") or []
                if rows:
                    repo.replace_papers(ctx.project_id, rows)
                    return cx, rows
            except Exception:  # noqa: BLE001
                pass  # 缓存损坏则照常重跑

    if ctx.paper_type == "yikeyilian":
        path, rows = _gen_yikeyilian(ctx, source, diff)
    elif ctx.paper_type == "kaogang_100":
        path, rows = _gen_kaogang(ctx, source, diff)
    elif ctx.paper_type == "shuangxi":
        path, rows = _gen_shuangxi(ctx, source, diff)
    else:
        path, rows = _gen_generic(ctx, source, diff)

    # 缓存本次产物：papers JSON（供切换来源复用卷量）+ 本地生成额外备份一份 xlsx。
    _write_plan_cache(ctx, source, path, rows)
    if source == "ocr":
        try:
            import shutil
            shutil.copyfile(path, _ocr_cache_xlsx(ctx))
        except Exception:  # noqa: BLE001
            pass

    # 导出规划表到 桌面/输出结果/生产规划/{产品名}/{省份}_{考类}/
    from engine import archive
    archive.export_planning_artifact(ctx, path)
    return path, rows
