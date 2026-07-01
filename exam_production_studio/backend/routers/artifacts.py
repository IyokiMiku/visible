"""产物归档与下载（阶段七，设计文档 §5.4）。"""
from __future__ import annotations

import os
import re
import subprocess
import sys
import zipfile
from pathlib import Path

from fastapi import APIRouter
from fastapi.responses import FileResponse

from engine import archive
from ._common import fail, load_ctx, ok

router = APIRouter(prefix="/api/projects", tags=["artifacts"])

_ILLEGAL = re.compile(r'[\\/:*?"<>|]')


def _safe_filename(name: str, fallback: str) -> str:
    cleaned = _ILLEGAL.sub("", name or "").strip()
    return cleaned or fallback


def _rel(ctx, p: Path) -> str:
    try:
        return str(p.relative_to(ctx.root))
    except ValueError:
        return str(p)


def _within(root: Path, target: Path) -> bool:
    """target 是否位于 root 目录树内（解析软链/.. 后判断，堵前缀绕过）。"""
    try:
        target.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


@router.get("/{project_id}/artifacts")
def list_artifacts(project_id: str):
    try:
        ctx = load_ctx(project_id)
    except KeyError:
        return fail("项目不存在", status=404)
    groups: dict[str, list] = {"成品": [], "质检报告": [], "其他": []}
    out = ctx.dir("生成结果")
    if out.exists():
        for f in sorted(out.rglob("*")):
            if f.is_file():
                item = {"name": f.name, "path": _rel(ctx, f), "size": f.stat().st_size}
                if f.suffix == ".docx":
                    groups["成品"].append(item)
                else:
                    groups["其他"].append(item)
    rep = ctx.dir("质检报告")
    if rep.exists():
        for f in sorted(rep.glob("*.md")):
            groups["质检报告"].append({"name": f.name, "path": _rel(ctx, f), "size": f.stat().st_size})
    return ok(groups)


@router.get("/{project_id}/artifacts/zip")
def zip_artifacts(project_id: str):
    try:
        ctx = load_ctx(project_id)
    except KeyError:
        return fail("项目不存在", status=404)
    out = ctx.dir("生成结果")
    out.mkdir(parents=True, exist_ok=True)
    zip_path = out / f"{_safe_filename(ctx.name, project_id)}_打包.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in out.rglob("*"):
            if f.is_file() and f != zip_path:
                zf.write(f, f.relative_to(out))
    return FileResponse(str(zip_path), filename=zip_path.name)


@router.post("/{project_id}/artifacts/open")
def open_output_folder(project_id: str):
    """在本机资源管理器/访达中打开该项目的归档输出目录（不存在则创建）。

    仅适用于后端与用户同机的本地运行场景。
    """
    try:
        ctx = load_ctx(project_id)
    except KeyError:
        return fail("项目不存在", status=404)
    dest = archive.dest_dir(ctx)
    try:
        dest.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        return fail(f"无法创建输出目录：{e}")
    try:
        if sys.platform.startswith("win"):
            os.startfile(str(dest))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(dest)])
        else:
            subprocess.Popen(["xdg-open", str(dest)])
    except Exception as e:  # noqa: BLE001
        return fail(f"打开文件夹失败：{e}")
    return ok({"path": str(dest)})


@router.get("/{project_id}/artifacts/download")
def download_artifact(project_id: str, path: str):
    try:
        ctx = load_ctx(project_id)
    except KeyError:
        return fail("项目不存在", status=404)
    target = (ctx.root / path).resolve()
    # 路径安全：必须位于项目树内（用 relative_to 判断，避免前缀绕过）
    if not _within(ctx.root, target) or not target.is_file():
        return fail("文件不存在或路径非法", status=404)
    return FileResponse(str(target), filename=target.name)


@router.get("/{project_id}/artifacts/tree")
def artifact_tree(project_id: str, base: str = "04_生成输出"):
    """列出项目树内某个基准目录下的全部文件/子目录（供中间文件浏览页）。

    默认列 04_生成输出。返回扁平列表，前端据 path 自行拼树 / 调 download 下载。
    """
    try:
        ctx = load_ctx(project_id)
    except KeyError:
        return fail("项目不存在", status=404)
    base_dir = (ctx.root / base).resolve()
    if not _within(ctx.root, base_dir):
        return fail("路径非法", status=400)
    entries: list[dict] = []
    if base_dir.exists():
        for p in sorted(base_dir.rglob("*"), key=lambda x: str(x).lower()):
            is_dir = p.is_dir()
            try:
                size = 0 if is_dir else p.stat().st_size
                mtime = p.stat().st_mtime
            except OSError:
                size, mtime = 0, 0
            entries.append({
                "name": p.name,
                "path": _rel(ctx, p),
                "is_dir": is_dir,
                "size": size,
                "mtime": mtime,
                "suffix": "" if is_dir else p.suffix.lower(),
            })
    return ok({"base": base, "entries": entries})


@router.get("/{project_id}/artifacts/preview-xlsx")
def preview_xlsx(project_id: str, path: str, max_rows: int = 500, max_cols: int = 50):
    """把项目树内某个 xlsx（规划表/映射表等）读成二维数组，供前端表格预览。"""
    try:
        ctx = load_ctx(project_id)
    except KeyError:
        return fail("项目不存在", status=404)
    target = (ctx.root / path).resolve()
    if not _within(ctx.root, target) or not target.is_file() or target.suffix.lower() != ".xlsx":
        return fail("文件不存在或不是 xlsx", status=404)
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(target), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return fail(f"无法读取 xlsx：{exc}")
    try:
        sheets: list[dict] = []
        for ws in wb.worksheets:
            rows: list[list] = []
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx >= max_rows:
                    break
                rows.append(["" if v is None else v for v in row[:max_cols]])
            sheets.append({"name": ws.title, "rows": rows})
    finally:
        wb.close()
    return ok({"path": path, "sheets": sheets})
