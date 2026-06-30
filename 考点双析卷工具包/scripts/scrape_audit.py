"""
学科网审核页面 - AI质检结果爬虫

功能：
  打开浏览器 → 手动登录 → 自动逐条点击"日志"→"预览质检结果"
  → 提取错误信息表格 → 保存为 Excel

操作流程：
  列表页 → 点击"日志"按钮 → 点击"预览质检结果"链接 → 读取质检表格 → 关闭弹窗 → 下一条

用法：
  1. 安装依赖: pip install selenium openpyxl
  2. 运行: python scrape_audit.py
  3. 在弹出的浏览器中手动登录
  4. 登录完成后回到终端按回车

依赖：
  pip install selenium openpyxl
"""

import time
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    ElementClickInterceptedException,
    StaleElementReferenceException,
)

try:
    from webdriver_manager.microsoft import EdgeChromiumDriverManager
    USE_WDM = True
except ImportError:
    USE_WDM = False
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ════════════════════ 配置 ════════════════════
TARGET_URL = "https://rbm.xkw.com/#/resource/search"
WAIT_TIMEOUT = 10
PAGE_LOAD_DELAY = 2
OUTPUT_FILE = "AI文档质检报告汇总.xlsx"
# ════════════════════════════════════════════════


def create_driver():
    """创建 Edge 浏览器实例"""
    options = webdriver.EdgeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    if USE_WDM:
        service = Service(EdgeChromiumDriverManager().install())
        driver = webdriver.Edge(service=service, options=options)
    else:
        driver = webdriver.Edge(options=options)

    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"},
    )
    return driver


def wait_for_login(driver):
    """等待用户手动登录"""
    print("正在打开目标网页...", flush=True)
    driver.get(TARGET_URL)
    print("\n" + "=" * 60, flush=True)
    print("  浏览器已打开，请手动登录学科网账号。", flush=True)
    print("  登录后确认已进入资料搜索页面，再回到此终端按回车。", flush=True)
    print("=" * 60, flush=True)
    print(flush=True)
    input("按回车键继续爬取 >>> ")
    print("继续执行...", flush=True)
    time.sleep(PAGE_LOAD_DELAY)


def safe_click(driver, element):
    """安全点击，处理被遮挡的情况"""
    try:
        element.click()
    except ElementClickInterceptedException:
        driver.execute_script("arguments[0].click();", element)


def close_all_dialogs(driver):
    """关闭所有打开的弹窗/对话框，并等待它们完全消失"""
    max_rounds = 5
    for _ in range(max_rounds):
        closed_any = False

        try:
            close_buttons = driver.find_elements(
                By.CSS_SELECTOR,
                ".el-dialog__headerbtn, .el-message-box__headerbtn, "
                ".el-drawer__close-btn, button.close, .modal-close"
            )
            for btn in reversed(close_buttons):
                try:
                    if btn.is_displayed():
                        safe_click(driver, btn)
                        time.sleep(0.5)
                        closed_any = True
                except Exception:
                    continue
        except Exception:
            pass

        try:
            close_btns = driver.find_elements(By.XPATH,
                "//button[contains(@class,'el-button') and .//span[text()='关闭']]"
            )
            for btn in close_btns:
                try:
                    if btn.is_displayed():
                        safe_click(driver, btn)
                        time.sleep(0.5)
                        closed_any = True
                except Exception:
                    continue
        except Exception:
            pass

        if not closed_any:
            break

    # 等待所有弹窗完全消失
    try:
        WebDriverWait(driver, 5).until(
            lambda d: all(
                "display: none" in (w.get_attribute("style") or "")
                or not w.is_displayed()
                for w in d.find_elements(By.CSS_SELECTOR, ".el-dialog__wrapper")
            )
        )
    except (TimeoutException, StaleElementReferenceException):
        pass
    time.sleep(0.5)


def get_total_records(driver):
    """获取总记录数"""
    try:
        pagination_text = driver.find_element(
            By.CSS_SELECTOR, ".el-pagination__total, .total-text"
        ).text
        match = re.search(r"(\d+)", pagination_text)
        if match:
            return int(match.group(1))
    except Exception:
        pass
    return 0


def get_table_rows(driver):
    """获取当前页的表格行"""
    selectors = [
        ".el-table__row",
        "table tbody tr",
        ".el-table__body tbody tr",
        "tbody tr",
        ".ant-table-row",
        ".list-item",
        "tr[class*='row']",
    ]

    for selector in selectors:
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, selector))
            )
            rows = driver.find_elements(By.CSS_SELECTOR, selector)
            if rows:
                print(f"    [调试] 使用选择器 '{selector}' 找到 {len(rows)} 行", flush=True)
                return rows
        except TimeoutException:
            continue

    # 如果所有选择器都失败，打印页面中可能的表格结构
    print("    [调试] 所有预设选择器均未匹配，正在分析页面结构...", flush=True)
    try:
        tables = driver.find_elements(By.CSS_SELECTOR, "table")
        print(f"    [调试] 页面中找到 {len(tables)} 个 <table> 元素", flush=True)
        for idx, t in enumerate(tables):
            trs = t.find_elements(By.CSS_SELECTOR, "tr")
            print(f"    [调试]   table[{idx}]: {len(trs)} 行, class='{t.get_attribute('class')}'", flush=True)

        divs = driver.find_elements(By.CSS_SELECTOR, "div[class*='table'], div[class*='list']")
        print(f"    [调试] 含 'table'/'list' 的 div: {len(divs)} 个", flush=True)
        for d in divs[:5]:
            print(f"    [调试]   class='{d.get_attribute('class')}'", flush=True)
    except Exception as e:
        print(f"    [调试] 分析失败: {e}", flush=True)

    return []


def click_log_button(driver, row):
    """点击某一行的"日志"按钮"""
    try:
        # 方式1: 查找行内含"日志"文字的按钮
        log_btn = row.find_element(
            By.XPATH, ".//button[contains(.,'日志')] | .//span[contains(.,'日志')]/parent::button"
        )
        safe_click(driver, log_btn)
        time.sleep(PAGE_LOAD_DELAY)
        return True
    except NoSuchElementException:
        pass

    try:
        # 方式2: 操作列的第三个按钮（预览/修改/日志）
        buttons = row.find_elements(By.CSS_SELECTOR, "td:last-child button, td:last-child .el-button")
        for btn in buttons:
            if "日志" in btn.text:
                safe_click(driver, btn)
                time.sleep(PAGE_LOAD_DELAY)
                return True
    except Exception:
        pass

    return False


def click_preview_qc_result(driver):
    """在日志弹窗中点击"预览质检结果"链接"""
    try:
        # 等待日志弹窗出现
        WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".el-dialog, .el-drawer, .modal"))
        )
        time.sleep(1)

        # 查找"预览质检结果"链接
        link = WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//a[contains(text(),'预览质检结果')] | "
                "//span[contains(text(),'预览质检结果')] | "
                "//a[contains(text(),'质检结果')]"
            ))
        )
        safe_click(driver, link)
        time.sleep(PAGE_LOAD_DELAY)
        return True

    except TimeoutException:
        print("    未找到'预览质检结果'链接（可能该条目无AI质检记录）")
        return False


def extract_qc_table(driver):
    """从质检结果弹窗中提取错误表格数据"""
    errors = []

    try:
        WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.presence_of_element_located((
                By.XPATH,
                "//*[contains(text(),'AI质检结果') or contains(text(),'AI文档质检')]"
            ))
        )
        time.sleep(1)
    except TimeoutException:
        pass

    # 尝试多种表格选择器
    table_selectors = [
        ".el-dialog:last-of-type .el-table__row",
        ".el-dialog--center .el-table__row",
        ".el-dialog:last-of-type table tbody tr",
        ".el-dialog table tbody tr",
        ".modal table tbody tr",
        "table tbody tr",
    ]

    rows = []
    for selector in table_selectors:
        try:
            found = driver.find_elements(By.CSS_SELECTOR, selector)
            if found:
                rows = found
                break
        except Exception:
            continue

    if not rows:
        try:
            rows = driver.find_elements(
                By.XPATH,
                "//table[.//th[contains(text(),'错误类型')] or "
                ".//td[contains(text(),'错误类型')]]/tbody/tr"
            )
        except Exception:
            pass

    if not rows:
        try:
            dialogs = driver.find_elements(By.CSS_SELECTOR, ".el-dialog__wrapper, .el-dialog")
            for dialog in dialogs:
                if dialog.is_displayed() and "质检" in dialog.text:
                    text = dialog.text
                    return parse_qc_text(text)
        except Exception:
            pass
        return errors

    for row in rows:
        try:
            cells = row.find_elements(By.CSS_SELECTOR, "td")
            if len(cells) >= 3:
                error = {
                    "错误类型": cells[0].text.strip(),
                    "具体位置": cells[1].text.strip(),
                    "具体描述": cells[2].text.strip(),
                }
                if any(error.values()):
                    errors.append(error)
            elif len(cells) == 2:
                errors.append({
                    "错误类型": "",
                    "具体位置": cells[0].text.strip(),
                    "具体描述": cells[1].text.strip(),
                })
        except StaleElementReferenceException:
            continue

    return errors


def parse_qc_text(text):
    """从纯文本中解析质检结果（兜底方案）"""
    errors = []
    lines = text.split("\n")

    skip_keywords = {"错误类型", "具体位置", "具体描述", "AI质检结果",
                     "AI文档质检", "关闭", "确定"}

    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not line or any(kw == line for kw in skip_keywords):
            i += 1
            continue

        # 尝试识别三列数据模式
        if i + 2 < len(lines):
            type_text = line
            pos_text = lines[i + 1].strip()
            desc_text = lines[i + 2].strip()

            if (type_text and pos_text and desc_text and
                    not any(kw in type_text for kw in skip_keywords)):
                errors.append({
                    "错误类型": type_text,
                    "具体位置": pos_text,
                    "具体描述": desc_text,
                })
                i += 3
                continue

        i += 1

    return errors


def get_resource_id(row):
    """获取资料ID"""
    try:
        cells = row.find_elements(By.CSS_SELECTOR, "td")
        if cells:
            id_text = cells[0].text.strip() if cells[0].text.strip() else cells[1].text.strip()
            match = re.search(r"\d+", id_text)
            if match:
                return match.group(0)
    except Exception:
        pass
    return ""


def get_resource_name(row):
    """获取资料名称 —— 从行内定位"基本属性"列的标题 <span> 元素"""
    try:
        # 标题位于 <span style="vertical-align: middle; ... color: black"> 内
        spans = row.find_elements(
            By.CSS_SELECTOR, 'span[style*="color: black"][style*="vertical-align"]'
        )
        for span in spans:
            txt = span.text.strip()
            if txt and len(txt) > 5:
                return txt

        # 备选：找行内文本最长的单元格，取第一行
        cells = row.find_elements(By.CSS_SELECTOR, "td")
        best_text = ""
        for cell in cells:
            try:
                full_text = cell.text.strip()
                if len(full_text) > len(best_text):
                    best_text = full_text
            except StaleElementReferenceException:
                continue
        if best_text:
            first_line = best_text.split("\n")[0].strip()
            return first_line if first_line else best_text[:100]
    except Exception:
        pass
    return ""


def go_to_next_page(driver):
    """翻到下一页"""
    try:
        next_btn = driver.find_element(
            By.CSS_SELECTOR, ".el-pagination .btn-next"
        )
        if "disabled" in next_btn.get_attribute("class"):
            return False
        safe_click(driver, next_btn)
        time.sleep(PAGE_LOAD_DELAY)
        return True
    except Exception:
        return False


def _extract_volume_number(name):
    """从试卷名称中提取卷号数字，用于排序（如"第62卷 ..." → 62）"""
    match = re.search(r"第(\d+)[卷章节套]", name)
    if match:
        return int(match.group(1))
    match = re.search(r"(\d+)", name)
    return int(match.group(1)) if match else float("inf")


def save_results(results, filename):
    """保存结果到 Excel（按试卷序号排序，相邻试卷交替底色）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "质检报告"

    # 按试卷名称中的序号排序
    sorted_results = sorted(results, key=lambda r: _extract_volume_number(r.get("资料名称", "")))

    # 样式
    header_font = Font(name="宋体", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))
    body_font = Font(name="宋体", size=10)
    light_blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    headers = ["序号", "资料ID", "资料名称", "错误类型", "具体位置", "具体描述", "已修改"]
    for j, hd in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=hd)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # 按试卷名称分组交替底色
    color_index = 0
    prev_name = None

    for i, record in enumerate(sorted_results, 1):
        cur_name = record.get("资料名称", "")
        if cur_name != prev_name:
            if prev_name is not None:
                color_index += 1
            prev_name = cur_name
        row_fill = light_blue_fill if color_index % 2 == 0 else white_fill

        row_data = [
            (1, i, center),
            (2, record.get("资料ID", ""), center),
            (3, cur_name, left_wrap),
            (4, record.get("错误类型", ""), center),
            (5, record.get("具体位置", ""), center),
            (6, record.get("具体描述", ""), left_wrap),
            (7, "", center),
        ]
        for col, value, align in row_data:
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.font = body_font
            cell.alignment = align
            cell.border = thin_border
            cell.fill = row_fill

    # "已修改"列添加下拉勾选
    if len(sorted_results) > 0:
        dv = DataValidation(type="list", formula1='"✔"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"G2:G{len(sorted_results) + 1}")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 55
    ws.column_dimensions["G"].width = 10

    ws.freeze_panes = "A2"
    wb.save(filename)


def main():
    import sys
    print("=" * 60, flush=True)
    print("  学科网 AI质检结果 爬虫工具", flush=True)
    print("=" * 60, flush=True)
    print(flush=True)
    print("流程：列表页 → 日志 → 预览质检结果 → 提取错误表格", flush=True)
    print(flush=True)

    print("正在启动浏览器...", flush=True)
    driver = create_driver()
    print("浏览器启动成功！", flush=True)
    all_results = []

    try:
        wait_for_login(driver)

        # 确保在资料搜索页
        if "resource/search" not in driver.current_url:
            print("正在导航到资料搜索页...")
            driver.get(TARGET_URL)
            time.sleep(PAGE_LOAD_DELAY + 2)

        total = get_total_records(driver)
        print(f"\n页面显示共 {total} 条记录" if total else "\n已就绪")

        while True:
            count_input = input("\n请输入本次要爬取的条数（输入 0 或 q 结束）：").strip()
            if count_input.lower() == "q" or count_input == "0":
                break

            if not count_input.isdigit() or int(count_input) <= 0:
                print("请输入有效的正整数。")
                continue

            target_count = int(count_input)
            print(f"\n将从当前页面开始爬取 {target_count} 条记录...")

            processed = 0
            page_num = 1
            # 记录已提取过的错误，用于去除累积沿用的旧错误
            seen_errors = set()

            while processed < target_count:
                print(f"\n{'─' * 40}")
                print(f"第 {page_num} 页 (已处理 {processed}/{target_count})")
                print(f"{'─' * 40}")

                rows = get_table_rows(driver)
                if not rows:
                    print("未找到表格行，爬取结束。")
                    break

                row_count = len(rows)
                print(f"当前页 {row_count} 条记录")

                for i in range(row_count):
                    if processed >= target_count:
                        break

                    rows = get_table_rows(driver)
                    if not rows or i >= len(rows):
                        break

                    row = rows[i]
                    resource_id = get_resource_id(row)
                    resource_name = get_resource_name(row)

                    print(f"\n  [{processed+1}/{target_count}] ID: {resource_id}")
                    print(f"    名称: {resource_name[:50]}")

                    # 步骤1: 点击"日志"
                    if not click_log_button(driver, row):
                        print("    ⚠ 未找到日志按钮，跳过")
                        continue

                    # 步骤2: 点击"预览质检结果"
                    if not click_preview_qc_result(driver):
                        close_all_dialogs(driver)
                        time.sleep(1)
                        processed += 1
                        continue

                    # 步骤3: 提取质检表格
                    raw_errors = extract_qc_table(driver)

                    # 步骤4: 提取完后立即关闭所有弹窗
                    close_all_dialogs(driver)
                    time.sleep(1)

                    # 去重：因弹窗累积，当前提取结果包含之前所有试卷的错误
                    # 用差集只保留本条试卷新增的错误
                    new_errors = []
                    for err in raw_errors:
                        key = (err["错误类型"], err["具体位置"], err["具体描述"])
                        if key not in seen_errors:
                            new_errors.append(err)
                    # 将本次所有原始错误（含旧的）加入 seen_errors，供下一条去重
                    for err in raw_errors:
                        seen_errors.add((err["错误类型"], err["具体位置"], err["具体描述"]))

                    if new_errors:
                        print(f"    找到 {len(new_errors)} 条质检报错 (原始 {len(raw_errors)} 条，去重 {len(raw_errors)-len(new_errors)} 条):")
                        for err in new_errors:
                            print(f"      - [{err['错误类型']}] {err['具体位置']}: {err['具体描述'][:40]}")
                            all_results.append({
                                "资料ID": resource_id,
                                "资料名称": resource_name,
                                "错误类型": err["错误类型"],
                                "具体位置": err["具体位置"],
                                "具体描述": err["具体描述"],
                            })
                    else:
                        print(f"    无质检报错或未能提取 (原始 {len(raw_errors)} 条均为重复)"
                              if raw_errors else "    无质检报错或未能提取")

                    processed += 1

                    # 每处理5条保存一次
                    if len(all_results) % 5 == 0 and all_results:
                        save_results(all_results, OUTPUT_FILE)

                # 如果还没处理够，翻页继续
                if processed < target_count:
                    if not go_to_next_page(driver):
                        print("\n已到最后一页，无更多记录。")
                        break
                    page_num += 1

            print(f"\n本轮爬取完成，已处理 {processed} 条。")
            if all_results:
                save_results(all_results, OUTPUT_FILE)
                print(f"当前共 {len(all_results)} 条报错已保存到 {OUTPUT_FILE}")

        # 最终保存
        if all_results:
            save_results(all_results, OUTPUT_FILE)
            print(f"\n{'=' * 60}")
            print(f"  爬取完成！共获取 {len(all_results)} 条质检报错")
            print(f"  已保存到: {OUTPUT_FILE}")
            print(f"{'=' * 60}")
        else:
            print("\n未获取到任何质检数据。")

    except KeyboardInterrupt:
        print("\n\n用户中断...")
        if all_results:
            save_results(all_results, OUTPUT_FILE)
            print(f"已保存 {len(all_results)} 条记录到 {OUTPUT_FILE}")
    except Exception as e:
        print(f"\n发生错误: {e}")
        import traceback
        traceback.print_exc()
        if all_results:
            save_results(all_results, OUTPUT_FILE)
            print(f"已保存 {len(all_results)} 条记录到 {OUTPUT_FILE}")
    finally:
        input("\n按回车关闭浏览器...")
        driver.quit()


if __name__ == "__main__":
    main()
