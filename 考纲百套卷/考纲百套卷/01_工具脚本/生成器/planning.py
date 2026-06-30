"""考纲百套卷规划表读取模块。

该模块把 ``04_生成输出/考点规划总表`` 下的 xlsx 规划表解析为统一的
试卷索引，供命令行选择指定卷号，并识别其属于考点训练卷、专题训练卷
还是课程综合卷。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
import re

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

try:
    from docx import Document
except ImportError:  # pragma: no cover
    Document = None


HEADER_ROW = 6
DATA_START_ROW = 7

COL_MODULE = 1
COL_TOPIC = 2
COL_POINT_NAME = 3
COL_POINT_CONTENT = 4
COL_POINT_COUNT = 5
COL_POINT_PAPER = 6
COL_TOPIC_COUNT = 7
COL_TOPIC_PAPER = 8
COL_COURSE_COUNT = 9
COL_COURSE_PAPER = 10

POINT_PAPER_TYPE = "考点训练卷"
TOPIC_PAPER_TYPE = "专题训练卷"
COURSE_PAPER_TYPE = "课程综合卷"

FLAT_PLAN_HEADERS = [
    "卷号",
    "省份",
    "考试类别",
    "课程",
    "模块",
    "专题",
    "考点",
    "题型",
    "题量",
    "难度",
    "题源关键词",
]

_FLAT_PLAN_FIELD_MAP = {
    "卷号": "paper_no",
    "省份": "province",
    "考试类别": "exam_category",
    "课程": "subject",
    "模块": "module",
    "专题": "topic",
    "考点": "knowledge_points",
    "题型": "question_type",
    "题量": "question_count",
    "难度": "difficulty",
    "题源关键词": "source_keywords",
}


@dataclass
class PlanningMeta:
    """规划表元数据。"""

    path: Path
    sheet_name: str
    title: str = ""
    subtitle: str = ""
    exam_info: str = ""
    exam_table_title: str = ""
    exam_type: str = ""
    province: str = ""
    exam_category: str = ""
    all_courses: list[str] = field(default_factory=list)
    question_type_summaries: dict[str, list[dict[str, str]]] = field(default_factory=dict)


@dataclass
class PlanRow:
    """规划表中一条 C 列考点行。"""

    row_no: int
    module: str = ""
    topic: str = ""
    point_name: str = ""
    point_content: str = ""
    point_count: int = 0
    point_paper_ref: str = ""
    topic_count: int = 0
    topic_paper_ref: str = ""
    course_count: int = 0
    course_paper_ref: str = ""
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass
class BlueprintRow:
    """细目表中一题对应的约束行。"""

    question_no: int | None = None
    question_no_raw: str = ""
    question_type: str = ""
    difficulty: str = ""
    content: str = ""
    point_name: str = ""
    knowledge_point: str = ""
    requirement: str = ""
    intent: str = ""
    paper_type: str = ""
    paper_label: str = ""
    province: str = ""
    module: str = ""
    topic: str = ""
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass
class PaperPlan:
    """按卷号索引后的单卷规划。"""

    paper_no: int
    paper_label: str
    paper_type: str
    module: str
    topic: str = ""
    point_name: str = ""
    point_content: str = ""
    rows: list[PlanRow] = field(default_factory=list)
    meta: PlanningMeta | None = None
    paper_ref: str = ""
    blueprint_path: Path | None = None
    blueprint_rows: list[BlueprintRow] = field(default_factory=list)
    blueprint_warnings: list[str] = field(default_factory=list)


@dataclass
class PlanItem:
    """兼容旧流程的一行规划任务。"""

    paper_no: str = ""
    paper_type: str = ""
    province: str = ""
    exam_category: str = ""
    subject: str = ""
    module: str = ""
    topic: str = ""
    knowledge_points: str = ""
    question_type: str = ""
    question_count: int = 0
    difficulty: str = ""
    source_keywords: str = ""
    raw: dict[str, Any] = field(default_factory=dict)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def _normalize_plan_header(text: Any) -> str:
    return re.sub(r"[\s\n\r　/／（）()\-—_]+", "", _clean_text(text))


_FLAT_PLAN_HEADER_ALIASES = {
    _normalize_plan_header(header): header for header in FLAT_PLAN_HEADERS
}
_FLAT_PLAN_HEADER_ALIASES.update(
    {
        "卷型": "卷号",
        "试卷编号": "卷号",
        "地区": "省份",
        "考类": "考试类别",
        "考试类型": "考试类别",
        "专业类": "考试类别",
        "科目": "课程",
        "教材": "课程",
        "学科": "课程",
        "一级主题": "模块",
        "单元": "模块",
        "知识模块": "模块",
        "主题": "专题",
        "二级主题": "专题",
        "章节": "专题",
        "知识点": "考点",
        "考纲知识点": "考点",
        "题目类型": "题型",
        "数量": "题量",
        "小题数": "题量",
        "难度要求": "难度",
        "难度比例": "难度",
        "关键词": "题源关键词",
        "API关键词": "题源关键词",
        "题库关键词": "题源关键词",
        "题源": "题源关键词",
    }
)


def _canonical_flat_header(value: Any) -> str | None:
    return _FLAT_PLAN_HEADER_ALIASES.get(_normalize_plan_header(value))


def _normalize_paper_no(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return format_paper_label(int(float(text)))
    numbers = parse_paper_numbers(text)
    if len(numbers) == 1:
        return format_paper_label(numbers[0])
    return text


def _to_int(value: Any, default: int = 0) -> int:
    text = _clean_text(value)
    if not text:
        return default
    try:
        return int(float(text))
    except Exception:
        match = re.search(r"\d+", text)
        return int(match.group(0)) if match else default


def _build_merged_value_map(ws) -> dict[tuple[int, int], Any]:
    merged_values: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = top_left_value
    return merged_values


def _cell_value(ws, merged_values: dict[tuple[int, int], Any], row: int, col: int) -> Any:
    value = ws.cell(row, col).value
    if value not in (None, ""):
        return value
    return merged_values.get((row, col), "")


def _parse_meta_from_filename(path: Path) -> tuple[str, str]:
    stem = path.stem
    suffix = "_考点规划总表"
    if stem.endswith(suffix):
        stem = stem[: -len(suffix)]
    parts = [part.strip() for part in stem.split("_") if part.strip()]
    if len(parts) >= 2:
        return parts[0], parts[1]
    return "", ""


def _parse_meta_from_title(title: str) -> tuple[str, str]:
    text = _clean_text(title).strip("《》")
    match = re.search(
        r"(内蒙古自治区|新疆维吾尔自治区|西藏自治区|广西壮族自治区|宁夏回族自治区|[一-龥]+(?:省|市|自治区)|内蒙古|新疆|西藏|广西|宁夏)(.+?类)",
        text,
    )
    if match:
        return match.group(1), match.group(2)
    return "", ""


def _extract_question_type_summaries(ws) -> dict[str, list[dict[str, str]]]:
    """读取规划表底部“各课程题型结构”汇总。"""
    summaries: dict[str, list[dict[str, str]]] = {}
    current_course = ""
    in_summary = False

    for row_no in range(1, ws.max_row + 1):
        row_values = [_clean_text(ws.cell(row_no, col).value) for col in range(1, min(ws.max_column, 6) + 1)]
        joined = " ".join(row_values)
        if "各课程题型结构" in joined:
            in_summary = True
            continue
        if not in_summary:
            continue

        first, second = (row_values + ["", ""])[:2]
        if second == "题型":
            current_course = re.sub(r"[（(].*?[)）]", "", first).strip()
            if current_course:
                summaries.setdefault(current_course, [])
            continue

        if current_course and second and second != "题型":
            summaries.setdefault(current_course, []).append(
                {
                    "type": second,
                    "count": row_values[2] if len(row_values) > 2 else "",
                    "score_per": row_values[3] if len(row_values) > 3 else "",
                    "subtotal": row_values[4] if len(row_values) > 4 else "",
                    "total": row_values[5] if len(row_values) > 5 else "",
                }
            )

    return summaries


def _load_meta(path: Path, ws) -> PlanningMeta:
    province, exam_category = _parse_meta_from_filename(path)
    title = _clean_text(ws.cell(1, 1).value)
    if not province or not exam_category:
        province_from_title, category_from_title = _parse_meta_from_title(title)
        province = province or province_from_title
        exam_category = exam_category or category_from_title

    exam_parts = [_clean_text(ws.cell(3, col).value) for col in range(1, COL_COURSE_PAPER + 1)]
    exam_info = " ".join(part for part in exam_parts if part)
    exam_table_title = _clean_text(ws.cell(2, 1).value)
    exam_type = _clean_text(ws.cell(2, 2).value)
    return PlanningMeta(
        path=path,
        sheet_name=ws.title,
        title=title,
        subtitle=_clean_text(ws.cell(5, 1).value),
        exam_info=exam_info,
        exam_table_title=exam_table_title,
        exam_type=exam_type,
        province=province,
        exam_category=exam_category,
        question_type_summaries=_extract_question_type_summaries(ws),
    )


def _detect_flat_plan_header(ws, max_scan_rows: int = 20) -> tuple[int, dict[str, int]] | None:
    max_row = min(ws.max_row, max_scan_rows)
    for row_no in range(1, max_row + 1):
        col_map: dict[str, int] = {}
        for col_no in range(1, ws.max_column + 1):
            canonical = _canonical_flat_header(ws.cell(row_no, col_no).value)
            if canonical and canonical not in col_map:
                col_map[canonical] = col_no
        if all(header in col_map for header in FLAT_PLAN_HEADERS):
            return row_no, col_map
    return None


def _looks_like_flat_plan_workbook(ws, max_scan_rows: int = 20) -> bool:
    max_row = min(ws.max_row, max_scan_rows)
    for row_no in range(1, max_row + 1):
        matched_headers = {
            canonical
            for col_no in range(1, ws.max_column + 1)
            if (canonical := _canonical_flat_header(ws.cell(row_no, col_no).value))
        }
        if "卷号" in matched_headers and len(matched_headers) >= 6:
            return True
    return False


def _flat_header_error(ws) -> str:
    seen: list[str] = []
    for row_no in range(1, min(ws.max_row, 20) + 1):
        row_values = [_clean_text(ws.cell(row_no, col_no).value) for col_no in range(1, ws.max_column + 1)]
        row_values = [value for value in row_values if value]
        if row_values:
            seen.append(f"第{row_no}行：" + "、".join(row_values))
    seen_text = "\n".join(seen[:5]) or "未识别到非空表头行"
    return (
        "规划表表头缺失，无法按固定表头读取。\n"
        "期望表头：" + "、".join(FLAT_PLAN_HEADERS) + "\n"
        "实际前几行：\n" + seen_text
    )


def load_flat_plan(plan_path: str | Path) -> list[PlanItem]:
    """读取最终固定 11 列生产规划表。"""
    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，请先安装规划表读取依赖。")

    path = Path(plan_path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        detected = _detect_flat_plan_header(ws)
        if detected is None:
            raise ValueError(_flat_header_error(ws))
        header_row, col_map = detected
        items: list[PlanItem] = []
        for row_no in range(header_row + 1, ws.max_row + 1):
            values = {
                header: _clean_text(ws.cell(row_no, col_map[header]).value)
                for header in FLAT_PLAN_HEADERS
            }
            if not any(values.values()):
                continue
            if not values["卷号"]:
                raise ValueError(f"规划表第{row_no}行缺少卷号。")
            raw = dict(values)
            raw["row_no"] = row_no
            items.append(
                PlanItem(
                    paper_no=_normalize_paper_no(values["卷号"]),
                    province=values["省份"],
                    exam_category=values["考试类别"],
                    subject=values["课程"],
                    module=values["模块"],
                    topic=values["专题"],
                    knowledge_points=values["考点"],
                    question_type=values["题型"],
                    question_count=_to_int(values["题量"]),
                    difficulty=values["难度"],
                    source_keywords=values["题源关键词"],
                    raw=raw,
                )
            )
        return items
    finally:
        wb.close()


def is_flat_plan_workbook(plan_path: str | Path) -> bool:
    """判断工作簿是否使用最终固定 11 列表头。"""
    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，请先安装规划表读取依赖。")
    path = Path(plan_path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return _detect_flat_plan_header(ws) is not None
    finally:
        wb.close()


def looks_like_flat_plan_workbook(plan_path: str | Path) -> bool:
    """判断工作簿是否像最终固定表，但可能缺少部分列。"""
    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，请先安装规划表读取依赖。")
    path = Path(plan_path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return _looks_like_flat_plan_workbook(ws)
    finally:
        wb.close()


def parse_paper_numbers(ref: str) -> list[int]:
    """解析“第1卷”“第34-36卷”“第34、35、36卷”等卷号文本。"""
    text = _clean_text(ref)
    if not text:
        return []
    translate = str.maketrans({
        "－": "-",
        "–": "-",
        "—": "-",
        "～": "-",
        "~": "-",
        "，": ",",
        "、": ",",
        "；": ",",
        ";": ",",
        " ": "",
        "　": "",
    })
    text = text.translate(translate).replace("第", "").replace("卷", "")
    text = re.sub(r"\s+", "", text)

    result: list[int] = []
    for part in filter(None, text.split(",")):
        range_match = re.fullmatch(r"(\d+)-(\d+)", part)
        if range_match:
            start, end = (int(range_match.group(1)), int(range_match.group(2)))
            if start > end:
                start, end = end, start
            result.extend(range(start, end + 1))
            continue
        if part.isdigit():
            result.append(int(part))

    deduped: list[int] = []
    seen: set[int] = set()
    for number in result:
        if number not in seen:
            deduped.append(number)
            seen.add(number)
    return deduped


def format_paper_label(paper_no: int) -> str:
    return f"第{paper_no}卷"


def _add_unique(index: dict[int, PaperPlan], paper: PaperPlan) -> None:
    existing = index.get(paper.paper_no)
    if existing is None:
        index[paper.paper_no] = paper
        return
    if existing.paper_type != paper.paper_type:
        raise ValueError(
            f"卷号冲突：{format_paper_label(paper.paper_no)} 同时属于 "
            f"{existing.paper_type} 和 {paper.paper_type}"
        )


def build_paper_index(rows: list[PlanRow], meta: PlanningMeta) -> dict[int, PaperPlan]:
    """按 F/H/J 三类卷号列构建卷号索引。"""
    index: dict[int, PaperPlan] = {}

    for row in rows:
        for paper_no in parse_paper_numbers(row.point_paper_ref):
            _add_unique(
                index,
                PaperPlan(
                    paper_no=paper_no,
                    paper_label=format_paper_label(paper_no),
                    paper_type=POINT_PAPER_TYPE,
                    module=row.module,
                    topic=row.topic,
                    point_name=row.point_name,
                    point_content=row.point_content,
                    rows=[row],
                    meta=meta,
                    paper_ref=row.point_paper_ref,
                ),
            )

    topic_groups: dict[tuple[str, str, str], list[PlanRow]] = {}
    for row in rows:
        if row.topic_paper_ref:
            topic_groups.setdefault((row.module, row.topic, row.topic_paper_ref), []).append(row)
    for (module, topic, paper_ref), group_rows in topic_groups.items():
        for paper_no in parse_paper_numbers(paper_ref):
            _add_unique(
                index,
                PaperPlan(
                    paper_no=paper_no,
                    paper_label=format_paper_label(paper_no),
                    paper_type=TOPIC_PAPER_TYPE,
                    module=module,
                    topic=topic,
                    rows=group_rows,
                    meta=meta,
                    paper_ref=paper_ref,
                ),
            )

    course_groups: dict[tuple[str, str], list[PlanRow]] = {}
    for row in rows:
        if row.course_paper_ref:
            course_groups.setdefault((row.module, row.course_paper_ref), []).append(row)
    for (module, paper_ref), group_rows in course_groups.items():
        for paper_no in parse_paper_numbers(paper_ref):
            _add_unique(
                index,
                PaperPlan(
                    paper_no=paper_no,
                    paper_label=format_paper_label(paper_no),
                    paper_type=COURSE_PAPER_TYPE,
                    module=module,
                    rows=group_rows,
                    meta=meta,
                    paper_ref=paper_ref,
                ),
            )

    return dict(sorted(index.items()))


def load_planning_workbook(plan_path: str | Path) -> tuple[PlanningMeta, list[PlanRow], dict[int, PaperPlan]]:
    """读取考点规划总表，返回元数据、考点行和按卷号索引的试卷规划。"""
    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，请先安装规划表读取依赖。")

    path = Path(plan_path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        ws = wb[wb.sheetnames[0]]
        merged_values = _build_merged_value_map(ws)
        meta = _load_meta(path, ws)
        rows: list[PlanRow] = []

        for row_no in range(DATA_START_ROW, ws.max_row + 1):
            module = _clean_text(_cell_value(ws, merged_values, row_no, COL_MODULE))
            if module.replace(" ", "") in {"合计", "合计:"}:
                break

            point_name = _clean_text(_cell_value(ws, merged_values, row_no, COL_POINT_NAME))
            point_paper_ref = _clean_text(_cell_value(ws, merged_values, row_no, COL_POINT_PAPER))
            if not point_name and not point_paper_ref:
                continue

            topic = _clean_text(_cell_value(ws, merged_values, row_no, COL_TOPIC))
            point_content = _clean_text(_cell_value(ws, merged_values, row_no, COL_POINT_CONTENT))
            point_count = _to_int(_cell_value(ws, merged_values, row_no, COL_POINT_COUNT))
            topic_count = _to_int(_cell_value(ws, merged_values, row_no, COL_TOPIC_COUNT))
            topic_paper_ref = _clean_text(_cell_value(ws, merged_values, row_no, COL_TOPIC_PAPER))
            course_count = _to_int(_cell_value(ws, merged_values, row_no, COL_COURSE_COUNT))
            course_paper_ref = _clean_text(_cell_value(ws, merged_values, row_no, COL_COURSE_PAPER))

            rows.append(
                PlanRow(
                    row_no=row_no,
                    module=module,
                    topic=topic,
                    point_name=point_name,
                    point_content=point_content,
                    point_count=point_count,
                    point_paper_ref=point_paper_ref,
                    topic_count=topic_count,
                    topic_paper_ref=topic_paper_ref,
                    course_count=course_count,
                    course_paper_ref=course_paper_ref,
                    raw={
                        "知识模块": module,
                        "考纲一级标题（专题名称）": topic,
                        "考点名称（考点训练卷名称）": point_name,
                        "掌握/理解/了解考点内容": point_content,
                        "考点训练卷数量": point_count,
                        "考点训练卷卷号": point_paper_ref,
                        "专题训练卷数量": topic_count,
                        "专题训练卷卷号": topic_paper_ref,
                        "课程综合卷数量": course_count,
                        "课程综合卷卷号": course_paper_ref,
                    },
                )
            )
    finally:
        wb.close()

    paper_index = build_paper_index(rows, meta)
    seen: set[str] = set()
    all_courses: list[str] = []
    for row in rows:
        if row.module and row.module not in seen:
            seen.add(row.module)
            all_courses.append(row.module)
    meta.all_courses = all_courses
    return meta, rows, paper_index


def validate_paper_index(paper_index: dict[int, PaperPlan]) -> list[str]:
    warnings: list[str] = []
    if not paper_index:
        warnings.append("未解析到任何卷号。")
        return warnings
    numbers = sorted(paper_index)
    missing = [number for number in range(numbers[0], numbers[-1] + 1) if number not in paper_index]
    if missing:
        warnings.append("缺少卷号：" + "、".join(format_paper_label(number) for number in missing))
    return warnings


def _safe_filename_part(text: str) -> str:
    cleaned = _clean_text(text)
    return re.sub(r'[<>:"/\\|?*]+', "_", cleaned)


def blueprint_filename_for_paper(paper: PaperPlan) -> str:
    """返回专题/课程综合卷对应细目表的标准文件名。"""
    province = _safe_filename_part(paper.meta.province if paper.meta else "")
    if paper.paper_type == TOPIC_PAPER_TYPE:
        subject = _safe_filename_part(paper.topic)
        return f"{TOPIC_PAPER_TYPE}_{paper.paper_label}_{province}_{subject}.docx"
    if paper.paper_type == COURSE_PAPER_TYPE:
        subject = _safe_filename_part(paper.module)
        return f"{COURSE_PAPER_TYPE}_{paper.paper_label}_{province}_{subject}.docx"
    return ""


def _blueprint_search_dir(paper: PaperPlan, search_dir: Path | None = None) -> Path | None:
    if search_dir is not None:
        return Path(search_dir)
    if paper.meta is None:
        return None
    return paper.meta.path.parent


def find_blueprint_for_paper(paper: PaperPlan, search_dir: Path | None = None) -> Path | None:
    """按标准命名优先查找细目表；失败时在当前考类目录下做保守兜底搜索。"""
    if paper.paper_type not in {TOPIC_PAPER_TYPE, COURSE_PAPER_TYPE}:
        return None
    base_dir = _blueprint_search_dir(paper, search_dir)
    if base_dir is None:
        return None

    expected = blueprint_filename_for_paper(paper)
    exact = base_dir / expected
    if exact.exists() and exact.is_file():
        return exact

    candidates = [
        path
        for path in base_dir.rglob("*.docx")
        if path.is_file()
        and not path.name.startswith("~$")
        and paper.paper_type in path.name
        and paper.paper_label in path.name
    ]
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        paper.blueprint_warnings.append(
            "找到多个可能的细目表，未自动选择：" + "；".join(str(path) for path in candidates)
        )
    return None


def _normalize_header(text: str) -> str:
    return re.sub(r"[\s\n\r　/／（）()\-—_]+", "", _clean_text(text))


_HEADER_ALIASES = {
    "卷别": "paper_type",
    "类型": "paper_type",
    "卷号": "paper_label",
    "地区": "province",
    "省份": "province",
    "课程名称": "module",
    "课程": "module",
    "知识模块": "module",
    "专题名称": "topic",
    "专题": "topic",
    "题号": "question_no",
    "序号": "question_no",
    "题型": "question_type",
    "题目类型": "question_type",
    "难度": "difficulty",
    "考查内容": "content",
    "考察内容": "content",
    "内容": "content",
    "对应考点": "point_name",
    "考点": "point_name",
    "知识点": "knowledge_point",
    "对应知识点": "knowledge_point",
    "考纲要求": "requirement",
    "要求": "requirement",
    "出题意图覆盖说明": "intent",
    "覆盖说明": "intent",
    "出题意图": "intent",
}


def _header_key(text: str) -> str | None:
    normalized = _normalize_header(text)
    return _HEADER_ALIASES.get(normalized)


def _row_texts(row) -> list[str]:
    return [_clean_text(cell.text) for cell in row.cells]


def _question_no(value: str) -> int | None:
    match = re.search(r"\d+", _clean_text(value))
    return int(match.group(0)) if match else None


def _looks_like_question_row(texts: list[str]) -> bool:
    if not texts:
        return False
    return _question_no(texts[0]) is not None


def _is_section_row(texts: list[str]) -> bool:
    values = [text for text in texts if text]
    if not values:
        return False
    unique_values = set(values)
    return len(unique_values) == 1 and _question_no(values[0]) is None


def _make_blueprint_row(data: dict[str, str], paper: PaperPlan | None) -> BlueprintRow:
    meta = paper.meta if paper else None
    question_no_raw = data.get("question_no", "")
    return BlueprintRow(
        question_no=_question_no(question_no_raw),
        question_no_raw=question_no_raw,
        question_type=data.get("question_type", ""),
        difficulty=data.get("difficulty", ""),
        content=data.get("content", ""),
        point_name=data.get("point_name", ""),
        knowledge_point=data.get("knowledge_point", ""),
        requirement=data.get("requirement", ""),
        intent=data.get("intent", ""),
        paper_type=data.get("paper_type", "") or (paper.paper_type if paper else ""),
        paper_label=data.get("paper_label", "") or (paper.paper_label if paper else ""),
        province=data.get("province", "") or (meta.province if meta else ""),
        module=data.get("module", "") or (paper.module if paper else ""),
        topic=data.get("topic", "") or (paper.topic if paper else ""),
        raw=data,
    )


def _parse_blueprint_table(table, paper: PaperPlan | None) -> list[BlueprintRow]:
    rows: list[BlueprintRow] = []
    header_map: dict[int, str] = {}
    current_section_type = ""

    for table_row in table.rows:
        texts = _row_texts(table_row)
        if not any(texts):
            continue

        keys = [_header_key(text) for text in texts]
        recognized = {idx: key for idx, key in enumerate(keys) if key}
        if recognized and ("question_no" in recognized.values() or len(recognized) >= 2):
            header_map = recognized
            continue

        if _is_section_row(texts):
            current_section_type = next(text for text in texts if text)
            continue

        data: dict[str, str] = {}
        if header_map:
            for idx, key in header_map.items():
                if idx < len(texts):
                    data[key] = texts[idx]
        elif len(texts) >= 3 and _looks_like_question_row(texts):
            data = {
                "question_no": texts[0],
                "question_type": current_section_type,
                "difficulty": texts[1],
                "content": texts[2],
                "knowledge_point": texts[2],
            }
        else:
            continue

        if not data.get("question_no") and not _looks_like_question_row(texts):
            continue
        if current_section_type and not data.get("question_type"):
            data["question_type"] = current_section_type
        row = _make_blueprint_row(data, paper)
        if row.question_no is None and not row.question_no_raw:
            continue
        rows.append(row)

    return rows


def parse_blueprint_docx(path: str | Path, paper: PaperPlan | None = None) -> list[BlueprintRow]:
    """解析细目表 DOCX 中的表格为逐题约束。"""
    if Document is None:
        raise RuntimeError("缺少 python-docx，请先安装细目表读取依赖。")
    doc = Document(str(path))
    rows: list[BlueprintRow] = []
    for table in doc.tables:
        rows.extend(_parse_blueprint_table(table, paper))
    return rows


def validate_blueprint_for_paper(paper: PaperPlan) -> list[str]:
    warnings: list[str] = []
    if paper.paper_type not in {TOPIC_PAPER_TYPE, COURSE_PAPER_TYPE}:
        return warnings
    if paper.blueprint_path is None:
        warnings.append(f"未找到细目表：{blueprint_filename_for_paper(paper)}")
        return warnings
    if not paper.blueprint_rows:
        warnings.append(f"细目表未解析到题目行：{paper.blueprint_path}")
        return warnings

    question_numbers = [row.question_no for row in paper.blueprint_rows if row.question_no is not None]
    if len(question_numbers) != len(paper.blueprint_rows):
        warnings.append("细目表存在无法识别题号的行。")
    if len(set(question_numbers)) != len(question_numbers):
        warnings.append("细目表题号存在重复。")
    if question_numbers:
        expected = list(range(min(question_numbers), max(question_numbers) + 1))
        if sorted(question_numbers) != expected:
            warnings.append("细目表题号不连续。")

    for row in paper.blueprint_rows:
        missing = []
        if not row.question_type:
            missing.append("题型")
        if not row.difficulty:
            missing.append("难度")
        if not (row.content or row.knowledge_point):
            missing.append("考查内容/对应知识点")
        if missing:
            warnings.append(f"第{row.question_no_raw or row.question_no}题缺少：{','.join(missing)}")

    if paper.paper_type == TOPIC_PAPER_TYPE:
        blueprint_text = "\n".join(
            f"{row.content}\n{row.point_name}\n{row.knowledge_point}" for row in paper.blueprint_rows
        )
        missing_points = [row.point_name for row in paper.rows if row.point_name and row.point_name not in blueprint_text]
        if missing_points:
            warnings.append("专题细目表可能未覆盖考点：" + "、".join(missing_points))
    return warnings


def attach_blueprint_for_paper(paper: PaperPlan, search_dir: Path | None = None) -> None:
    """为专题训练卷/课程综合卷查找并解析细目表。"""
    paper.blueprint_path = None
    paper.blueprint_rows = []
    paper.blueprint_warnings = []
    if paper.paper_type not in {TOPIC_PAPER_TYPE, COURSE_PAPER_TYPE}:
        return

    path = find_blueprint_for_paper(paper, search_dir)
    if path is None:
        paper.blueprint_warnings.extend(validate_blueprint_for_paper(paper))
        return

    paper.blueprint_path = path
    try:
        paper.blueprint_rows = parse_blueprint_docx(path, paper)
    except Exception as exc:
        paper.blueprint_warnings.append(f"细目表解析失败：{exc}")
        return
    paper.blueprint_warnings.extend(validate_blueprint_for_paper(paper))


def attach_blueprints_for_papers(papers: list[PaperPlan], search_dir: Path | None = None) -> list[str]:
    warnings: list[str] = []
    for paper in papers:
        attach_blueprint_for_paper(paper, search_dir)
        warnings.extend(f"{paper.paper_label}：{warning}" for warning in paper.blueprint_warnings)
    return warnings


def load_plan(plan_path: str | Path) -> list[PlanItem]:
    """读取规划表并生成统一 PlanItem 列表。"""
    if is_flat_plan_workbook(plan_path) or looks_like_flat_plan_workbook(plan_path):
        return load_flat_plan(plan_path)

    meta, _, paper_index = load_planning_workbook(plan_path)
    items: list[PlanItem] = []
    for paper in paper_index.values():
        knowledge_points = "\n".join(row.point_content for row in paper.rows if row.point_content)
        items.append(
            PlanItem(
                paper_no=paper.paper_label,
                paper_type=paper.paper_type,
                province=meta.province,
                exam_category=meta.exam_category,
                subject=paper.module,
                module=paper.module,
                topic=paper.topic or paper.point_name,
                knowledge_points=knowledge_points,
                question_count=sum(row.point_count for row in paper.rows),
                raw={"paper": paper, "rows": paper.rows, "meta": meta},
            )
        )
    return items


def group_by_paper(items: list[PlanItem]) -> dict[str, list[PlanItem]]:
    """按卷号归并规划任务。"""
    grouped: dict[str, list[PlanItem]] = {}
    for item in items:
        key = item.paper_no or "未编号"
        grouped.setdefault(key, []).append(item)
    return grouped
