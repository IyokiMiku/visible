"""学科网课程/题型/知识点 ID 解析器。

从 02_配置资源/学科网映射/ 中读取所有大类映射表，
根据中文课程名和题型名返回对应的 courseId、typeId。
支持全部大类（装备制造、电子信息、财经商贸等）。

知识点匹配支持两种模式：
  1. 精确/模糊字符串匹配（search_kpoints）
  2. AI 辅助批量映射（build_kpoint_map / resolve_kpoints）
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any, Callable

BASE_DIR = Path(__file__).resolve().parent.parent.parent  # 项目根目录
MAPPING_DIR = BASE_DIR / "02_配置资源" / "学科网映射"
_CACHE_DIR = MAPPING_DIR / "kpoint_cache"

# 题型中文名 → typeId 后缀的归一化映射（用于模糊匹配）
_TYPE_SUFFIX_MAP = {
    "单选题": "01", "单选": "01", "单项选择题": "01",
    "多选题": "02", "多选": "02", "多项选择题": "02", "多项选择": "02",
    "判断题": "03", "判断": "03", "是非题": "03",
    "填空题": "04", "填空": "04",
    "简答题": "05", "简答": "05", "问答题": "05",
    "作图题": "06", "作图": "06", "画图题": "06",
    "综合应用题": "07", "综合题": "07", "综合": "07",
    "分析计算题": "08", "计算题": "08", "分析计算": "08",
    "名词解释题": "09", "名词解释": "09",
}

# 全局缓存：避免重复解析同一个 md 文件
_cache: dict[str, tuple[dict[str, int], dict[str, dict[str, str]]]] = {}


def _list_categories() -> list[Path]:
    """返回所有大类映射 md 文件路径。"""
    return sorted(MAPPING_DIR.glob("*.md"))


def _ensure_cache(md_path: Path):
    if str(md_path) in _cache:
        return
    courses: dict[str, int] = {}
    types_by_course: dict[str, dict[str, str]] = {}
    if not md_path.exists():
        _cache[str(md_path)] = (courses, types_by_course)
        return

    text = md_path.read_text(encoding="utf-8")
    # 提取所有课程行
    for match in re.finditer(r"\|\s*(\d+)\s*\|\s*(\S+)\s*\|", text):
        course_id = int(match.group(1))
        course_name = match.group(2).strip()
        courses[course_name] = course_id

        # 提取该课程的题型表
        section_re = re.compile(
            rf"###\s*.*?\(courseId={course_id}\).*?\n\n(.*?)(?=\n###|\Z)", re.S
        )
        section_match = section_re.search(text)
        types: dict[str, str] = {}
        if section_match:
            for tmatch in re.finditer(r"\|\s*(\d+)\s*\|\s*(\S+)\s*\|\s*(yes|no)\s*\|", section_match.group(1)):
                types[tmatch.group(2).strip()] = tmatch.group(1)
        types_by_course[course_name] = types

    _cache[str(md_path)] = (courses, types_by_course)


def resolve_course(course_name: str, category: str | None = None) -> int | None:
    """根据中文课程名返回 courseId。

    category 为 None 时自动搜索全部大类映射文件。
    """
    if category:
        md_paths = [MAPPING_DIR / f"{category}.md"]
    else:
        md_paths = _list_categories()

    for md_path in md_paths:
        _ensure_cache(md_path)
        courses, _ = _cache[str(md_path)]
        if course_name in courses:
            return courses[course_name]
        for name, cid in courses.items():
            if course_name in name or name in course_name:
                return cid
    return None


def resolve_type(course_name: str, type_name: str, category: str | None = None) -> str | None:
    """根据课程名+中文题型名返回 typeId。category 为 None 时自动搜索全部大类。

    支持模糊匹配："单选题"→"单项选择题"，"综合题"→"综合应用题"等。
    """
    if category:
        md_paths = [MAPPING_DIR / f"{category}.md"]
    else:
        md_paths = _list_categories()

    for md_path in md_paths:
        _ensure_cache(md_path)
        courses, types_by_course = _cache[str(md_path)]
        if course_name not in courses:
            continue
        types = types_by_course.get(course_name, {})

        if type_name in types:
            return types[type_name]

        # 模糊匹配：尝试多个后缀
        suffixes_to_try = []
        tn = type_name
        for _ in range(5):
            suffix = _TYPE_SUFFIX_MAP.get(tn)
            if suffix:
                suffixes_to_try.append(suffix)
                tn = suffix
            else:
                break

        for suffix in suffixes_to_try:
            for tname, tid in types.items():
                if tid.endswith(suffix):
                    return tid
    return None


def resolve_all_courses(category: str | None = None) -> dict[str, int]:
    """返回全部大类或指定大类的课程名→courseId 映射。"""
    if category:
        md_paths = [MAPPING_DIR / f"{category}.md"]
    else:
        md_paths = _list_categories()

    result: dict[str, int] = {}
    for md_path in md_paths:
        _ensure_cache(md_path)
        courses, _ = _cache[str(md_path)]
        result.update(courses)
    return result


def resolve_all_types(course_name: str, category: str | None = None) -> dict[str, str]:
    """返回指定课程的全部题型名→typeId 映射。"""
    if category:
        md_paths = [MAPPING_DIR / f"{category}.md"]
    else:
        md_paths = _list_categories()

    for md_path in md_paths:
        _ensure_cache(md_path)
        _, types_by_course = _cache[str(md_path)]
        if course_name in types_by_course:
            return types_by_course[course_name]
    return {}


def load_kpoint_tree(course_name: str, category: str | None = None) -> list[dict[str, Any]]:
    """从 knowledge_points/*.md 中加载指定课程的知识点树。

    每个 md 文件内以代码块形式存储树形结构，格式如下：
    ```
    电路的基本概念 (87644)
    ├── 电路的组成与电路模型 (87645)
    │   ├── 电路组成的基本要素 (87646)
    │   └── 电路模型及电路图 (87647)
    └── 电路的基本物理量 (87650)
    ```

    返回: [{id, name, parent_id, level}, ...]
    跨所有大类 md 文件搜索该课程。
    """
    md_dir = MAPPING_DIR / "knowledge_points"
    if not md_dir.exists():
        return []

    nodes: list[dict[str, Any]] = []
    seen_ids: set[int] = set()

    for md_path in sorted(md_dir.glob("*.md")):
        try:
            content = md_path.read_text(encoding="utf-8")
        except Exception:
            continue

        # 定位课程段落：### 课程名
        section_start = content.find(f"### {course_name}")
        if section_start == -1:
            # 模糊匹配
            for m in re.finditer(r"^### (.+)$", content, re.MULTILINE):
                hdr = m.group(1).strip()
                if hdr in course_name or course_name in hdr:
                    section_start = m.start()
                    break
        if section_start == -1:
            continue

        # 提取段落到下一个 ###
        next_section = content.find("\n### ", section_start + 1)
        section = content[section_start:next_section if next_section != -1 else None]

        # 提取代码块中的树
        code_block = re.search(r"```\n(.*?)```", section, re.S)
        if not code_block:
            continue

        tree_text = code_block.group(1)
        # 用缩进重建层级：按 4 空格 / 2 全角 / ─├└│ 等符号计算
        parent_stack: list[tuple[int, int]] = []  # [(indent, kpoint_id)]

        for line in tree_text.split("\n"):
            m = re.search(r"(.+?)\s*\((\d+)\)", line)
            if not m:
                continue
            name = m.group(1).strip()
            # 去掉树形符号前缀
            name = re.sub(r"^[├└│─\s]+", "", name).strip()
            kpoint_id = int(m.group(2))
            if kpoint_id in seen_ids:
                continue
            seen_ids.add(kpoint_id)

            # 计算缩进：统计行首非ASCII符号数量
            indent = 0
            for ch in line:
                if ch in " │├└─":
                    indent += 1
                elif ch.strip():
                    break
                else:
                    indent += 1
            indent = indent // 4  # 每4字符一级

            # 找到父节点
            parent_id: int | None = None
            level = 1
            while parent_stack and parent_stack[-1][0] >= indent:
                parent_stack.pop()
            if parent_stack:
                parent_id = parent_stack[-1][1]
                level = len(parent_stack) + 1

            parent_stack.append((indent, kpoint_id))
            nodes.append({"id": kpoint_id, "name": name, "parent_id": parent_id, "level": level})

    return nodes


def search_kpoints(query: str, course_name: str, category: str | None = None) -> list[dict[str, Any]]:
    """在指定课程的知识树中按名称模糊搜索知识点。

    匹配策略：
    1. 精确包含匹配（query 是节点名的子串）
    2. 若未命中，尝试关键词拆分匹配（query 中每个词都在节点名中）
    """
    nodes = load_kpoint_tree(course_name, category)
    if not nodes:
        return []

    # 包含匹配
    results = [n for n in nodes if query in n["name"]]
    if results:
        return sorted(results, key=lambda n: n["level"])

    # 关键词拆分匹配
    keywords = [k for k in re.split(r"[、，,、\s]+", query) if len(k) >= 2]
    if keywords:
        results = [n for n in nodes if all(k in n["name"] for k in keywords)]
    return sorted(results, key=lambda n: n["level"])


# ============================================================
# AI 辅助批量知识点映射
# ============================================================

_KPOINT_CACHE: dict[str, dict[str, list[int]]] = {}  # {course_name: {key: [kpointId, ...]}}


def _load_cache(course_name: str) -> dict[str, list[int]]:
    """加载知识点映射缓存文件。"""
    if course_name in _KPOINT_CACHE:
        return _KPOINT_CACHE[course_name]
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = _CACHE_DIR / f"{course_name}_kpoint_map.json"
    if cache_path.exists():
        _KPOINT_CACHE[course_name] = json.loads(cache_path.read_text(encoding="utf-8"))
    else:
        _KPOINT_CACHE[course_name] = {}
    return _KPOINT_CACHE[course_name]


def _save_cache(course_name: str) -> None:
    """保存知识点映射缓存到文件。"""
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = _CACHE_DIR / f"{course_name}_kpoint_map.json"
    cache_path.write_text(
        json.dumps(_KPOINT_CACHE.get(course_name, {}), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def resolve_kpoints(
    query: str,
    course_name: str,
    category: str | None = None,
) -> list[int]:
    """从缓存中查询知识点 ID 列表。先查缓存再走 AI。"""
    cache = _load_cache(course_name)
    return cache.get(query, [])


def load_mapping_table(province: str, exam_category: str) -> dict[str, list[int]]:
    """加载完整映射表，返回 {试卷序号: [知识点ID, ...]}。

    读取 AI 生成的映射表 xlsx，然后自动聚合专题训练卷和课程综合卷的
    kpointIds。如果映射表中 B 列已填则直接使用，否则从规划表中查找组内
    考点训练卷的 ID 合并去重。

    映射表路径：04_生成输出/生产规划/{省份} {考类}/{省份}_{考类}_映射表.xlsx
    """
    from pathlib import Path
    import openpyxl

    BASE = Path(__file__).resolve().parent.parent.parent
    plan_dir = BASE / "04_生成输出" / "生产规划" / f"{province} {exam_category}"
    xlsx_path = plan_dir / f"{province}_{exam_category}_映射表.xlsx"

    if not xlsx_path.exists():
        return {}

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["知识点映射"]
    except Exception:
        return {}

    # 第一遍：读取 AI 匹配结果 + 记录映射方式 + 收集未填的聚合行
    ai_mapping: dict[str, list[int]] = {}       # 考点训练卷（AI匹配，B列已填）
    aggregated_keys: list[str] = []             # 聚合行（B列待填）
    all_rows: list[tuple[str, str, str]] = []   # 全部行数据 (卷号, ID串, 映射方式)

    for row in ws.iter_rows(min_row=2, values_only=True):
        key = str(row[0]).strip() if row[0] else ""
        val = str(row[1]).strip() if row[1] else ""
        method = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if not key or key.startswith("#"):
            continue
        all_rows.append((key, val, method))
        if val:
            ids = [int(x.strip()) for x in val.split(",") if x.strip().isdigit()]
            if method == "AI匹配":
                ai_mapping[key] = ids

    # 第二遍：对 聚合 行做自动聚合；对 AI生成 行留空
    result: dict[str, list[int]] = {}
    for key, val, method in all_rows:
        if val:
            ids = [int(x.strip()) for x in val.split(",") if x.strip().isdigit()]
            result[key] = ids
        elif method == "AI生成":
            # 标记为 AI 生成，不填 ID，由调用方走 AI 直接生成流程
            result[key] = []
        elif method == "聚合":
            # B 列为空 + 聚合 → 自动聚合
            aggregated = _aggregate_kpoints(key, ai_mapping, plan_dir, province, exam_category)
            if aggregated:
                result[key] = aggregated
        # 其他情况（B列空且无明确标记）→ 不填

    return result


def get_mapping_ai_generate_papers(province: str, exam_category: str) -> set[str]:
    """返回映射表中 C 列标记为「AI生成」的试卷序号集合。

    这些试卷的知识树节点与考纲考点不匹配，应跳过 API 拉题，直接由 AI 生成。
    """
    from pathlib import Path
    import openpyxl

    BASE = Path(__file__).resolve().parent.parent.parent
    plan_dir = BASE / "04_生成输出" / "生产规划" / f"{province} {exam_category}"
    xlsx_path = plan_dir / f"{province}_{exam_category}_映射表.xlsx"
    if not xlsx_path.exists():
        return set()

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["知识点映射"]
    except Exception:
        return set()

    ai_keys: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = str(row[0]).strip() if row[0] else ""
        method = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if key and method == "AI生成":
            ai_keys.add(key)
    return ai_keys


def _aggregate_kpoints(
    paper_key: str,
    ai_mapping: dict[str, list[int]],
    plan_dir: Path,
    province: str,
    exam_category: str,
) -> list[int]:
    """根据规划表将 专题/课程综合卷 的 kpointIds 聚合出来。

    策略：
    1. 读取规划表，找到该卷所在 专题(B列) 或 课程(A列) 下的全部考点训练卷
    2. 从 ai_mapping 中取出对应卷号的 ID，合并去重
    """
    import openpyxl

    # 查找规划表 xlsx
    plan_files = list(plan_dir.glob(f"{province}_{exam_category}_考点规划总表.xlsx"))
    if not plan_files:
        return []
    plan_path = plan_files[0]

    try:
        wb = openpyxl.load_workbook(plan_path, data_only=True)
        ws = wb.active
    except Exception:
        return []

    # 收集规划表中每一行的信息：卷号 → (课程, 专题)
    paper_info: dict[str, tuple[str, str, str]] = {}  # {卷号: (课程, 专题, 卷型)}
    # 收集卷型的分布：哪些卷号是 考点训练卷
    point_volumes: set[str] = set()

    for row in ws.iter_rows(min_row=7, values_only=True):  # 第7行起为数据行
        module = str(row[0]).strip() if row[0] else ""      # A列：课程
        topic = str(row[1]).strip() if row[1] else ""        # B列：专题
        f_vol = str(row[5]).strip() if len(row) > 5 and row[5] else ""   # F列：考点训练卷卷号
        h_vol = str(row[7]).strip() if len(row) > 7 and row[7] else ""   # H列：专题训练卷卷号
        j_vol = str(row[9]).strip() if len(row) > 9 and row[9] else ""   # J列：课程综合卷卷号

        if f_vol:
            point_volumes.add(f_vol)
            paper_info[f_vol] = (module, topic, "point")
        if h_vol:
            paper_info[h_vol] = (module, topic, "topic")
        if j_vol:
            # j_vol 可能是 "1,2" 或 "1-3" 格式，需要拆分
            vol_list = _parse_paper_numbers(j_vol)
            for v in vol_list:
                vol_key = f"第{v}卷"
                paper_info[vol_key] = (module, topic, "course")

    if paper_key not in paper_info:
        return []

    module, topic, paper_type = paper_info[paper_key]

    if paper_type == "topic":
        # 专题训练卷：聚合该专题下全部考点训练卷
        source_volumes = [
            k for k, (m, t, pt) in paper_info.items()
            if pt == "point" and t == topic and m == module
        ]
    elif paper_type == "course":
        # 课程综合卷：聚合该课程下全部考点训练卷
        source_volumes = [
            k for k, (m, t, pt) in paper_info.items()
            if pt == "point" and m == module
        ]
    else:
        return []

    ids: list[int] = []
    seen: set[int] = set()
    for vol in sorted(source_volumes):
        for kid in ai_mapping.get(vol, []):
            if kid not in seen:
                ids.append(kid)
                seen.add(kid)

    return ids


def _parse_paper_numbers(raw: str) -> list[str]:
    """解析 "1,2,3" 或 "1-3" 为卷号列表 ['1','2','3']。
    也兼容 "第1-3卷"、"第1卷,第2卷" 等带前后缀的格式。
    """
    import re
    result: list[str] = []
    # 先统一去掉 "第" 前缀和 "卷" 后缀
    cleaned = re.sub(r'^第|第(?=\d)', '', raw.strip())
    cleaned = re.sub(r'卷(?=$|\s*[,，])', '', cleaned)
    parts = re.split(r"[，,]", cleaned)
    for part in parts:
        part = part.strip().rstrip("卷")
        if not part:
            continue
        if "-" in part:
            m = re.match(r"(\d+)\s*-\s*(\d+)", part)
            if m:
                result.extend(str(i) for i in range(int(m.group(1)), int(m.group(2)) + 1))
        else:
            result.append(part)
    return result


def build_kpoint_map(
    course_name: str,
    queries: list[str],
    llm_call: Callable[[str], str],
    category: str | None = None,
    batch_size: int = 20,
) -> dict[str, list[int]]:
    """用 AI 批量匹配知识点文本 → kpointIds，结果写入缓存。

    Args:
        course_name: 课程名（如"电工技术基础与技能"）
        queries: 规划表 D 列考点文本列表（如 ["理解三相对称交流电源的概念和特点", ...]）
        llm_call: 调用 AI 的函数，签名 (prompt) -> response_text
        batch_size: 每批处理的考点数

    Returns:
        {考点文本: [kpointId, ...], ...}
    """
    nodes = load_kpoint_tree(course_name, category)
    if not nodes:
        print(f"错误：{course_name} 知识点树 CSV 不存在，请先放入 02_配置资源/学科网映射/knowledge_points/")
        return {}

    cache = _load_cache(course_name)
    pending = [q for q in queries if q not in cache]
    if not pending:
        print(f"全部 {len(queries)} 个考点已在缓存中。")
        return cache

    print(f"待匹配考点: {len(pending)} 个（已缓存 {len(queries) - len(pending)} 个）")

    # 构建知识树摘要
    tree_text = "\n".join(
        f"{n['id']}\t{'　' * (n['level'] - 1)}{n['name']}"
        for n in nodes
    )

    matched = 0
    for i in range(0, len(pending), batch_size):
        batch = pending[i : i + batch_size]
        query_text = "\n".join(f"{idx + 1}. {q}" for idx, q in enumerate(batch))

        prompt = f"""根据知识树内容，为下列考点找到匹配的知识点 ID。

知识树（id\\t层级缩进\\t名称）：
{tree_text}

待匹配考点：
{query_text}

要求：
1. 每个考点可能对应 1 个或多个知识点 ID。如果考点涵盖多个子知识点的内容，列出所有匹配的 ID。
2. 只输出 JSON，格式：{{"考点原文": [id1, id2, ...], ...}}
3. 如果某个考点在知识树中确实找不到匹配，对应值为空列表 []。
4. 注意相近概念的区别（如"线电压"≠"相电压"，各自匹配各自的知识点）。
5. 只输出 JSON，不要输出其他解释。"""

        try:
            response = llm_call(prompt)
            # 处理 AI 返回的 JSON
            response = response.strip()
            if response.startswith("```"):
                response = re.sub(r"^```(?:json)?\s*|\s*```$", "", response)
            result = json.loads(response)
            for query_text_batch, ids in result.items():
                if isinstance(ids, list) and all(isinstance(x, int) for x in ids):
                    cache[query_text_batch] = ids
                    matched += len(ids)
            _save_cache(course_name)
            print(f"  进度: {min(i + batch_size, len(pending))}/{len(pending)}")
        except Exception as exc:
            print(f"  批次 {i // batch_size + 1} 失败: {exc}")

    print(f"匹配完成: {matched} 个 kpointIds，缓存已保存。")
    return cache
