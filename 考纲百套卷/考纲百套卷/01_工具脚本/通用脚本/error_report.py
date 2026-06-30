"""
试卷报错采集工具

功能：
  1. 从网页复制报错表格（Ctrl+C），程序自动从剪贴板读取
  2. 询问试卷序号，标注来源
  3. 所有记录按试卷号排序，汇总输出到 xlsx 文件
  4. xlsx 中含"已修改"勾选列，方便用户标记修改进度

用法：
  python ocr_report.py
  在网页上选中报错表格 → Ctrl+C → 切回程序按 Enter

依赖：
  pip install openpyxl
"""

import sys
import os
import re

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("请安装依赖: pip install openpyxl")
    sys.exit(1)

_ILLEGAL_XML_CHARS_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]"
)


def _sanitize(value):
    """移除 XML 非法控制字符，防止 openpyxl 写入时报错"""
    if isinstance(value, str):
        return _ILLEGAL_XML_CHARS_RE.sub("", value)
    return value


# ════════════════════ 剪贴板读取 ════════════════════


def get_clipboard_text():
    """从系统剪贴板获取文本"""
    import tkinter as tk
    root = tk.Tk()
    root.withdraw()
    try:
        text = root.clipboard_get()
    except tk.TclError:
        text = ""
    root.destroy()
    return text


# ════════════════════ 表格解析 ════════════════════

SKIP_KEYWORDS = {"退回理由", "退稿理由", "错误类型", "具体位置", "具体描述",
                 "AI文档质检", "退回", "退稿", "勾选"}


def parse_clipboard_table(text):
    """
    解析从网页复制的表格文本。
    网页表格复制后通常每行用换行分隔，列之间用 Tab 分隔。
    返回 [{"错误类型":..., "具体位置":..., "具体描述":...}, ...]
    """
    lines = text.strip().split("\n")
    records = []

    for line in lines:
        line = line.strip()
        if not line:
            continue
        if any(kw in line for kw in SKIP_KEYWORDS):
            continue

        parts = line.split("\t")
        parts = [p.strip() for p in parts if p.strip()]

        if not parts:
            continue

        # 网页表格复制后可能有不同列数，智能适配
        rec = {"错误类型": "", "具体位置": "", "具体描述": ""}

        if len(parts) >= 4:
            # 完整四列：复选框列 | 错误类型 | 具体位置 | 具体描述
            rec["错误类型"] = parts[1]
            rec["具体位置"] = parts[2]
            rec["具体描述"] = parts[3]
        elif len(parts) == 3:
            rec["错误类型"] = parts[0]
            rec["具体位置"] = parts[1]
            rec["具体描述"] = parts[2]
        elif len(parts) == 2:
            rec["错误类型"] = parts[0]
            rec["具体描述"] = parts[1]
        elif len(parts) == 1:
            rec["具体描述"] = parts[0]

        if any(rec.values()):
            records.append(rec)

    return records


# ════════════════════ XLSX 输出 ════════════════════


def save_to_xlsx(all_records, output_path):
    """
    将所有记录按试卷号排序后写入 xlsx。
    all_records: [(试卷号, [记录]), ...]
    """
    all_records.sort(key=lambda r: r[0])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "质检报告"

    header_font = Font(name="宋体", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))
    body_font = Font(name="宋体", size=10)
    odd_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1",
                           fill_type="solid")

    headers = ["试卷序号", "问题类型", "具体位置", "问题详细描述", "已修改"]
    for j, hd in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=hd)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    row_idx = 2
    for exam_num, records in all_records:
        is_odd = exam_num % 2 == 1
        for rec in records:
            cells_data = [
                (1, exam_num, center),
                (2, rec.get("错误类型", ""), center),
                (3, rec.get("具体位置", ""), center),
                (4, rec.get("具体描述", ""), left_wrap),
                (5, "", center),
            ]
            for col, value, align in cells_data:
                cell = ws.cell(row=row_idx, column=col, value=_sanitize(value))
                cell.font = body_font
                cell.alignment = align
                cell.border = thin_border
                if is_odd:
                    cell.fill = odd_fill

            row_idx += 1

    # 第5列添加下拉勾选
    if row_idx > 2:
        dv = DataValidation(type="list", formula1='"✔"', allow_blank=True)
        dv.prompt = "修改完成后选择 ✔"
        dv.promptTitle = "标记已修改"
        ws.add_data_validation(dv)
        dv.add(f"E2:E{row_idx - 1}")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 10

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


# ════════════════════ 主程序 ════════════════════


def main():
    out_dir = os.environ.get("EXAM_TOOL_WORKDIR")
    if not out_dir or not os.path.isdir(out_dir):
        if getattr(sys, "frozen", False):
            out_dir = os.path.dirname(sys.executable)
        else:
            out_dir = os.path.dirname(os.path.abspath(__file__)) or "."

    output_path = os.path.join(out_dir, "AI文档质检报告汇总.xlsx")
    all_records = []

    print("=" * 50)
    print("  试卷报错采集工具")
    print("=" * 50)
    print()
    print("操作说明：")
    print("  1. 按 Enter 开始")
    print("  2. 输入试卷序号")
    print("  3. 在网页上复制报错表格（Ctrl+C），回到本程序按 Enter")
    print("  4. 重复以上步骤采集更多试卷")
    print('  5. 输入 q 结束采集')
    print()
    print(f"  每次录入后自动保存到: {output_path}")
    print()

    while True:
        cmd = input("按 Enter 开始录入（输入 q 结束采集）：").strip()
        if cmd.lower() == "q":
            break

        # 先问试卷序号
        while True:
            num_str = input("  请输入试卷序号（数字）：").strip()
            if num_str.isdigit() and int(num_str) > 0:
                exam_num = int(num_str)
                break
            print("  请输入有效的正整数。")

        # 再让用户复制并粘贴
        input(f"  第 {exam_num} 套 — 请复制网页报错表格后按 Enter 读取：")

        text = get_clipboard_text()
        if not text.strip():
            print("  剪贴板为空，请先在网页上复制表格内容。\n")
            continue

        records = parse_clipboard_table(text)
        if not records:
            print("  未解析到有效记录，请确认已正确复制表格内容。\n")
            continue

        print(f"  解析到 {len(records)} 条报错记录：")
        for i, rec in enumerate(records, 1):
            t = rec['错误类型']
            p = rec['具体位置']
            d = rec['具体描述'][:50]
            print(f"    {i}. [{t}] {p} — {d}")

        found = False
        for idx, (n, recs) in enumerate(all_records):
            if n == exam_num:
                all_records[idx] = (n, recs + records)
                found = True
                break
        if not found:
            all_records.append((exam_num, records))

        # 每次录入后立即保存，防止后续出错丢失数据
        try:
            save_to_xlsx(all_records, output_path)
            print(f"  已保存到 {output_path}")
        except Exception as e:
            print(f"  ⚠ 保存失败: {e}")
            print("  数据仍在内存中，下次录入时会重新尝试保存。")

        total = sum(len(r) for _, r in all_records)
        print(f"  当前共采集 {total} 条记录，"
              f"涉及 {len(all_records)} 套试卷。\n")

    if not all_records:
        print("未采集任何记录，程序退出。")
        return

    # 最终再保存一次（确保排序等最终状态写入）
    try:
        save_to_xlsx(all_records, output_path)
    except Exception as e:
        print(f"最终保存失败: {e}")

    total = sum(len(r) for _, r in all_records)
    print(f"\n报告已生成: {output_path}")
    print(f"共 {len(all_records)} 套试卷，{total} 条报错记录，已按试卷号排序。")

    input("\n按 Enter 退出...")


if __name__ == "__main__":
    main()
