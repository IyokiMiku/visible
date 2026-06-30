"""重新生成映射表，按 L3叶子→L2父节点 优先级匹配 kpoint。

用法：
  python regenerate_mapping.py
"""
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE / "01_工具脚本" / "学科网API拉题移植版"))
from kpoint_resolver import load_kpoint_tree, search_kpoints

PLAN_DIR = BASE / "04_生成输出" / "生产规划" / "重庆市 电子信息类"
PLAN_XLSX = PLAN_DIR / "重庆市_电子信息类_考点规划总表.xlsx"
MAPPING_XLSX = PLAN_DIR / "重庆市_电子信息类_映射表.xlsx"

# ---- 1. 读规划表 ----
import openpyxl

wb = openpyxl.load_workbook(PLAN_XLSX, data_only=True)
ws = wb.active

# 收集数据行（处理合并单元格）
prev_module = ""
prev_topic = ""
rows_data = []

for row in ws.iter_rows(min_row=7, values_only=True):
    cells = list(row)
    # 处理合并单元格：空值继承上一行
    module = str(cells[0]).strip() if cells[0] else prev_module
    topic = str(cells[1]).strip() if cells[1] else prev_topic
    point_name = str(cells[2]).strip() if len(cells) > 2 and cells[2] else ""
    point_content = str(cells[3]).strip() if len(cells) > 3 and cells[3] else ""
    f_vol = str(cells[5]).strip() if len(cells) > 5 and cells[5] else ""
    h_vol = str(cells[7]).strip() if len(cells) > 7 and cells[7] else ""
    j_vol = str(cells[9]).strip() if len(cells) > 9 and cells[9] else ""

    if module:
        prev_module = module
    if topic:
        prev_topic = topic

    # 跳过汇总行
    if "→" in module or "：" in module and "分" in module:
        continue

    rows_data.append({
        "module": module, "topic": topic, "point_name": point_name,
        "point_content": point_content,
        "f_vol": f_vol, "h_vol": h_vol, "j_vol": j_vol,
    })

print(f"规划表数据行: {len(rows_data)}")

# ---- 2. 匹配 kpoint（L3优先→L2→L1兜底但警告） ----
def extract_keywords(d_text: str) -> list[str]:
    """从 D 列提取核心关键词"""
    # 去掉序号和掌握/理解/了解前缀
    cleaned = re.sub(r'^\d+[\.\、．)\s]+', '', d_text.strip())
    cleaned = re.sub(r'^(掌握|理解|了解)\s*', '', cleaned)
    # 按标点拆分
    parts = re.split(r'[，,、;；。]', cleaned)
    return [p.strip() for p in parts if len(p.strip()) >= 2]

def match_kpoint(point_name: str, d_text: str, tree: list[dict], course: str) -> tuple[list[int], str]:
    """匹配 kpoint：先精确→模糊→NLP拆词，L3优先→L2→不可L1"""
    all_matched_ids = set()
    
    candidates = [point_name] + extract_keywords(d_text)
    
    for kw in candidates:
        kw = kw.strip()
        if len(kw) < 2:
            continue
        
        # 策略1：精确包含匹配
        results = [r for r in search_kpoints(kw, course) if r["level"] > 1]
        if results:
            for r in results:
                all_matched_ids.add(r["id"])
            break
        
        # 策略2：拆词匹配 — 将 kw 拆成单字/双字，找包含所有关键词的节点
        chars = list(kw.replace('与', ' ').replace('和', ' ').replace('的', ' '))
        key_parts = [c for c in kw if c not in '与和的及之']
        for node in tree:
            if node["level"] == 1:
                continue
            if all(c in node["name"] for c in key_parts[:4]):  # 至少前4个关键字符都在
                all_matched_ids.add(node["id"])
        if all_matched_ids:
            break
    
    if not all_matched_ids:
        return [], "未匹配"
    
    return sorted(all_matched_ids), "AI匹配"

# ---- 3. 生成映射表 ----
mapping_rows = []  # [(试卷序号, 知识点ID字符串, 映射方式, 备注)]

for course_name in ["电工技术基础与技能", "电子技术基础与技能", "单片机技术与应用"]:
    tree = load_kpoint_tree(course_name)
    if not tree:
        print(f"  [WARN] {course_name}: 无知识树")
        continue
    
    # 该课程的考点训练卷（兼容"单片机技术及应用"和"单片机技术与应用"）
    course_papers = [r for r in rows_data if (
        r["module"] == course_name or
        r["module"].replace("及","与") == course_name.replace("及","与")
    ) and r["f_vol"] and r["point_content"]]
    
    for p in course_papers:
        ids, method = match_kpoint(p["point_name"], p["point_content"], tree, course_name)
        id_str = ",".join(str(i) for i in ids) if ids else ""
        
        note = ""
        if method == "未匹配":
            note = f"知识树无匹配节点"
        elif "排除L1" in method:
            note = f"已排除L1根节点，匹到{len(ids)}个L2/L3节点"
        
        mapping_rows.append({
            "key": p["f_vol"],
            "ids": id_str,
            "method": "AI匹配" if ids else "AI生成",
            "note": note,
            "module": course_name,
            "topic": p["topic"],
        })
        print(f"  {p['f_vol']} {p['point_name']}: {len(ids)} IDs → {method}")

# 专题训练卷 & 课程综合卷（聚合行）
for r in rows_data:
    if r["h_vol"]:
        mapping_rows.append({
            "key": r["h_vol"], "ids": "", "method": "聚合",
            "note": f"专题训练卷，聚合自{r['topic']}",
            "module": r["module"], "topic": r["topic"],
        })
    if r["j_vol"]:
        mapping_rows.append({
            "key": r["j_vol"], "ids": "", "method": "聚合",
            "note": f"课程综合卷，聚合自{r['module']}",
            "module": r["module"], "topic": r["topic"],
        })

# ---- 4. 写入 xlsx（如原文件被占用则写临时名） ----
mwb = openpyxl.Workbook()
mws = mwb.active
mws.title = "知识点映射"
mws.append(["试卷序号", "知识点 ID", "映射方式", "备注"])

for row in mapping_rows:
    mws.append([row["key"], row["ids"], row["method"], row["note"]])

mws.column_dimensions['A'].width = 15
mws.column_dimensions['B'].width = 40
mws.column_dimensions['C'].width = 12
mws.column_dimensions['D'].width = 40

MAPPING_XLSX.parent.mkdir(parents=True, exist_ok=True)
try:
    mwb.save(MAPPING_XLSX)
    print(f"\n[OK] {MAPPING_XLSX}")
except PermissionError:
    alt = MAPPING_XLSX.with_name("重庆市_电子信息类_映射表_NEW.xlsx")
    mwb.save(alt)
    print(f"\n[OK] 原文件被占用，已保存到: {alt}")
print(f"   共 {len(mapping_rows)} 行")
