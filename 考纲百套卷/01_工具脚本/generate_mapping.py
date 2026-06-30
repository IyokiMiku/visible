"""
为重庆市电子信息类生成映射表 xlsx v6
修复: tokenization bug + override逻辑
"""
import re, json
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

BASE = Path(r"C:\Users\zxxk\Desktop\wyy\考纲百套卷")

MANUAL_OVERRIDE = {
    '电工技术基础与技能': {
        '通路、断路、短路三种电路状态的特点': '87648',
        '电容充放电': '87723',
        '线圈的储能公式': '87759,87760',
        '三相对称交流电源的概念和特点': '87835',
    },
    '电子技术基础与技能': {
        '集成运放输入电压': '89525',
        '截止失真和饱和失真': '89511',
        '选频、鉴频、调频、变频等概念': '89585,89608',
        'A/D转换和D/A转换': '89759',
    },
    '单片机技术及应用': {
        '单片机最小系统': '94316',
        'C51语言的基本语法': '94292',
        'C51语言的': '94292',
        '顺序结构、分支结构、循环结构': '94292',
        '中断的概念': '94306',
        '单片机中断': '94306',
        '串行': '94312',
    },
}

def parse_knowledge_tree(fp, course_names):
    text = fp.read_text(encoding='utf-8')
    result = {}
    for c in course_names:
        m = re.search(rf'### {re.escape(c)}\n\n```\n(.*?)```', text, re.DOTALL)
        if not m: continue
        entries = {}
        for line in m.group(1).split('\n'):
            m2 = re.search(r'(.+?)\s*\((\d+)\)\s*$', line)
            if m2:
                entries[re.sub(r'^[│├└─\s]+','', m2.group(1).strip())] = m2.group(2)
        result[c] = entries
        print(f'  {c}: {len(entries)} nodes')
    return result

def read_planning_table(fp):
    wb = load_workbook(fp, data_only=True); ws = wb.active
    hdr = None
    for r in range(1, ws.max_row+1):
        if '知识模块' in str(ws.cell(r,1).value or ''): hdr = r; break
    merged = {}
    for mc in ws.merged_cells.ranges:
        for r in range(mc.min_row, mc.max_row+1):
            for c in range(mc.min_col, mc.max_col+1):
                merged[(r,c)] = mc.min_row
    rows = []; la=lb=lg=lh=li=lj=''
    for ri in range(hdr+1, ws.max_row+1):
        def gv(c):
            v = ws.cell(ri,c).value
            if v is not None: return str(v).strip()
            if (ri,c) in merged:
                vv = ws.cell(merged[(ri,c)],c).value
                return str(vv).strip() if vv else ''
            return ''
        a = gv(1) or la; c = gv(3); d = gv(4); f = gv(6)
        if a.startswith('合'): break
        if not c and not d and not f: continue
        b = gv(2) or lb; gv7 = gv(7) or lg; h = gv(8) or lh
        i = gv(9) or li; j = gv(10) or lj
        if h: lh=h; lg=gv7
        if j: lj=j; li=i
        if a: la=a
        if b: lb=b
        rows.append({'A':a,'B':b,'C':c,'D':d,'F':f,'G':gv7,'H':h,'I':i,'J':j})
    print(f'Planning: {len(rows)} rows')
    return rows

def expand_range(js):
    if not js: return []
    inner = js.replace('第','').replace('卷','')
    m = re.match(r'(\d+)-(\d+)', inner)
    if m: return [f'第{i}卷' for i in range(int(m.group(1)), int(m.group(2))+1)]
    return [js]

MODIFIERS = ['基本','常用','主要','一般','正确','典型','常见','简单']

def tokenize(text):
    text = text.strip()
    for sep in ['、','，',',','；',';','。','．','与','和','及','的']:
        text = text.replace(sep, '|')
    return [t.strip() for t in text.split('|') if len(t.strip()) >= 2]

def fuzzy_match(text, course_tree):
    if not text or len(text) < 2: return None
    clean = text
    for mod in MODIFIERS: clean = clean.replace(mod, '')
    clean = clean.strip()
    
    # 1. exact
    if text in course_tree: return course_tree[text]
    if clean != text and clean in course_tree: return course_tree[clean]
    
    # 2. substring
    for kname, kid in course_tree.items():
        if len(kname) >= 3 and len(text) >= 3:
            if text in kname or kname in text: return kid
            if clean and clean != text and (clean in kname or kname in clean): return kid
    
    # 3. token
    tokens = tokenize(text)
    tokens.sort(key=len, reverse=True)
    for tok in tokens[:10]:
        if len(tok) < 2: continue
        for kname, kid in course_tree.items():
            if tok in kname: return kid
    
    # 4. n-gram 3-5
    for n in range(5,2,-1):
        for i in range(len(clean)-n+1):
            ng = clean[i:i+n]
            for kname, kid in course_tree.items():
                if ng in kname: return kid
    
    return None

def match_d_text(d_text, course_tree, override=None):
    if not d_text or not course_tree: return [], []
    
    # Split entries
    entries = re.split(r'\n', d_text)
    if len(entries) <= 1: entries = re.split(r'(?=\d+\.\s*)', d_text)
    entries = [e.strip() for e in entries if e.strip()]
    
    all_ids, all_um = [], []
    for entry in entries:
        clean = re.sub(r'^\d+\.\s*', '', entry)
        clean = re.sub(r'^(了解并掌握|理解并掌握|了解|理解|掌握)\s*', '', clean)
        clean = clean.rstrip('。.；;，,')
        if len(clean) < 2: continue
        
        # Check manual override (for clean text)
        if override:
            found_ov = False
            for key, ids in override.items():
                if key in clean:
                    for id_ in ids.split(','):
                        if id_ not in all_ids: all_ids.append(id_)
                    found_ov = True; break
            if found_ov: continue
        
        # Auto match
        kid = fuzzy_match(clean, course_tree)
        if kid:
            if kid not in all_ids: all_ids.append(kid); continue
        
        # Sub-entry split
        subs = re.split(r'[；;，,]', clean)
        matched = False
        for sub in subs:
            s = sub.strip()
            if len(s) < 2: continue
            # Override check for sub
            if override:
                found_ov2 = False
                for key, ids in override.items():
                    if key in s:
                        for id_ in ids.split(','):
                            if id_ not in all_ids: all_ids.append(id_)
                        found_ov2 = True; break
                if found_ov2: matched = True; break
            kid = fuzzy_match(s, course_tree)
            if kid:
                if kid not in all_ids: all_ids.append(kid)
                matched = True
        if not matched:
            all_um.append(clean)
    return all_ids, all_um

COURSE_MAP = {
    '电工技术基础与技能': '电工技术基础与技能',
    '电子技术基础与技能': '电子技术基础与技能',
    '单片机技术及应用': '单片机技术与应用',
}

def main():
    print('=== Knowledge Tree ===')
    ktree = parse_knowledge_tree(
        BASE/'02_配置资源/学科网映射/knowledge_points/装备制造大类.md',
        list(COURSE_MAP.values()))
    
    print('\n=== Planning ===')
    rows = read_planning_table(BASE/'04_生成输出/生产规划/重庆市 电子信息类/重庆市_电子信息类_考点规划总表.xlsx')
    
    print('\n=== Papers ===')
    papers = []; sh = set(); sj = set()
    for row in rows:
        if row['F']:
            papers.append({'seq':row['F'],'type':'考点训练卷','course':row['A'],'topic':row['B'],'d_text':row['D']})
        if row['H'] and row['H'] not in sh:
            sh.add(row['H'])
            papers.append({'seq':row['H'],'type':'专题训练卷','course':row['A'],'topic':row['B'],'d_text':''})
        if row['J']:
            for e in expand_range(row['J']):
                if e not in sj:
                    sj.add(e)
                    papers.append({'seq':e,'type':'课程综合卷','course':row['A'],'topic':'','d_text':''})
    print(f'Total: {len(papers)}')
    for t in ['考点训练卷','专题训练卷','课程综合卷']:
        print(f'  {t}: {sum(1 for p in papers if p["type"]==t)}')
    
    print('\n=== Matching ===')
    um_list = []
    for p in papers:
        if p['type'] != '考点训练卷':
            p['kpoint_ids']=''; p['method']='聚合'; p['remark']=''; continue
        tc = COURSE_MAP.get(p['course'])
        if not tc or tc not in ktree:
            p['kpoint_ids']=''; p['method']='AI匹配'; p['remark']='无知识树'; continue
        ov = MANUAL_OVERRIDE.get(p['course'])
        ids, um = match_d_text(p['d_text'], ktree[tc], ov)
        p['kpoint_ids']=','.join(ids); p['method']='AI匹配'; p['remark']=''
        for u in um: um_list.append((p['seq'],p['course'],u))
    
    # Remarks
    tv = {}; cv = {}
    for row in rows:
        if row['F']:
            v = row['F'].replace('第','').replace('卷','')
            tv.setdefault(row['B'],[]).append(v)
            cv.setdefault(row['A'],[]).append(v)
    for p in papers:
        if p['type']=='专题训练卷':
            p['remark'] = f"聚合自第{'、'.join(tv.get(p['topic'],[]))}卷"
        elif p['type']=='课程综合卷':
            p['remark'] = f"聚合自第{'、'.join(cv.get(p['course'],[]))}卷"
    
    # Write
    print('\n=== Writing ===')
    wb = Workbook(); ws = wb.active; ws.title = '知识点映射'
    hf = Font(bold=True, size=11); hfi = PatternFill('solid', fgColor='D9E1F2')
    bd = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
    for ci, h in enumerate(['试卷序号','知识点ID','映射方式','备注'], 1):
        c = ws.cell(1,ci,h); c.font=hf; c.fill=hfi; c.border=bd
        c.alignment=Alignment(horizontal='center',vertical='center')
    ws.column_dimensions['A'].width=14; ws.column_dimensions['B'].width=70
    ws.column_dimensions['C'].width=12; ws.column_dimensions['D'].width=60
    
    for i, p in enumerate(papers):
        r = i + 2
        ws.cell(r,1,p['seq']).border=bd; ws.cell(r,1).alignment=Alignment(vertical='center')
        ws.cell(r,2,p['kpoint_ids']).border=bd; ws.cell(r,2).alignment=Alignment(wrap_text=True,vertical='center')
        ws.cell(r,3,p['method']).border=bd; ws.cell(r,3).alignment=Alignment(horizontal='center',vertical='center')
        ws.cell(r,4,p.get('remark','')).border=bd; ws.cell(r,4).alignment=Alignment(wrap_text=True,vertical='center')
    
    if um_list:
        nr = len(papers)+3
        ws.cell(nr,1,'# 未匹配列表').font=Font(bold=True,color='FF0000')
        for j,(s,co,e) in enumerate(um_list):
            ws.cell(nr+1+j,1,s); ws.cell(nr+1+j,2,f'[{co}] {e}')
    
    out = BASE / '04_生成输出/生产规划/重庆市 电子信息类/重庆市_电子信息类_映射表.xlsx'
    wb.save(str(out))
    
    ai = sum(1 for p in papers if p['method']=='AI匹配')
    ok = sum(1 for p in papers if p['method']=='AI匹配' and p['kpoint_ids'])
    ag = sum(1 for p in papers if p['method']=='聚合')
    print(f'\nDone: {out.name}')
    print(f'  AI匹配: {ai} (成功: {ok}/{ai})')
    print(f'  聚合: {ag}')
    print(f'  未匹配: {len(um_list)}')
    if um_list:
        print('\n未匹配条目:')
        for s,co,e in um_list: print(f'  {s} [{co}] {e[:80]}')

if __name__=='__main__':
    main()
