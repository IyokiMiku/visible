"""Write planning-table workbooks."""

import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 样式常量
FILL_HEADER = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
FILL_COURSE = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
FILL_SECTION = PatternFill(start_color="00000000", end_color="00000000", fill_type=None)
FILL_NORMAL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
FILL_IMPORTANT = PatternFill(start_color="FFFFD7D7", end_color="FFFFD7D7", fill_type="solid")

FONT_HEADER = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
FONT_COURSE = Font(name="微软雅黑", size=10, bold=True)
FONT_SECTION = Font(name="微软雅黑", size=10, bold=True)
FONT_NORMAL = Font(name="微软雅黑", size=9)
FONT_TITLE = Font(name="微软雅黑", size=12, bold=True)

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _strip_knowledge_trailing_period(text):
    """B列考纲知识点不在句尾追加句号；只去掉末尾句号，保留原文其他内容。"""
    return re.sub(r"[。．.]+\s*$", "", str(text or "").strip())


def write_planning_xlsx(output_path, title, config_line, info_line, topics, textbooks):
    """生成带格式的规划表 xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考点规划表"

    # 列宽设置
    col_widths = [6, 60, 18, 8, 28, 10, 6, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # === 表头区域 ===
    # Row 1: 标题
    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, title)
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: 配置行
    ws.merge_cells("A2:H2")
    ws.cell(2, 1, config_line).font = Font(name="微软雅黑", size=9)

    # Row 3: 信息行
    ws.merge_cells("A3:H3")
    ws.cell(3, 1, info_line).font = Font(name="微软雅黑", size=9)

    # Row 4: 空行
    # Row 5: 表头
    headers = ["序号", "考纲知识点", "试卷主题", "级别", "题型", "难度", "套数", "考纲标号"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(5, col, h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    # === 数据区域 ===
    row = 6
    for item in topics:
        if item["type"] == "course":
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row, 1, item["text"])
            c.font = FONT_COURSE
            c.fill = FILL_COURSE
            c.alignment = Alignment(vertical="center")
            row += 1

        elif item["type"] == "section":
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row, 1, f"  {item['text']}")
            c.font = FONT_SECTION
            c.alignment = Alignment(vertical="center")
            row += 1

        elif item["type"] == "topic":
            fill = FILL_IMPORTANT if item["level"] == "极重要" else FILL_NORMAL
            values = [
                item["seq"],
                _strip_knowledge_trailing_period(item["knowledge"]),
                item["theme"],
                item["level"],
                item["question_types"],
                item["difficulty"],
                item["sets"],
                item["exam_ref"],
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row, col, val)
                c.font = FONT_NORMAL
                c.fill = fill
                c.border = THIN_BORDER
                if col in (1, 4, 6, 7):
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(vertical="center", wrap_text=True)
            row += 1

    # 保存
    wb.save(output_path)
    return row - 6  # 数据行数
