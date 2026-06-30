"""内蒙古机电类一课一练考点规划表生成器
无考纲模式：依据教材目录自建主题并扩写知识点。
"""
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
import sys

from planning_assets import prepare_planning_assets, validate_question_plan

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUTPUT_DIR = os.path.join(BASE_DIR, "04_生成输出", "考点规划表", "内蒙古", "机电类")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# === 样式 ===
FONT_TITLE = Font(name="微软雅黑", size=12, bold=True)
FONT_HEADER = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
FONT_COURSE = Font(name="微软雅黑", size=10, bold=True)
FONT_SECTION = Font(name="微软雅黑", size=9, bold=True)
FONT_DATA = Font(name="微软雅黑", size=9)
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_COURSE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_SECTION = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_IMPORTANT = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_STANDARD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"))
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

DIFFICULTY = "80:10:10"
QTYPES_COMPUTE = "单选4+判断2+填空2+计算1+综合1"
QTYPES_NONCOMP = "单选4+判断2+填空2+简答1+综合1"
QUESTION_TOTAL = 10

CN_NUMS = "一二三四五六七八九十"

def has_computation(desc):
    kw = ["计算","公式","估算","定量","代入","求解","方程","算出",
          "功率","阻抗","电流","电压","电阻","感抗","容抗","分压",
          "分流","等效","欧姆定律","基尔霍夫","戴维宁","叠加","谐振",
          "变压","变比","安培力","感应电动势","磁通","充放电",
          "惠斯通","电桥","伏安法","有功","无功","视在","频率","误差"]
    return any(k in desc for k in kw)

def get_qtypes(desc, override_qtypes=None):
    if override_qtypes:
        return override_qtypes
    return QTYPES_COMPUTE if has_computation(desc) else QTYPES_NONCOMP

def get_level(desc):
    if desc.startswith("了解"): return "标准"
    return "重要"


def create_planning(book_data, filename, qtypes_override=None, total_questions=QUESTION_TOTAL):
    """生成xlsx：课程-节标题-数据行的层次结构"""
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, w in {"A":6,"B":60,"C":18,"D":8,"E":30,"F":10,"G":6,"H":16}.items():
        ws.column_dimensions[col].width = w

    # Row 1-3: 标题、配置、教材
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value=book_data["title"]); c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A2:H2")
    config_text = f"题型：{qtypes_override} | 难度：{DIFFICULTY}" if qtypes_override else book_data["config"]
    c = ws.cell(row=2, column=1, value=config_text); c.font = FONT_DATA
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A3:H3")
    c = ws.cell(row=3, column=1, value=book_data["textbook"]); c.font = FONT_DATA
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 5: 列标题
    for ci, h in enumerate(["序号","考纲知识点","试卷主题","级别","题型","难度","套数","考纲标号"], 1):
        c = ws.cell(row=5, column=ci, value=h); c.font = FONT_HEADER
        c.fill = FILL_HEADER; c.alignment = ALIGN_CENTER

    row = 6; seq = 1
    for course in book_data["courses"]:
        # 课程行 (merged)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=course["name"])
        c.font = FONT_COURSE; c.fill = FILL_COURSE; c.alignment = ALIGN_CENTER
        row += 1

        for sec in course["sections"]:
            # 节标题行：序号列为空，知识点列写"一、xxx"
            ws.cell(row=row, column=1, value="").font = FONT_SECTION
            c = ws.cell(row=row, column=2, value=sec["title"]); c.font = FONT_SECTION
            c.alignment = ALIGN_LEFT
            for ci in range(1,9):
                ws.cell(row=row, column=ci).fill = FILL_SECTION
            row += 1

            for t in sec["topics"]:
                topic_name, knowledge_desc = t[0], t[1]
                level = get_level(knowledge_desc)
                qtypes = get_qtypes(knowledge_desc, qtypes_override)
                validate_question_plan(qtypes, total_questions, label=f"{topic_name}题型配置")
                exam_ref = f"教材{seq}"
                suffix_list = ["(一)","(二)"] if level=="重要" else [""]
                for suffix in suffix_list:
                    theme = f"{topic_name}{suffix}" if suffix else topic_name
                    ws.cell(row=row, column=1, value=seq).font = FONT_DATA
                    ws.cell(row=row, column=1).alignment = ALIGN_CENTER
                    c = ws.cell(row=row, column=2, value=re.sub(r"[。．.]+\s*$", "", str(knowledge_desc).strip()))
                    c.font = FONT_DATA; c.alignment = ALIGN_LEFT
                    c = ws.cell(row=row, column=3, value=theme)
                    c.font = FONT_DATA; c.alignment = ALIGN_CENTER
                    c = ws.cell(row=row, column=4, value=level)
                    c.font = FONT_DATA; c.alignment = ALIGN_CENTER
                    c = ws.cell(row=row, column=5, value=qtypes)
                    c.font = FONT_DATA; c.alignment = ALIGN_CENTER
                    c = ws.cell(row=row, column=6, value=DIFFICULTY)
                    c.font = FONT_DATA; c.alignment = ALIGN_CENTER
                    ws.cell(row=row, column=7, value=1).font = FONT_DATA
                    ws.cell(row=row, column=7).alignment = ALIGN_CENTER
                    c = ws.cell(row=row, column=8, value=exam_ref)
                    c.font = FONT_DATA; c.alignment = ALIGN_CENTER
                    fill = FILL_IMPORTANT if level=="重要" else FILL_STANDARD
                    for ci in range(1,9):
                        ws.cell(row=row, column=ci).fill = fill
                    seq += 1; row += 1

    # 边框
    for r in range(1, row):
        for ci in range(1,9):
            ws.cell(row=r, column=ci).border = THIN_BORDER

    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path); wb.close()
    return path, seq-1


# ======== 数据定义 ========
CONFIG = "题型：单选4+判断2+填空2+计算1+综合1（有计算）/ 单选4+判断2+填空2+简答1+综合1（无计算） | 难度：80:10:10"

DG = { # 电工技术基础与技能
    "title": "内蒙古自治区机电类《一课一练》考点规划表",
    "config": CONFIG,
    "textbook": "参考教材：《电工技术基础与技能》（高教版·第3版）周邵敏",
    "courses": [
        {"name":"绪论","sections":[
            {"title":"一、电路基本概念","topics":[
                ("电路的基本概念","了解电路的概念、组成及电路图的表示方法，认识电源、负载、导线和开关等基本组成部分。"),
                ("电路模型","理解电路模型的概念，会建立简单实际电路的电路模型，区分理想元件与实际元件的差异。"),
            ]}
        ]},
        {"name":"第1章 直流电路基础","sections":[
            {"title":"一、电流与电阻","topics":[
                ("电流的形成","理解电流的形成条件与方向规定，掌握电流的定义及单位换算。"),
                ("电阻与温度关系","理解电阻的概念及影响因素(R=ρL/S)，了解电阻温度系数及其工程意义。"),
            ]},
            {"title":"二、欧姆定律与电功率","topics":[
                ("部分电路欧姆定律","掌握部分电路欧姆定律I=U/R，会计算简单电路中的电压、电流、电阻关系。"),
                ("电功与电功率","掌握电功W=UIt和电功率P=UI的计算，理解焦耳定律Q=I²Rt的含义和应用。"),
            ]}
        ]},
        {"name":"第2章 简单直流电路","sections":[
            {"title":"一、全电路欧姆定律与串并联","topics":[
                ("电动势与全电路欧姆定律","理解电动势的概念，掌握全电路欧姆定律I=E/(R+r)，会分析端电压随负载变化的关系。"),
                ("电阻的串联","掌握电阻串联电路的特点（等电流、分压），会计算等效电阻和分压关系。"),
                ("电阻的并联","掌握电阻并联电路的特点（等电压、分流），会计算等效电阻1/R=1/R1+1/R2和分流关系。"),
                ("电阻的混联","掌握混联电路的分析方法，能按要求逐步化简，计算总等效电阻和各支路电流。"),
            ]},
            {"title":"二、测量与电位","topics":[
                ("万用表的使用","了解万用表的结构和测量原理，掌握测量电压、电流、电阻的正确操作方法。"),
                ("电桥测量电阻","了解惠斯通电桥的结构和平衡条件，会用直流电桥精确测量电阻值。"),
                ("电路中各点电位的计算","掌握电位的概念和参考点选取原则，会计算电路中各点的电位值。"),
            ]}
        ]},
        {"name":"第3章 复杂直流电路","sections":[
            {"title":"一、基本定律与定理","topics":[
                ("基尔霍夫定律","理解支路、节点、回路的概念，掌握基尔霍夫电流定律(∑I=0)和电压定律(∑U=0)的表述及应用。"),
                ("支路电流法","掌握支路电流法的解题步骤，会用基尔霍夫定律列方程求解复杂电路。"),
            ]},
            {"title":"二、网络定理","topics":[
                ("叠加定理","理解叠加定理的适用条件（线性电路），会用叠加定理分析多电源电路。"),
                ("戴维宁定理","理解二端网络和有源二端网络的概念，会用戴维宁定理将复杂电路等效为电压源串联电阻。"),
                ("两种电源模型的等效变换","理解电压源和电流源模型，掌握两者等效变换的条件和步骤。"),
            ]}
        ]},
        {"name":"第4章 电容","sections":[
            {"title":"一、电容器基础","topics":[
                ("电容器与电容","理解电容器的结构和电容的定义C=Q/U，掌握平行板电容器电容C=εS/d的计算。"),
                ("电容器的连接","掌握电容器串联和并联的特点，会计算等效电容和分压关系。"),
                ("电容器的充放电","理解电容器的充电和放电过程，了解时间常数τ=RC的含义和电容器储能公式。"),
            ]}
        ]},
        {"name":"第5章 磁场和磁路","sections":[
            {"title":"一、磁场基础","topics":[
                ("电流的磁效应与磁场","理解磁场的基本概念和磁感线，会用安培定则判断电流产生的磁场方向。"),
                ("磁场的主要物理量","掌握磁感应强度B、磁通Φ、磁导率μ、磁场强度H的定义、单位及相互关系。"),
            ]},
            {"title":"二、磁力与磁路","topics":[
                ("磁场对通电导线的作用力","掌握安培力F=BILsinθ的计算，会用左手定则判断安培力方向。"),
                ("铁磁性物质的磁化","理解铁磁性物质的磁化过程、磁化曲线和磁滞回线，区分硬磁和软磁材料。"),
                ("磁路的基本概念","理解磁路与电路的类比关系，掌握磁路欧姆定律Φ=F/Rm。"),
            ]}
        ]},
        {"name":"第6章 电磁感应","sections":[
            {"title":"一、电磁感应定律","topics":[
                ("电磁感应现象与方向","理解电磁感应现象产生的条件，会用楞次定律判断感应电流的方向。"),
                ("法拉第电磁感应定律","掌握法拉第电磁感应定律E=NΔΦ/Δt，会计算感应电动势的大小。"),
            ]},
            {"title":"二、自感互感与涡流","topics":[
                ("自感","理解自感现象、自感系数L和自感电动势，了解磁场能量的计算。"),
                ("互感","理解互感现象和互感系数M，了解互感在变压器和互感器中的应用。"),
                ("涡流与集肤效应","了解涡流的产生、利弊及应用，理解集肤效应对交流导体的影响。"),
            ]}
        ]},
        {"name":"第7章 单相正弦交流电","sections":[
            {"title":"一、交流电基本概念","topics":[
                ("正弦交流电的产生","理解正弦交流电的产生原理，掌握正弦量的三要素（最大值、角频率、初相位）。"),
                ("交流电的物理量","理解最大值、有效值、平均值的关系（有效值=最大值/√2），掌握周期、频率、角频率的关系。"),
            ]},
            {"title":"二、正弦量的表示","topics":[
                ("正弦量的表示方法","掌握解析式、波形图和矢量图三种表示正弦量的方法及其相互转换。"),
                ("非正弦周期波","了解非正弦周期波的概念和组成，理解谐波分析的基本思路。"),
            ]}
        ]},
        {"name":"第8章 正弦交流电路","sections":[
            {"title":"一、单一参数电路","topics":[
                ("纯电阻电路","理解纯电阻电路中电流与电压同相的关系，掌握有功功率P=UI的计算。"),
                ("纯电感电路","理解感抗XL=2πfL的概念，掌握纯电感电路中电流滞后电压90°的相位关系和功率特点。"),
                ("纯电容电路","理解容抗XC=1/(2πfC)的概念，掌握纯电容电路中电流超前电压90°的相位关系和功率特点。"),
            ]},
            {"title":"二、串联电路与功率","topics":[
                ("RL串联电路","掌握RL串联电路的阻抗Z=√(R²+XL²)、阻抗角和功率因数的计算，会画矢量图。"),
                ("RC串联电路","掌握RC串联电路的阻抗Z=√(R²+XC²)、阻抗角和功率因数的计算，会画矢量图。"),
                ("RLC串联电路","掌握RLC串联电路的分析方法，会判断电路呈感性、容性或阻性。"),
                ("交流电路的功率","掌握有功功率、无功功率和视在功率的概念及三者关系S²=P²+Q²，理解功率因数的意义和提高方法。"),
            ]}
        ]},
        {"name":"第9章 谐振电路","sections":[
            {"title":"一、谐振电路","topics":[
                ("串联谐振","理解串联谐振的条件(f=1/(2π√LC))、特点和品质因数Q，了解串联谐振的选频应用。"),
                ("并联谐振","理解并联谐振的条件和特点，了解并联谐振在选频电路中的应用。"),
            ]}
        ]},
        {"name":"第10章 三相正弦交流电路","sections":[
            {"title":"一、三相电源与负载","topics":[
                ("三相电源","理解三相交流电动势的产生和对称性，掌握三相电源的星形和三角形接法及线电压与相电压的关系。"),
                ("三相负载的连接","掌握三相负载星形接法(Y)和三角形接法(Δ)的特点，会计算对称负载的线电流和相电流。"),
                ("三相电路的功率","掌握对称三相电路有功功率P=√3ULILcosφ的计算，了解三相四线制中性线的作用。"),
            ]},
            {"title":"二、安全用电与变压器","topics":[
                ("安全用电","了解触电原因和安全电压，掌握保护接地和保护接零的区别及适用场合，了解电气火灾的预防。"),
                ("变压器","了解变压器的结构和工作原理（变压比=匝数比），掌握电压、电流、阻抗变换的关系。"),
            ]}
        ]},
        {"name":"第11章 暂态过程","sections":[
            {"title":"一、暂态分析","topics":[
                ("换路定律","理解暂态过程的概念和换路定律uc(0+)=uc(0-)、iL(0+)=iL(0-)，会计算初始值。"),
                ("RC电路的暂态过程","理解RC电路充放电的规律和三要素法，会分析RC电路暂态响应的变化趋势。"),
            ]}
        ]},
    ]
}

DZ = { # 电工电子技术与技能
    "title": "内蒙古自治区机电类《一课一练》考点规划表",
    "config": CONFIG,
    "textbook": "参考教材：《电工电子技术与技能》（高教版·第3版）程周",
    "courses": [
        {"name":"第1章 电路的基本概念与基本定律","sections":[
            {"title":"一、电路基础知识","topics":[
                ("电路的组成与电路模型","了解电路的组成和各部分作用，理解电路模型的概念，会建立简单实际电路的电路模型。"),
                ("电流与电压","理解电流的形成条件、方向规定，掌握电压与电位的关系及参考方向的设定。"),
            ]},
            {"title":"二、欧姆定律与电功率","topics":[
                ("电阻与欧姆定律","掌握部分电路欧姆定律和电阻的定义，会进行简单电路的计算。"),
                ("电功与电功率","掌握电功W=UIt和电功率P=UI的计算，理解焦耳定律和额定值的含义。"),
            ]}
        ]},
        {"name":"第2章 直流电路分析","sections":[
            {"title":"一、电阻电路","topics":[
                ("电阻串并联电路","掌握电阻串联、并联及混联电路的等效变换和分析计算方法。"),
            ]},
            {"title":"二、网络分析方法","topics":[
                ("基尔霍夫定律","掌握基尔霍夫电流定律和电压定律，会用节点法或回路法分析直流电路。"),
                ("电压源与电流源","理解理想电压源和电流源的特性，掌握两种电源模型的等效变换。"),
                ("叠加定理与戴维宁定理","会用叠加定理分析多电源线性电路，会用戴维宁定理简化电路分析。"),
            ]}
        ]},
        {"name":"第3章 磁场与电磁感应","sections":[
            {"title":"一、磁场与电磁力","topics":[
                ("磁场的基本概念","理解磁场的产生和性质，掌握磁感应强度、磁通、磁导率等物理量的含义。"),
                ("电流的磁效应与安培力","会用安培定则判断电流磁场，会用左手定则判断安培力方向，会计算安培力大小。"),
            ]},
            {"title":"二、电磁感应","topics":[
                ("电磁感应定律","理解电磁感应现象，掌握法拉第电磁感应定律和楞次定律，会计算感应电动势。"),
                ("自感与互感","理解自感和互感现象及其在电路中的应用，了解变压器的工作原理。"),
            ]}
        ]},
        {"name":"第4章 单相正弦交流电路","sections":[
            {"title":"一、交流电基础","topics":[
                ("正弦交流电的基本概念","理解正弦交流电的产生，掌握三要素（最大值、频率、初相）和有效值的含义。"),
            ]},
            {"title":"二、交流电路分析","topics":[
                ("单一参数交流电路","掌握纯电阻、纯电感、纯电容电路中电压与电流的相位关系和功率特点。"),
                ("RLC串联电路","掌握RLC串联电路的分析方法，会计算阻抗、电流、功率因数和各元件功率。"),
                ("功率因数的提高","理解功率因数的含义和提高方法，了解并联电容器补偿的原理和计算。"),
            ]}
        ]},
        {"name":"第5章 三相交流电路","sections":[
            {"title":"一、三相电路","topics":[
                ("三相电源与三相负载","理解对称三相电源的特点，掌握三相负载星形和三角形接法的分析和计算。"),
                ("三相电路的功率","掌握三相电路有功功率的计算，了解三相四线制中性线的作用和安全用电常识。"),
            ]}
        ]},
        {"name":"第6章 变压器与电动机","sections":[
            {"title":"一、变压器","topics":[
                ("变压器的原理与应用","理解变压器的工作原理，掌握电压、电流、阻抗变换的计算，了解常用变压器的类型。"),
            ]},
            {"title":"二、电动机","topics":[
                ("三相异步电动机","了解三相异步电动机的结构和旋转磁场，理解工作原理和铭牌参数，了解起动和调速方法。"),
                ("单相异步电动机","了解单相异步电动机的类型和特点，理解电容分相起动和罩极式电机的工作原理。"),
            ]}
        ]},
        {"name":"第7章 半导体器件","sections":[
            {"title":"一、半导体基础","topics":[
                ("半导体基础知识与PN结","了解半导体的导电特性，理解PN结的形成和单向导电性。"),
            ]},
            {"title":"二、二极管与三极管","topics":[
                ("二极管及其应用","了解二极管的结构、伏安特性和主要参数，会分析整流、滤波和稳压电路。"),
                ("三极管及其特性","了解三极管的结构、电流放大原理和工作状态，掌握输入输出特性曲线。"),
            ]}
        ]},
        {"name":"第8章 放大电路基础","sections":[
            {"title":"一、基本放大电路","topics":[
                ("基本放大电路","理解共射极放大电路的组成和各元件作用，掌握静态工作点的设置和估算方法。"),
                ("放大电路的性能指标","理解放大倍数、输入电阻、输出电阻的含义，会估算基本放大电路的主要指标。"),
            ]},
            {"title":"二、集成运放","topics":[
                ("集成运算放大器","了解集成运放的特点和理想特性，掌握反相、同相比例运算电路的分析方法。"),
            ]}
        ]},
        {"name":"第9章 直流稳压电源","sections":[
            {"title":"一、整流滤波","topics":[
                ("整流电路","理解半波、桥式整流电路的组成和工作原理，会估算输出电压和电流。"),
                ("滤波电路","了解电容滤波、电感滤波电路的原理，会分析滤波效果的影响因素。"),
            ]},
            {"title":"二、稳压电路","topics":[
                ("稳压电路","了解硅稳压管稳压电路和集成稳压器的工作原理及典型应用。"),
            ]}
        ]},
        {"name":"第10章 数字电路基础","sections":[
            {"title":"一、逻辑基础","topics":[
                ("数字信号与数制","理解模拟信号与数字信号的区别，掌握二进制、十进制的转换方法。"),
                ("基本逻辑门电路","理解与、或、非、与非、或非、异或门的逻辑功能和符号，会写真值表。"),
            ]},
            {"title":"二、组合与时序逻辑","topics":[
                ("组合逻辑电路","了解组合逻辑电路的分析和设计方法，会进行逻辑函数的化简。"),
                ("触发器","理解基本RS、JK、D触发器的逻辑功能和特性，了解触发器的应用。"),
            ]}
        ]},
        {"name":"第11章 电力电子技术基础","sections":[
            {"title":"一、晶闸管与可控整流","topics":[
                ("晶闸管的结构与原理","了解晶闸管的结构、导通和关断条件，理解其伏安特性。"),
                ("可控整流电路","了解单相半波和桥式可控整流电路的工作原理和移相控制的概念。"),
            ]},
            {"title":"二、交流调压与变频","topics":[
                ("交流调压与变频","了解交流调压电路的基本原理和变频技术的概念。"),
            ]}
        ]},
    ]
}

CL = { # 电气测量技术
    "title": "内蒙古自治区机电类《一课一练》考点规划表",
    "config": CONFIG,
    "textbook": "参考教材：《电气测量技术》（高教版·第四版）文春帆",
    "courses": [
        {"name":"第1章 测量基础知识","sections":[
            {"title":"一、测量基本概念","topics":[
                ("测量的基本概念","了解测量的定义和分类，理解测量方法和测量误差的来源及分类。"),
                ("测量误差与数据处理","掌握绝对误差、相对误差和引用误差的计算，了解系统误差和随机误差的处理方法。"),
                ("测量仪表的选用","了解测量仪表的准确度等级，会正确选择仪表量程和精度。"),
            ]}
        ]},
        {"name":"第2章 电流与电压的测量","sections":[
            {"title":"一、直流测量","topics":[
                ("直流电流的测量","掌握电流表的使用方法和量程扩展（分流器），会正确连接电流表进行测量。"),
                ("直流电压的测量","掌握电压表的使用方法和量程扩展（分压器），会正确连接电压表进行测量。"),
            ]},
            {"title":"二、交流测量与万用表","topics":[
                ("交流电流与电压的测量","了解交流仪表的特点，掌握交流电流表和电压表的使用方法和注意事项。"),
                ("万用表的原理与使用","理解万用表的基本结构和工作原理，掌握测量交直流电压、电流和电阻的方法。"),
            ]}
        ]},
        {"name":"第3章 电阻的测量","sections":[
            {"title":"一、电阻测量方法","topics":[
                ("伏安法测电阻","理解伏安法测电阻的原理，掌握电流表内接法和外接法的选择及误差分析。"),
                ("电桥法测电阻","理解惠斯通电桥和开尔文电桥的平衡条件，会用直流电桥精确测量电阻。"),
                ("绝缘电阻的测量","了解兆欧表的结构和工作原理，掌握测量绝缘电阻的正确操作方法。"),
            ]}
        ]},
        {"name":"第4章 功率与电能的测量","sections":[
            {"title":"一、功率测量","topics":[
                ("直流功率的测量","掌握用电压表电流表法和功率表法测量直流功率的方法。"),
                ("单相交流功率的测量","理解电动系功率表的结构和工作原理，会正确连接功率表测量单相交流功率。"),
                ("三相功率的测量","掌握一瓦特表法和二瓦特表法测量三相功率的原理及适用条件。"),
            ]},
            {"title":"二、电能测量","topics":[
                ("电能的测量","了解感应系电能表的结构和工作原理，掌握电能表的正确接线和读数方法。"),
            ]}
        ]},
        {"name":"第5章 频率与相位的测量","sections":[
            {"title":"一、频率与相位","topics":[
                ("频率的测量","了解频率计和示波器法测量频率的原理和方法。"),
                ("相位的测量","理解相位的含义，会用双踪示波器和相位表测量两个同频率信号的相位差。"),
            ]}
        ]},
        {"name":"第6章 电子测量仪器","sections":[
            {"title":"一、常用仪器","topics":[
                ("示波器的原理与使用","了解示波器的基本结构和工作原理，掌握测量电压、周期、频率和相位的方法。"),
                ("信号发生器","了解信号发生器的功能和输出参数，会正确设置输出信号的波形、频率和幅度。"),
                ("电子电压表与频率计","了解电子电压表和数字频率计的工作原理及使用方法。"),
            ]}
        ]},
        {"name":"第7章 非电量的电气测量","sections":[
            {"title":"一、传感器基础","topics":[
                ("传感器概述","了解传感器的组成和分类，理解静态特性（灵敏度、线性度、迟滞）的含义。"),
            ]},
            {"title":"二、常用传感器","topics":[
                ("温度传感器","了解热电偶和热电阻的测温原理，掌握常用温度传感器的使用和选型。"),
                ("压力与位移传感器","了解电阻应变式、电感式和电容式传感器的原理及其在压力、位移测量中的应用。"),
            ]}
        ]},
        {"name":"第8章 自动检测与数据采集","sections":[
            {"title":"一、检测系统","topics":[
                ("自动检测系统","了解自动检测系统的组成和分类，理解信号调理、A/D转换和数据处理的基本流程。"),
                ("数据采集系统","了解数据采集系统的组成和各部分功能，了解虚拟仪器的基本概念。"),
            ]}
        ]},
    ]
}


def main():
    parser = argparse.ArgumentParser(description="生成内蒙古自治区机电类一课一练考点规划表")
    parser.add_argument("--qt", help="统一题型配置；不填时按是否含计算自动使用10题方案")
    parser.add_argument("--total-questions", type=int, default=10, help="每张试卷总题量，默认10")
    parser.add_argument("--style-mode", choices=["auto", "template", "skip"], default="auto", help="规划表生成后真题风格库准备方式")
    parser.add_argument("--type-config-mode", choices=["auto", "template", "skip"], default="template", help="规划表生成后题型定义JSON准备方式")
    parser.add_argument("--refresh-type-config", action="store_true", help="允许覆盖已存在的题型定义JSON")
    args = parser.parse_args()

    if args.qt:
        try:
            validate_question_plan(args.qt, args.total_questions)
        except ValueError as exc:
            print(f"错误：{exc}")
            sys.exit(1)
    else:
        for label, qtypes in (("有计算默认题型", QTYPES_COMPUTE), ("无计算默认题型", QTYPES_NONCOMP)):
            try:
                validate_question_plan(qtypes, args.total_questions, label=label)
            except ValueError as exc:
                print(f"错误：{exc}")
                sys.exit(1)

    print("正在生成内蒙古机电类考点规划表...\n")
    total = 0
    for book, fname in [
        (DG, "内蒙古机电类_电工技术基础与技能_一课一练考点规划表.xlsx"),
        (DZ, "内蒙古机电类_电工电子技术与技能_一课一练考点规划表.xlsx"),
        (CL, "内蒙古机电类_电气测量技术_一课一练考点规划表.xlsx"),
    ]:
        path, count = create_planning(book, fname, qtypes_override=args.qt, total_questions=args.total_questions)
        print(f"  ✓ {os.path.basename(path)}: {count} 个主题")
        total += count

    asset_qtypes = args.qt or QTYPES_COMPUTE
    prepare_planning_assets(
        "内蒙古自治区机电类",
        asset_qtypes,
        total_questions=args.total_questions,
        style_mode=args.style_mode,
        type_config_mode=args.type_config_mode,
        textbooks=[DG["textbook"], DZ["textbook"], CL["textbook"]],
        refresh_type_config=args.refresh_type_config,
    )

    print(f"\n总计: {total} 个主题")
    print(f"输出目录: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
