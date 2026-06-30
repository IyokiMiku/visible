# -*- coding: utf-8 -*-
"""优化规划表试卷主题。

默认读取当前项目结构下的规划表，也可用 --input 指定文件：

    python 01_工具脚本/规划表/optimize_themes.py --input "04_生成输出/考点规划表/重庆市/机械加工类/重庆市机械加工类_一课一练考点规划表.xlsx"
"""
import argparse
import json
import sys
import time
from pathlib import Path

import openpyxl
from openai import OpenAI

if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = Path(__file__).resolve().parents[2]
CONFIG_PATH = BASE_DIR / "02_配置资源" / "config.json"
DEFAULT_INPUT_XLSX = BASE_DIR / "04_生成输出" / "考点规划表" / "重庆市" / "机械加工类" / "重庆市机械加工类_一课一练考点规划表.xlsx"

SYSTEM_PROMPT = "你是命题规划助手。为每个考点提炼≤8字极简主题名。\n\n规则：\n- 是对知识点的浓缩总结，不是缩写\n- 去掉把握/熟悉/了解/理解/能/会/认识等动词\n- 去掉概念/知识/方法/原理/定义等后缀\n- 像一节课的名称，例如极限与配合、齿轮参数计算\n\n输出纯JSON格式，键名为themes，数组元素含seq和theme。\n\n好例：\n- 认识机器组成及各组成部分的作用 → 机器组成\n- 掌握圆柱齿轮主要参数及几何尺寸的计算 → 齿轮参数计算\n- 了解铁碳合金相图及其应用 → 铁碳合金相图\n- 掌握钢的热处理工艺 → 钢的热处理\n- 了解碳素钢合金钢和铸铁的分类牌号性能和应用 → 钢铁材料分类"

BATCH_SIZE = 25


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def collect_records(ws):
    records = []
    for row_idx in range(6, ws.max_row + 1):
        seq = ws.cell(row_idx, 1).value
        knowledge = ws.cell(row_idx, 2).value
        theme = ws.cell(row_idx, 3).value
        if seq and knowledge and theme:
            themes = str(theme)
            suffix = ""
            if themes.endswith("（一）"):
                suffix = "（一）"
            elif themes.endswith("（二）"):
                suffix = "（二）"
            core = themes.replace("（一）", "").replace("（二）", "")
            records.append({
                "row": row_idx,
                "seq": int(seq),
                "knowledge": str(knowledge),
                "theme_core": core,
                "suffix": suffix,
            })
    return records


def needs_theme_fix(record):
    core = record["theme_core"]
    knowledge = record["knowledge"]
    if len(core) > 8:
        return True
    if len(core) >= 5 and core in knowledge:
        return len(core) > 4
    return False


def optimize_themes(input_xlsx):
    config = load_config()
    client = OpenAI(api_key=config["api_key"], base_url=config["base_url"])

    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active

    records = collect_records(ws)
    print(f"共 {len(records)} 条")

    need_fix = [r for r in records if needs_theme_fix(r)]
    print(f"需优化: {len(need_fix)} 条")

    if not need_fix:
        print("无需优化！")
        return

    updated = 0
    for i in range(0, len(need_fix), BATCH_SIZE):
        batch = need_fix[i:i + BATCH_SIZE]
        batch_num = i // BATCH_SIZE + 1
        total = (len(need_fix) + BATCH_SIZE - 1) // BATCH_SIZE
        print(f"\n第 {batch_num}/{total} 批 ({len(batch)} 条)...")

        items = [{"seq": r["seq"], "知识": r["knowledge"]} for r in batch]

        for attempt in range(3):
            try:
                resp = client.chat.completions.create(
                    model=config["model"],
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": json.dumps(items, ensure_ascii=False)},
                    ],
                    temperature=0.3,
                    max_tokens=4096,
                    response_format={"type": "json_object"},
                )
                content = resp.choices[0].message.content
                data = json.loads(content)

                theme_list = data.get("themes", [])
                if not theme_list:
                    for val in data.values():
                        if isinstance(val, list) and val and isinstance(val[0], dict):
                            theme_list = val
                            break

                if theme_list:
                    for item in theme_list:
                        seq_val = item.get("seq")
                        new_theme = str(item.get("theme", "")).strip()
                        if seq_val and new_theme and len(new_theme) <= 8:
                            for r in batch:
                                if r["seq"] == seq_val:
                                    full = new_theme + r["suffix"]
                                    ws.cell(r["row"], 3, full)
                                    print(f"  #{seq_val}: {full}")
                                    updated += 1
                    break

                keys = list(data.keys()) if isinstance(data, dict) else []
                raise ValueError(f"未找到themes: keys={keys}")
            except (json.JSONDecodeError, ValueError, KeyError) as e:
                print(f"  尝试 {attempt + 1}/3 失败: {e}")
                if attempt < 2:
                    time.sleep(2)
            except Exception as e:
                print(f"  尝试 {attempt + 1}/3 API错误: {e}")
                if attempt < 2:
                    time.sleep(2)
                else:
                    print(f"  !! 第{batch_num}批最终失败")

        time.sleep(0.3)

    wb.save(input_xlsx)
    print(f"\n完成！共更新 {updated} 条 → {input_xlsx}")


def main():
    parser = argparse.ArgumentParser(description="优化规划表试卷主题")
    parser.add_argument("--input", "-i", default=str(DEFAULT_INPUT_XLSX), help="规划表 xlsx 路径")
    args = parser.parse_args()

    input_xlsx = Path(args.input)
    if not input_xlsx.is_absolute():
        input_xlsx = BASE_DIR / input_xlsx
    if not input_xlsx.exists():
        print(f"错误：规划表不存在 {input_xlsx}")
        sys.exit(1)

    optimize_themes(input_xlsx)


if __name__ == "__main__":
    main()
