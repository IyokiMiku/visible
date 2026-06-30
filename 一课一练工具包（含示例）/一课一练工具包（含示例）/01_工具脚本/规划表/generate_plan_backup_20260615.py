"""从考纲 PDF 自动生成考点规划表 xlsx

使用方法：
  python generate_plan.py                           # 交互式选择 PDF
  python generate_plan.py --pdf "考纲.pdf"          # 指定 PDF
  python generate_plan.py --pdf "考纲.pdf" --title "重庆市汽车类"  # 指定标题前缀

生成逻辑：
  1. 读取考纲 PDF，解析出 课程→节→考点 结构
  2. 按关键词（掌握/熟悉/了解）自动判定每个考点的重要性
  3. 每个扫描到的考点单独生成行；极重要考点生成两卷，序号连续递增
  4. 输出带底纹、节名称的 xlsx 规划表
"""

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openai import OpenAI
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from planning_assets import (
    prepare_planning_assets,
    split_province_category as _asset_split_province_category,
    validate_question_plan,
)

BASE_DIR = Path(__file__).resolve().parents[2]
SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "02_配置资源" / "config.json"

# === PDF 文本提取 ===

def extract_pdf_text(pdf_path):
    """提取 PDF 全文文本（优先用 pdfplumber，退化用 PyPDF2）"""
    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except ImportError:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(pdf_path)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        except ImportError:
            print("错误：需要安装 pdfplumber 或 PyPDF2")
            print("  pip install pdfplumber")
            sys.exit(1)
    return text


# === 解析考纲结构 ===

def parse_exam_outline(text):
    """解析考纲文本为结构化数据：课程 → 节 → 考点列表

    返回:
        courses: [{name, sections: [{name, points: [{text, level}]}]}]
        textbooks: [{name, publisher, edition}]
    """
    lines = text.split("\n")
    courses = []
    textbooks = []
    current_course = None
    current_section = None

    # 查找"考试内容及要求"之后的内容
    start_idx = 0
    for i, line in enumerate(lines):
        if "考试内容及要求" in line:
            start_idx = i + 1
            break

    # 查找参考教材
    for i, line in enumerate(lines):
        if "参考教材" in line:
            for j in range(i + 1, min(i + 20, len(lines))):
                tb_line = lines[j].strip()
                # 匹配格式: 1．《机械基础》xxx主编，高等教育出版社，2024年8月第3版。
                m = re.match(
                    r'\d+[．.]\s*《(.+?)》.+?(高等教育出版社|机械工业出版社|人民邮电出版社|中国劳动社会保障出版社|重庆大学出版社|[\u4e00-\u9fa5]+出版社)[，,]\s*(\d{4})\s*年\s*\d+\s*月第\s*(\d+)\s*版',
                    tb_line
                )
                if m:
                    pub_name = m.group(2)
                    pub_short = {"高等教育出版社": "高教", "机械工业出版社": "机工",
                                 "人民邮电出版社": "人邮", "重庆大学出版社": "重大"}.get(pub_name, pub_name[:2])
                    textbooks.append({
                        "name": m.group(1),
                        "publisher": pub_short,
                        "edition": int(m.group(4)),
                    })
            # 优先使用"特别提醒"后的教材列表（通常是更新版本）
            for j in range(i + 1, min(i + 30, len(lines))):
                if "特别提醒" in lines[j] or "2026" in lines[j]:
                    newer_textbooks = []
                    for k in range(j + 1, min(j + 10, len(lines))):
                        tb_line2 = lines[k].strip()
                        m2 = re.match(
                            r'\d+[．.]\s*《(.+?)》.+?(高等教育出版社|机械工业出版社|人民邮电出版社|中国劳动社会保障出版社|重庆大学出版社|[\u4e00-\u9fa5]+出版社)[，,]\s*(\d{4})\s*年\s*\d+\s*月第\s*(\d+)\s*版',
                            tb_line2
                        )
                        if m2:
                            pub_name2 = m2.group(2)
                            pub_short2 = {"高等教育出版社": "高教", "机械工业出版社": "机工",
                                          "人民邮电出版社": "人邮", "重庆大学出版社": "重大"}.get(pub_name2, pub_name2[:2])
                            newer_textbooks.append({
                                "name": m2.group(1),
                                "publisher": pub_short2,
                                "edition": int(m2.group(4)),
                            })
                    if newer_textbooks:
                        textbooks = newer_textbooks
                    break
            break

    # 解析课程/节/考点
    i = start_idx
    while i < len(lines):
        line = lines[i].strip()

        # 跳过页码标记
        if re.match(r'^-+\s*\d+\s*(of|/)\s*\d+\s*-+$', line) or re.match(r'^\d+$', line):
            i += 1
            continue

        # 课程行: "课程一：机械基础" 或 "课程三：机械加工技术"
        m = re.match(r'课程[一二三四五六七八九十\d]+[：:]\s*(.+)', line)
        if m:
            current_course = {"name": m.group(1).strip(), "sections": []}
            courses.append(current_course)
            current_section = None
            i += 1
            continue

        # 节标题: "1．制图基本知识" 或 "绪论" 或 "1. 力系与平衡"
        m = re.match(r'(\d+)[．.、]\s*(.+)', line)
        if m and current_course is not None:
            sec_name = m.group(2).strip()
            # 排除考点行（以"了解/理解/掌握/熟悉/能/会"开头的不是节标题）
            if not re.match(r'(了解|理解|掌握|熟悉|能|会|认识)', sec_name):
                current_section = {"name": f"{m.group(1)}．{sec_name}", "points": []}
                current_course["sections"].append(current_section)
                i += 1
                continue

        # 独立的绪论等
        if line == "绪论" and current_course is not None:
            current_section = {"name": "绪论", "points": []}
            current_course["sections"].append(current_section)
            i += 1
            continue

        # 考点行: "（1）掌握xxx" 或 "(1) 了解xxx"
        m = re.match(r'[（(]\s*(\d+)\s*[）)]\s*(.+)', line)
        if m and current_section is not None:
            point_text = m.group(2).strip()
            # 判断认知层次
            if re.match(r'(掌握|熟练掌握)', point_text):
                level = "掌握"
            elif re.match(r'(熟悉|理解)', point_text):
                level = "理解"
            elif re.match(r'(能|会)', point_text):
                level = "应用"
            else:
                level = "了解"
            current_section["points"].append({"text": point_text, "level": level})
            i += 1
            continue

        # 到"参考教材"或"四、"时停止解析考点
        if "参考教材" in line or re.match(r'四[、.]', line):
            break

        i += 1

    return courses, textbooks


# === 重要性判定 ===

def assess_importance(section):
    """根据考点内容判断节的重要程度

    规则：
      - 含≥2个"掌握"级考点 → 极重要
      - 含≥1个"掌握"级或≥2个"理解/应用"级 → 重要
      - 其余 → 标准
    """
    points = section["points"]
    if not points:
        return "标准"

    master_count = sum(1 for p in points if p["level"] == "掌握")
    understand_count = sum(1 for p in points if p["level"] in ("理解", "应用"))

    if master_count >= 2:
        return "极重要"
    elif master_count >= 1 or understand_count >= 2:
        return "重要"
    else:
        return "标准"


def assess_point_importance(point):
    """根据单个考点的认知层次判断重要程度"""
    if point["level"] == "掌握":
        return "极重要"
    if point["level"] in ("理解", "应用"):
        return "重要"
    return "标准"


def make_theme_from_point(point_text):
    """根据单个考纲知识点生成兜底试卷主题名"""
    theme = point_text.strip()

    # 去掉常见认知层次动词，让主题聚焦在考点本身；“认识”常可直接形成主题，保留。
    theme = re.sub(r'^(熟练掌握|掌握|熟悉|理解|了解|能|会)', '', theme).strip()

    # 主题取考点的核心前半句，避免把多个要求都塞进标题。
    theme = re.split(r'[，,；;。]|及|和|与', theme, maxsplit=1)[0].strip()

    # 标题中常见的“的”多为连接词，去掉后更像主题名。
    theme = theme.replace("的", "")

    return theme or point_text.strip()


def collect_point_records(courses):
    """收集所有考点，供 AI 批量生成主题"""
    records = []
    for course_idx, course in enumerate(courses, 1):
        for sec_idx, section in enumerate(course["sections"]):
            for point_idx, point in enumerate(section["points"], 1):
                records.append({
                    "id": f"{course_idx}-{sec_idx}-{point_idx}",
                    "course": course["name"],
                    "section": section["name"],
                    "point": point["text"],
                })
    return records


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def _extract_json_object(text):
    """从模型输出中提取 JSON 对象文本，兼容被 ```json 包裹的情况。"""
    text = (text or "").strip()
    fence = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.S)
    if fence:
        return fence.group(1)
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        return text[start:end + 1]
    return text


def generate_ai_theme_map(courses, model=None):
    """调用 config.json 中的 OpenAI 兼容 API，根据每个考纲知识点生成简短试卷主题。"""
    records = collect_point_records(courses)
    if not records:
        return {}

    config = load_config()
    model = model or config.get("model")
    client = OpenAI(api_key=config["api_key"], base_url=config["base_url"])
    prompt = {
        "task": "为每个考纲知识点生成适合作为一课一练试卷标题的简短中文主题。",
        "requirements": [
            "必须逐条处理输入中的每一个考点，不能遗漏、合并或新增。",
            "主题应体现考纲知识点本身，而不是直接使用节名。",
            "主题要短，一般 4 到 10 个汉字；必要时可稍长。",
            "去掉掌握、熟悉、了解、理解、能、会等认知层次词。",
            "不要使用（一）（二）（1）（2）等卷次或序号后缀。",
            "不要输出‘试卷’‘练习’‘专题’等泛化词。",
            "如果考点包含多个并列要求，提炼最核心、最适合命题的主题。",
        ],
        "examples": [
            {"point": "认识机器的组成及各组成部分的作用", "theme": "认识机器组成"},
            {"point": "掌握平面图形尺寸标注的方法", "theme": "平面图形尺寸标注"},
            {"point": "了解常用金属材料的性能", "theme": "金属材料性能"},
        ],
        "output_format": {
            "themes": [
                {"id": "输入 items 中的 id", "theme": "生成的主题"}
            ]
        },
        "items": records,
    }

    system_prompt = (
        "你是职业教育考试命题规划助手，擅长把考纲知识点提炼成简短、准确、适合试卷主题栏使用的中文标题。"
        "请只返回一个合法 JSON 对象，不要添加 Markdown、解释或多余文字。"
    )
    user_prompt = json.dumps(prompt, ensure_ascii=False)

    for attempt in range(3):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                max_tokens=config.get("max_tokens", 8000),
                temperature=0.2,
            )
            content = response.choices[0].message.content
            data = json.loads(_extract_json_object(content))
            break
        except json.JSONDecodeError:
            print("错误：AI 返回内容不是有效 JSON")
            sys.exit(1)
        except Exception as e:
            print(f"错误：AI 主题生成调用失败 (第{attempt + 1}次): {e}")
            if attempt == 2:
                sys.exit(1)
            time.sleep((attempt + 1) * 8)

    valid_ids = {r["id"] for r in records}
    theme_map = {}
    for item in data.get("themes", []):
        item_id = item.get("id")
        theme = str(item.get("theme", "")).strip()
        if item_id in valid_ids and theme:
            theme_map[item_id] = theme

    missing = valid_ids - set(theme_map)
    if missing:
        print(f"警告：AI 未返回 {len(missing)} 个考点主题，将使用规则兜底生成。")

    return theme_map


# === 生成规划主题 ===

def generate_topics(courses, default_qt="单选5+填空3+综合2", default_diff="80:10:10", theme_map=None):
    """将解析的课程结构转化为规划表的主题行列表。

    每个扫描到的考点单独生成行；极重要考点生成两卷，序号连续递增。
    """
    topics = []
    seq = 1

    for course_idx, course in enumerate(courses, 1):
        course_header = f"课程{'一二三四五六七八九十'[course_idx - 1]}：{course['name']}"
        topics.append({"type": "course", "text": course_header})

        for sec_idx, section in enumerate(course["sections"]):
            # 节标题行
            topics.append({"type": "section", "text": section["name"]})

            if not section["points"]:
                continue

            for point_idx, point in enumerate(section["points"], 1):
                importance = assess_point_importance(point)
                exam_ref = f"课程{course_idx}§{sec_idx}({point_idx})"
                point_id = f"{course_idx}-{sec_idx}-{point_idx}"
                theme_base = (theme_map or {}).get(point_id) or make_theme_from_point(point["text"])
                copy_suffixes = ["一", "二"] if importance == "极重要" else [None]

                for copy_suffix in copy_suffixes:
                    if copy_suffix:
                        theme = f"{theme_base}（{copy_suffix}）"
                    else:
                        theme = theme_base

                    topics.append({
                        "type": "topic",
                        "seq": seq,
                        "knowledge": point["text"],
                        "theme": theme,
                        "level": importance,
                        "question_types": default_qt,
                        "difficulty": default_diff,
                        "sets": 1,
                        "exam_ref": exam_ref,
                    })
                    seq += 1

    return topics


# === 教材目录驱动规划表 ===

_PUBLISHER_SHORT = {
    "高等教育出版社": "高教", "高教版": "高教", "高教": "高教",
    "机械工业出版社": "机工", "机工版": "机工", "机工": "机工",
    "人民邮电出版社": "人邮", "人邮版": "人邮", "人邮": "人邮",
    "重庆大学出版社": "重大", "重大版": "重大", "重大": "重大",
}


def parse_pages_list(pages_str, default_end=None):
    """解析页码范围，如 1,3,5-8；不传且给定 default_end 时默认前 default_end 页。"""
    if not pages_str:
        if default_end is None:
            return []
        return list(range(1, default_end + 1))
    pages = []
    for part in str(pages_str).split(","):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^(\d+)\s*[-~]\s*(\d+)$", part)
        if m:
            start, end = int(m.group(1)), int(m.group(2))
            pages.extend(range(min(start, end), max(start, end) + 1))
        else:
            try:
                pages.append(int(part))
            except ValueError:
                pass
    return sorted(set(p for p in pages if p > 0))


def prompt_toc_pages_for_textbook(pdf_path):
    """交互询问单本教材的目录页范围。"""
    while True:
        print(f"请输入《{Path(pdf_path).stem}》目录所在页码范围（如 1-3 或 2,4-6）：")
        pages_input = input("> ").strip()
        pages = parse_pages_list(pages_input)
        if pages:
            return pages
        print("页码范围不能为空，且需包含有效正整数页码。")


def parse_textbook_filename(path):
    """从教材 PDF 文件名中提取书名、出版社简称、版次。"""
    stem = Path(path).stem
    name = stem
    publisher = ""
    edition = ""

    m = re.search(r"(.+?)[（(]([^）)]*?)(高教|机工|人邮|重大|[^）)]{2,8}出版社)?(?:版)?[·•\s_-]*第\s*(\d+)\s*版[^）)]*[）)]", stem)
    if m:
        name = m.group(1).strip()
        raw_pub = (m.group(2) or m.group(3) or "").strip(" ·•_-版")
        publisher = _PUBLISHER_SHORT.get(raw_pub, raw_pub[:2] if raw_pub else "")
        edition = m.group(4)
    else:
        m = re.search(r"(.+?)[（(]([^）)]+)[）)]", stem)
        if m:
            name = m.group(1).strip()
            inside = m.group(2)
            pub_m = re.search(r"(高教|机工|人邮|重大|[一-龥]+出版社)", inside)
            ed_m = re.search(r"第\s*(\d+)\s*版", inside)
            if pub_m:
                raw_pub = pub_m.group(1)
                publisher = _PUBLISHER_SHORT.get(raw_pub, raw_pub[:2])
            if ed_m:
                edition = ed_m.group(1)

    name = re.sub(r"[（(].*$", "", name).strip()
    return {"name": name or stem, "publisher": publisher or "待填", "edition": int(edition) if str(edition).isdigit() else "待填", "path": str(path)}


def _textbook_info_line(textbook):
    edition = textbook.get("edition", "待填")
    if isinstance(edition, int):
        edition_text = f"第{edition}版"
    elif str(edition).isdigit():
        edition_text = f"第{edition}版"
    else:
        edition_text = "第待填版"
    return f"参考教材：《{textbook.get('name', '待填写')}》{textbook.get('publisher', '待填')}{edition_text}"


def _clean_toc_title(text):
    text = re.sub(r"\.{2,}\s*\d+\s*$", "", text)
    text = re.sub(r"[·•]{2,}\s*\d+\s*$", "", text)
    text = re.sub(r"\s+\d+\s*$", "", text)
    text = re.sub(r"^(第[一二三四五六七八九十百\d]+[章节篇项目]|项目[一二三四五六七八九十\d]+|任务[一二三四五六七八九十\d]+|\d+(?:[．.]\d+)*|[一二三四五六七八九十]+[、.．])\s*", "", text)
    text = re.sub(r"^(学习任务|任务|项目|知识点|单元|模块)\s*", "", text)
    text = re.sub(r"\s+", " ", text).strip(" ：:.-—_\t")
    return text


def make_theme_from_toc(title):
    theme = _clean_toc_title(title)
    theme = re.sub(r"(的)?(概念|知识|方法|原理|定义|概述|基础|简介)$", "", theme)
    theme = theme.replace(" ", "")
    return theme[:12] or title.strip()[:12]


def _parse_toc_line(line):
    raw = line.strip()
    if not raw or raw in ("目录", "CONTENTS", "Contents"):
        return None
    raw = re.sub(r"\s+", " ", raw)
    raw = re.sub(r"[.·•…]{2,}", " ", raw)
    m_page = re.search(r"(?:\s|^)(\d{1,4})\s*$", raw)
    page = int(m_page.group(1)) if m_page else None
    text = re.sub(r"\s+\d{1,4}\s*$", "", raw).strip()

    chapter_patterns = [
        r"^(第[一二三四五六七八九十百\d]+[章篇])\s+(.+)$",
        r"^(项目[一二三四五六七八九十\d]+)\s+(.+)$",
        r"^(模块[一二三四五六七八九十\d]+)\s+(.+)$",
    ]
    for pat in chapter_patterns:
        m = re.match(pat, text)
        if m:
            return {"kind": "chapter", "title": f"{m.group(1)} {m.group(2).strip()}", "page": page}

    section_patterns = [
        r"^(\d+[．.]\d+(?:[．.]\d+)*)\s*(.+)$",
        r"^(任务[一二三四五六七八九十\d]+)\s+(.+)$",
        r"^([一二三四五六七八九十]+[、.．])\s*(.+)$",
        r"^(\d+[、.．])\s*(.+)$",
    ]
    for pat in section_patterns:
        m = re.match(pat, text)
        if m and len(m.group(2).strip()) >= 2:
            return {"kind": "section", "title": f"{m.group(1)} {m.group(2).strip()}", "page": page}
    return None


def parse_toc_text(toc_text):
    """将教材目录 OCR 文本解析为章/节/主题列表。"""
    items = []
    current_chapter = ""
    seen = set()
    for line in toc_text.splitlines():
        parsed = _parse_toc_line(line)
        if not parsed:
            continue
        title = parsed["title"]
        if parsed["kind"] == "chapter":
            current_chapter = title
            continue
        key = (current_chapter, title)
        if key in seen:
            continue
        seen.add(key)
        theme = make_theme_from_toc(title)
        if len(theme) < 2:
            continue
        items.append({
            "id": f"toc-{len(items) + 1}",
            "chapter": current_chapter or "教材目录",
            "section": title,
            "theme": theme,
            "page": parsed.get("page"),
        })
    return items


def ocr_textbook_toc(pdf_path, pages, output_dir, reuse=True, engine="auto", tessdata_dir=None,
                     dpi=2.5, layout="auto", preprocess=False, keep_images=False):
    """OCR 教材目录页，优先使用目录专用扫描器，缓存 toc_raw/toc_structured 并返回文本。"""
    output_dir = Path(output_dir)
    os.makedirs(output_dir, exist_ok=True)
    raw_path = output_dir / "toc_raw.txt"
    structured_path = output_dir / "toc_structured.json"
    if reuse and structured_path.exists() and raw_path.exists() and raw_path.read_text(encoding="utf-8", errors="ignore").strip():
        return raw_path.read_text(encoding="utf-8", errors="ignore")
    if reuse and raw_path.exists() and raw_path.read_text(encoding="utf-8", errors="ignore").strip():
        return raw_path.read_text(encoding="utf-8", errors="ignore")

    engine = (engine or "auto").lower()
    if engine in ("auto", "tesseract"):
        try:
            from textbook_toc_scanner import scan_textbook_toc_structured
            print(f"  使用 Tesseract 目录专用 OCR: {Path(pdf_path).name} 页码 {pages}")
            result = scan_textbook_toc_structured(
                pdf_path,
                pages=pages,
                tessdata_dir=tessdata_dir,
                dpi=dpi,
                output_dir=output_dir,
                reuse=False,
                keep_images=keep_images,
                layout=layout,
                preprocess=preprocess,
                verbose=False,
            )
            text = raw_path.read_text(encoding="utf-8", errors="ignore") if raw_path.exists() else "\n\n".join(result.get("page_texts", {}).values())
            raw_path.write_text(text, encoding="utf-8")
            return text
        except Exception as e:
            if engine == "tesseract":
                print(f"错误：Tesseract 目录 OCR 失败：{e}")
                sys.exit(1)
            print(f"  Tesseract 目录 OCR 不可用，回退 RapidOCR：{e}")

    ocr_dir = BASE_DIR / "01_工具脚本" / "OCR"
    if str(ocr_dir) not in sys.path:
        sys.path.insert(0, str(ocr_dir))
    try:
        from ocr_pdf import export_pdf_pages, run_ocr
    except ImportError:
        print("错误：无法导入 OCR/ocr_pdf.py")
        sys.exit(1)

    print(f"  OCR 教材目录页: {Path(pdf_path).name} 页码 {pages}")
    try:
        image_entries = export_pdf_pages(pdf_path, output_dir, pages=pages, zoom=dpi)
        run_ocr(image_entries, output_dir, min_score=0.0)
    except SystemExit:
        raise
    except Exception as e:
        print(f"错误：教材目录 OCR 失败：{e}")
        sys.exit(1)

    combined = output_dir / "combined.txt"
    text = combined.read_text(encoding="utf-8", errors="ignore") if combined.exists() else ""
    raw_path.write_text(text, encoding="utf-8")
    return text


def convert_scanner_entries_to_toc_items(entries):
    """将 textbook_toc_scanner 的结构化 entries 转为 generate_plan 的 toc_items。"""
    items = []
    current_chapter = "教材目录"
    seen = set()
    for entry in entries or []:
        level = int(entry.get("level") or 0)
        title = str(entry.get("title") or "").strip()
        if not title:
            continue
        if level == 1:
            current_chapter = title
            continue
        key = (current_chapter, title)
        if key in seen:
            continue
        seen.add(key)
        theme = make_theme_from_toc(title)
        if len(theme) < 2:
            continue
        items.append({
            "id": f"toc-{len(items) + 1}",
            "chapter": current_chapter,
            "section": title,
            "theme": theme,
            "page": entry.get("book_page"),
            "source_page": entry.get("source_page"),
            "raw_line": entry.get("raw_line", ""),
        })
    return items


def load_structured_toc_items(ocr_dir):
    """读取增强 OCR 产出的 toc_structured.json，并转为 toc_items。"""
    structured_path = Path(ocr_dir) / "toc_structured.json"
    if not structured_path.exists():
        return []
    try:
        data = json.loads(structured_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    entries = data.get("entries", [])
    return convert_scanner_entries_to_toc_items(entries)


def flatten_outline_points(courses, course_filter=None):
    """把考纲课程结构展开为可匹配的考点列表。"""
    records = []
    for course_idx, course in enumerate(courses, 1):
        if course_filter and course_filter not in course["name"] and course["name"] not in course_filter:
            continue
        for sec_idx, section in enumerate(course["sections"], 1):
            for point_idx, point in enumerate(section["points"], 1):
                records.append({
                    "id": f"{course_idx}-{sec_idx}-{point_idx}",
                    "course": course["name"],
                    "section": section["name"],
                    "text": point["text"],
                    "level": point["level"],
                    "exam_ref": f"课程{course_idx}§{sec_idx}({point_idx})",
                })
    return records


def _keyword_set(text):
    text = re.sub(r"^(熟练掌握|掌握|熟悉|理解|了解|能|会|认识)", "", text or "")
    text = re.sub(r"[，,。；;：:？！?、（）()《》<>\[\]【】\s]", "", text)
    stop = {"掌握", "熟悉", "理解", "了解", "认识", "应用", "方法", "概念", "作用", "特点", "分类", "要求", "进行", "常用"}
    units = {text[i:i + 2] for i in range(max(0, len(text) - 1))}
    units |= {text[i:i + 3] for i in range(max(0, len(text) - 2))}
    return {u for u in units if len(u) >= 2 and u not in stop}


def local_match_toc_to_outline(toc_items, outline_points):
    """本地关键词粗匹配：为每个目录项找最接近的考纲知识点。"""
    matches = {}
    for item in toc_items:
        item_text = item.get("chapter", "") + item.get("section", "") + item.get("theme", "")
        item_units = _keyword_set(item_text)
        scored = []
        for point in outline_points:
            point_units = _keyword_set(point["course"] + point["section"] + point["text"])
            shorter = min(len(item_units), len(point_units)) or 1
            score = len(item_units & point_units) / shorter
            if item.get("theme") and item["theme"] in point["text"]:
                score += 0.35
            cleaned_section = _clean_toc_title(point["section"])
            if cleaned_section and (cleaned_section in item.get("section", "") or cleaned_section in item.get("chapter", "")):
                score += 0.2
            if point.get("course") and (point["course"] in item.get("chapter", "") or item.get("chapter", "") in point["course"]):
                score += 0.1
            if score > 0:
                scored.append((score, point))
        scored.sort(key=lambda x: x[0], reverse=True)
        selected = [p for score, p in scored[:3] if score >= 0.28]
        matches[item["id"]] = selected
    return matches


def build_local_match_candidates(toc_items, outline_points, limit=5):
    """生成本地匹配候选和分数，用于校对报告。"""
    report = {}
    for item in toc_items:
        item_text = item.get("chapter", "") + item.get("section", "") + item.get("theme", "")
        item_units = _keyword_set(item_text)
        scored = []
        for point in outline_points:
            point_units = _keyword_set(point["course"] + point["section"] + point["text"])
            shorter = min(len(item_units), len(point_units)) or 1
            score = len(item_units & point_units) / shorter
            if item.get("theme") and item["theme"] in point["text"]:
                score += 0.35
            cleaned_section = _clean_toc_title(point["section"])
            if cleaned_section and (cleaned_section in item.get("section", "") or cleaned_section in item.get("chapter", "")):
                score += 0.2
            if score > 0:
                scored.append({"score": round(score, 4), "point": point})
        scored.sort(key=lambda x: x["score"], reverse=True)
        report[item["id"]] = scored[:limit]
    return report


def ai_match_toc_to_outline(toc_items, outline_points, model=None):
    """调用 config.json 中的 API，把教材目录项匹配到考纲知识点。"""
    if not toc_items or not outline_points:
        return {}
    config = load_config()
    model = model or config.get("model")
    client = OpenAI(api_key=config["api_key"], base_url=config["base_url"])
    prompt = {
        "task": "把教材目录条目匹配到最相关的考纲知识点。只允许选择给定的 point id，不得新增知识点。",
        "rules": [
            "每个目录条目可匹配0到3个考纲知识点。",
            "只有语义明确相关时才匹配；不确定时返回空数组。",
            "confidence 范围0-1，低于0.55视为待人工确认。",
        ],
        "toc_items": [{"id": i["id"], "chapter": i["chapter"], "section": i["section"], "theme": i["theme"]} for i in toc_items],
        "outline_points": [{"id": p["id"], "course": p["course"], "section": p["section"], "point": p["text"], "level": p["level"]} for p in outline_points],
        "output_format": {"matches": [{"toc_id": "toc-1", "matched_point_ids": ["1-1-1"], "confidence": 0.8, "reason": "简述理由"}]},
    }
    system_prompt = "你是职业教育教材目录与考试大纲对齐助手。请只返回合法 JSON 对象，不要输出 Markdown。"
    for attempt in range(3):
        try:
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": json.dumps(prompt, ensure_ascii=False)},
                ],
                max_tokens=config.get("max_tokens", 8000),
                temperature=0.1,
            )
            data = json.loads(_extract_json_object(response.choices[0].message.content))
            break
        except Exception as e:
            print(f"  AI 匹配失败 (第{attempt + 1}次): {e}")
            if attempt == 2:
                return {}
            time.sleep((attempt + 1) * 8)

    point_by_id = {p["id"]: p for p in outline_points}
    result = {}
    for item in data.get("matches", []):
        if float(item.get("confidence", 0) or 0) < 0.55:
            continue
        matched = [point_by_id[pid] for pid in item.get("matched_point_ids", []) if pid in point_by_id]
        if matched:
            result[item.get("toc_id")] = matched[:3]
    return result


def _merge_matches(local_matches, ai_matches):
    merged = dict(local_matches or {})
    for toc_id, points in (ai_matches or {}).items():
        if points:
            merged[toc_id] = points
    return merged


def write_toc_match_report(ocr_dir, toc_items, matches, local_candidates):
    """输出目录项与考纲匹配校对报告。"""
    report = []
    for item in toc_items:
        matched = matches.get(item["id"], [])
        candidates = local_candidates.get(item["id"], []) if local_candidates else []
        report.append({
            "toc_id": item["id"],
            "chapter": item.get("chapter", ""),
            "section": item.get("section", ""),
            "theme": item.get("theme", ""),
            "source_page": item.get("source_page"),
            "book_page": item.get("page"),
            "matched": [
                {"exam_ref": p.get("exam_ref"), "course": p.get("course"), "section": p.get("section"), "text": p.get("text"), "level": p.get("level")}
                for p in matched
            ],
            "local_candidates": [
                {"score": c.get("score"), "exam_ref": c.get("point", {}).get("exam_ref"), "text": c.get("point", {}).get("text")}
                for c in candidates
            ],
            "status": "matched" if matched else "待人工确认",
        })
    ocr_dir = Path(ocr_dir)
    ocr_dir.mkdir(parents=True, exist_ok=True)
    (ocr_dir / "toc_match_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    md = ["# 教材目录与考纲匹配报告\n"]
    for row in report:
        md.append(f"## {row['toc_id']} {row['chapter']} / {row['section']}")
        md.append(f"- 主题：{row['theme']}")
        md.append(f"- 来源页：PDF {row.get('source_page') or '未知'} / 书页 {row.get('book_page') or '未知'}")
        if row["matched"]:
            md.append("- 匹配考纲：")
            for p in row["matched"]:
                md.append(f"  - {p.get('exam_ref')}：{p.get('text')}")
        else:
            md.append("- 匹配考纲：待人工确认")
        md.append("")
    (ocr_dir / "toc_match_report.md").write_text("\n".join(md), encoding="utf-8")


def _highest_level(points):
    order = {"掌握": 3, "理解": 2, "应用": 2, "了解": 1}
    if not points:
        return "标准"
    best = max(points, key=lambda p: order.get(p.get("level"), 0))
    return assess_point_importance(best)


def generate_topics_from_textbook_toc(toc_items, matches, textbook, default_qt="单选5+填空3+综合2", default_diff="80:10:10"):
    """按教材目录生成规划表 topics，B列写匹配到的考纲知识点。"""
    topics = []
    seq = 1
    current_chapter = None
    topics.append({"type": "course", "text": f"教材：《{textbook.get('name', '未命名教材')}》"})
    for item in toc_items:
        if item["chapter"] != current_chapter:
            current_chapter = item["chapter"]
            topics.append({"type": "section", "text": current_chapter})
        points = matches.get(item["id"], [])
        knowledge = "；".join(p["text"] for p in points) if points else f"待补充：教材目录主题“{item['theme']}”未在考纲中精确匹配"
        exam_ref = "；".join(p["exam_ref"] for p in points) if points else "待人工确认"
        level = _highest_level(points)
        copy_suffixes = ["一", "二"] if level == "极重要" else [None]
        for copy_suffix in copy_suffixes:
            theme = f"{item['theme']}（{copy_suffix}）" if copy_suffix else item["theme"]
            topics.append({
                "type": "topic",
                "seq": seq,
                "knowledge": knowledge,
                "theme": theme,
                "level": level,
                "question_types": default_qt,
                "difficulty": default_diff,
                "sets": 1,
                "exam_ref": exam_ref,
            })
            seq += 1
    return topics


def _resolve_textbook_pdfs(args, province, category):
    if args.textbook_pdf:
        return [Path(args.textbook_pdf)]
    candidates = []
    if args.textbook_dir:
        candidates.append(Path(args.textbook_dir))
    if province and category:
        # 双击交互时按标题前缀优先查找：03_项目数据/参考资料/教材/省份/考类
        candidates.append(BASE_DIR / "03_项目数据" / "参考资料" / "教材" / province / category)
        candidates.append(BASE_DIR / "03_项目数据" / "教材" / province / category)
    for directory in candidates:
        if directory.exists():
            pdfs = sorted(p for p in directory.glob("*.pdf") if not p.name.startswith("~"))
            if pdfs:
                return pdfs
    return []


def _warn_missing_textbook_pdfs(expected_textbooks, pdfs):
    """提示考纲参考教材与实际教材 PDF 数量/名称不一致。"""
    if not expected_textbooks:
        return

    found = [parse_textbook_filename(p) for p in pdfs]
    found_names = [tb.get("name", "") for tb in found]
    missing = []
    for expected in expected_textbooks:
        expected_name = expected.get("name", "")
        if not expected_name:
            continue
        matched = any(
            expected_name in found_name or found_name in expected_name
            for found_name in found_names
            if found_name
        )
        if not matched:
            missing.append(expected)

    if len(pdfs) != len(expected_textbooks) or missing:
        print("\n警告：考试说明中的参考教材与当前教材 PDF 不完全一致。")
        print(f"  考试说明参考教材：{len(expected_textbooks)} 本；当前找到教材 PDF：{len(pdfs)} 本")
        if found:
            print("  已找到教材 PDF：")
            for tb in found:
                print(f"    - 《{tb.get('name', '未知教材')}》{tb.get('publisher', '待填')}第{tb.get('edition', '待填')}版")
        if missing:
            print("  可能缺失教材 PDF：")
            for tb in missing:
                print(f"    - 《{tb.get('name', '未知教材')}》{tb.get('publisher', '待填')}第{tb.get('edition', '待填')}版")
        print("  如需生成三本书对应的三个规划表，请把缺失教材 PDF 放入 --textbook-dir 指定目录。")


def _match_course_name(textbook, courses):
    name = textbook.get("name", "")
    for course in courses:
        if name and (name in course["name"] or course["name"] in name):
            return course["name"]
    return ""


def _default_ocr_dir(args, province, category, textbook):
    if args.ocr_output_dir:
        return Path(args.ocr_output_dir)
    return BASE_DIR / "03_项目数据" / "参考资料" / "教材OCR" / _safe_path_part(province or "未分类") / _safe_path_part(category or "未分类") / _safe_path_part(textbook.get("name", "教材"))


def run_textbook_driven_plan(args, courses, textbooks, title_prefix):
    """教材目录驱动模式主流程。"""
    province, category = _split_province_category(title_prefix)
    pdfs = _resolve_textbook_pdfs(args, province, category)
    if not pdfs:
        print("错误：未找到教材 PDF。请指定 --textbook-pdf 或 --textbook-dir。")
        sys.exit(1)
    _warn_missing_textbook_pdfs(textbooks, pdfs)

    generated = []
    shared_pages = parse_pages_list(args.toc_pages)
    for pdf_path in pdfs:
        if not pdf_path.exists():
            print(f"错误：教材文件不存在 {pdf_path}")
            sys.exit(1)
        textbook = parse_textbook_filename(pdf_path)
        course_name = _match_course_name(textbook, courses)
        outline_points = flatten_outline_points(courses, course_filter=course_name)
        if not outline_points:
            outline_points = flatten_outline_points(courses)

        ocr_dir = _default_ocr_dir(args, province, category, textbook)
        pages = shared_pages or prompt_toc_pages_for_textbook(pdf_path)
        toc_text = ocr_textbook_toc(
            str(pdf_path),
            pages,
            ocr_dir,
            reuse=args.reuse_ocr,
            engine=args.ocr_engine,
            tessdata_dir=args.tessdata,
            dpi=args.toc_dpi,
            layout=args.toc_layout,
            preprocess=args.toc_preprocess,
            keep_images=args.keep_toc_images,
        )
        toc_items = load_structured_toc_items(ocr_dir) or parse_toc_text(toc_text)
        if toc_items:
            (Path(ocr_dir) / "toc_items_for_plan.json").write_text(json.dumps(toc_items, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"  解析目录条目：{len(toc_items)} 条")
        if not toc_items:
            print("错误：未能从教材目录 OCR 文本解析出目录条目，请检查 toc_raw.txt 或调整 --toc-pages。")
            sys.exit(1)

        local_candidates = build_local_match_candidates(toc_items, outline_points)
        local_matches = local_match_toc_to_outline(toc_items, outline_points)
        ai_matches = {} if args.no_ai_match else ai_match_toc_to_outline(toc_items, outline_points, args.ai_model)
        matches = _merge_matches(local_matches, ai_matches)
        write_toc_match_report(ocr_dir, toc_items, matches, local_candidates)
        topics = generate_topics_from_textbook_toc(toc_items, matches, textbook, args.qt, args.diff)

        title = f"{title_prefix}《一课一练》考点规划表 v1"
        config_line = f"题型：{args.qt} | 难度：{args.diff}"
        info_line = _textbook_info_line(textbook)

        if args.output and len(pdfs) == 1:
            output_path = args.output
        else:
            base_dir = BASE_DIR / "04_生成输出" / "考点规划表" / _safe_path_part(province or "未分类") / _safe_path_part(category or "未分类")
            os.makedirs(base_dir, exist_ok=True)
            safe_name = _safe_path_part(f"{title_prefix}_{textbook['name']}_一课一练考点规划表")
            output_path = str(base_dir / f"{safe_name}.xlsx")
            if os.path.exists(output_path):
                i = 2
                while True:
                    candidate = str(base_dir / f"{safe_name}_v{i}.xlsx")
                    if not os.path.exists(candidate):
                        output_path = candidate
                        break
                    i += 1

        data_rows = write_planning_xlsx(output_path, title, config_line, info_line, topics, [textbook])
        prepare_planning_assets(
            title_prefix,
            args.qt,
            total_questions=args.total_questions,
            style_mode=args.style_mode,
            type_config_mode=args.type_config_mode,
            textbooks=[textbook],
            refresh_type_config=args.refresh_type_config,
        )
        generated.append(output_path)
        matched_count = sum(1 for item in toc_items if matches.get(item["id"]))
        print(f"✓ 已生成目录驱动规划表: {output_path}")
        print(f"  目录条目 {len(toc_items)} 条，匹配考纲 {matched_count} 条，写入 {data_rows} 行")

    print("\n目录驱动规划表生成完成：")
    for path in generated:
        print(f"  - {path}")


# === 生成 xlsx ===

# 样式常量
FILL_HEADER = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
FILL_COURSE = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
FILL_SECTION = PatternFill(start_color="00000000", end_color="00000000", fill_type=None)
FILL_NORMAL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
FILL_IMPORTANT = PatternFill(start_color="FFFFD7D7", end_color="FFFFD7D7", fill_type="solid")

FONT_HEADER = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
FONT_COURSE = Font(name="微软雅黑", size=10, bold=True)
FONT_SECTION = Font(name="微软雅黑", size=10, bold=True)
FONT_NORMAL = Font(name="微软雅黑", size=9)
FONT_TITLE = Font(name="微软雅黑", size=12, bold=True)

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _strip_knowledge_trailing_period(text):
    """B列考纲知识点不在句尾追加句号；只去掉末尾句号，保留原文其他内容。"""
    return re.sub(r"[。．.]+\s*$", "", str(text or "").strip())


def write_planning_xlsx(output_path, title, config_line, info_line, topics, textbooks):
    """生成带格式的规划表 xlsx"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考点规划表"

    # 列宽设置
    col_widths = [6, 60, 18, 8, 28, 10, 6, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # === 表头区域 ===
    # Row 1: 标题
    ws.merge_cells("A1:H1")
    c = ws.cell(1, 1, title)
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: 配置行
    ws.merge_cells("A2:H2")
    ws.cell(2, 1, config_line).font = Font(name="微软雅黑", size=9)

    # Row 3: 信息行
    ws.merge_cells("A3:H3")
    ws.cell(3, 1, info_line).font = Font(name="微软雅黑", size=9)

    # Row 4: 空行
    # Row 5: 表头
    headers = ["序号", "考纲知识点", "试卷主题", "级别", "题型", "难度", "套数", "考纲标号"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(5, col, h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    # === 数据区域 ===
    row = 6
    for item in topics:
        if item["type"] == "course":
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row, 1, item["text"])
            c.font = FONT_COURSE
            c.fill = FILL_COURSE
            c.alignment = Alignment(vertical="center")
            row += 1

        elif item["type"] == "section":
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row, 1, f"  {item['text']}")
            c.font = FONT_SECTION
            c.alignment = Alignment(vertical="center")
            row += 1

        elif item["type"] == "topic":
            fill = FILL_IMPORTANT if item["level"] == "极重要" else FILL_NORMAL
            values = [
                item["seq"],
                _strip_knowledge_trailing_period(item["knowledge"]),
                item["theme"],
                item["level"],
                item["question_types"],
                item["difficulty"],
                item["sets"],
                item["exam_ref"],
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row, col, val)
                c.font = FONT_NORMAL
                c.fill = fill
                c.border = THIN_BORDER
                if col in (1, 4, 6, 7):
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(vertical="center", wrap_text=True)
            row += 1

    # 保存
    wb.save(output_path)
    return row - 6  # 数据行数


def _split_province_category(title_prefix):
    """从标题前缀中解析省份和考类，如“重庆市机械加工类”，兼容自治区简称/全称。"""
    province, category = _asset_split_province_category(title_prefix)
    if province and category:
        return province, category
    text = (title_prefix or "").strip()
    m = re.search(r"([一-龥]+(?:省|市))([一-龥]+类)", text)
    if m:
        return m.group(1), m.group(2)
    return "", ""


def _safe_path_part(text):
    """清理路径非法字符。"""
    text = (text or "").strip()
    return re.sub(r'[\\/:*?"<>|\s]+', "_", text).strip("_") or "未分类"


# === 主流程 ===

def main():
    parser = argparse.ArgumentParser(description="从考纲 PDF 生成考点规划表")
    parser.add_argument("--pdf", "-p", help="考纲 PDF 文件路径")
    parser.add_argument("--title", "-t", help="规划表标题前缀（如'重庆市汽车类'）")
    parser.add_argument("--qt", default="单选5+填空3+综合2", help="默认题型配置")
    parser.add_argument("--total-questions", type=int, default=10, help="每张试卷总题量，默认10；题型配置数量合计必须等于该值")
    parser.add_argument("--diff", default="80:10:10", help="默认难度配比")
    parser.add_argument("--style-mode", choices=["auto", "template", "skip"], default="auto", help="规划表生成后真题风格库准备方式：auto调用API蒸馏，template只生成模板，skip跳过")
    parser.add_argument("--type-config-mode", choices=["auto", "template", "skip"], default="template", help="规划表生成后题型定义JSON准备方式：auto调用API总结，template生成模板，skip跳过")
    parser.add_argument("--refresh-type-config", action="store_true", help="允许覆盖已存在的题型定义JSON；默认只生成建议版不覆盖")
    parser.add_argument("--output", "-o", help="输出 xlsx 路径")
    parser.add_argument("--no-ai-theme", action="store_true", help="不调用 AI，使用规则兜底生成试卷主题")
    parser.add_argument("--ai-model", help="用于生成试卷主题的模型；默认使用 config.json 中的 model")
    parser.add_argument("--textbook-driven", action="store_true", help="启用教材目录 OCR 驱动模式：用教材目录命名试卷主题，再匹配考纲知识点")
    parser.add_argument("--textbook-dir", help="教材 PDF 所在目录；不填时按 教材/省份/考类 或 03_项目数据/参考资料/教材/省份/考类 自动查找")
    parser.add_argument("--textbook-pdf", help="只处理指定教材 PDF")
    parser.add_argument("--toc-pages", help="教材目录页范围，如 1-3 或 3,5-8；不填时每本教材单独询问")
    parser.add_argument("--ocr-output-dir", help="教材目录 OCR 缓存目录；默认 03_项目数据/参考资料/教材OCR/省份/考类/教材名")
    parser.add_argument("--reuse-ocr", action="store_true", default=True, help="复用已有 toc_raw.txt（默认启用）")
    parser.add_argument("--no-reuse-ocr", dest="reuse_ocr", action="store_false", help="强制重新 OCR 教材目录页")
    parser.add_argument("--no-ai-match", action="store_true", help="不调用 API 匹配目录与考纲，只使用本地关键词匹配")
    parser.add_argument("--ocr-engine", choices=["auto", "tesseract", "rapidocr"], default="auto", help="教材目录 OCR 引擎，默认 auto：优先 Tesseract 目录扫描器，失败回退 RapidOCR")
    parser.add_argument("--tessdata", help="Tesseract tessdata 目录（含 chi_sim.traineddata）")
    parser.add_argument("--toc-dpi", type=float, default=2.5, help="教材目录页渲染倍率，默认 2.5")
    parser.add_argument("--toc-layout", choices=["auto", "single", "double"], default="auto", help="教材目录版面：auto/single/double")
    parser.add_argument("--toc-preprocess", action="store_true", help="OCR 前对教材目录图片做灰度、对比度和二值化增强")
    parser.add_argument("--keep-toc-images", action="store_true", help="保留教材目录页渲染图片，便于人工校对")
    args = parser.parse_args()

    try:
        validate_question_plan(args.qt, args.total_questions)
    except ValueError as exc:
        print(f"错误：{exc}")
        sys.exit(1)

    # 选择 PDF
    pdf_path = args.pdf
    if not pdf_path:
        # 交互式选择
        print("请输入考纲 PDF 文件路径：")
        pdf_path = input("> ").strip().strip('"')

    if not os.path.exists(pdf_path):
        print(f"错误：找不到文件 {pdf_path}")
        sys.exit(1)

    # 标题
    title_prefix = args.title
    if not title_prefix:
        print("请输入规划表标题前缀（如'重庆市机械加工类'）：")
        title_prefix = input("> ").strip()

    # 双击运行时询问是否扫描教材目录。命令行已显式指定 --textbook-driven/--textbook-dir/--textbook-pdf 时不重复询问。
    if not args.textbook_driven and not args.textbook_dir and not args.textbook_pdf:
        print("是否需要扫描教材目录 PDF？输入 y/是 扫描教材目录；直接回车则只按考纲生成规划表：")
        answer = input("> ").strip().lower()
        if answer in ("y", "yes", "是", "需要", "扫描", "1"):
            args.textbook_driven = True
            province, category = _split_province_category(title_prefix)
            if province and category:
                textbook_dir = BASE_DIR / "03_项目数据" / "参考资料" / "教材" / province / category
                print(f"将按前缀查找教材 PDF：{textbook_dir}")
            else:
                print("警告：未能从标题前缀解析出省份和考类，请使用如'重庆市机械加工类'的格式。")

    print(f"\n正在解析考纲: {Path(pdf_path).name}")
    print("=" * 60)

    # 提取文本
    text = extract_pdf_text(pdf_path)
    if not text.strip():
        print("错误：PDF 文本为空，请确认文件可读")
        sys.exit(1)

    # 解析结构
    courses, textbooks = parse_exam_outline(text)
    print(f"解析完成：{len(courses)} 个课程")
    for c in courses:
        sec_count = len(c["sections"])
        point_count = sum(len(s["points"]) for s in c["sections"])
        print(f"  {c['name']}: {sec_count} 节, {point_count} 个考点")

    if textbooks:
        print(f"\n参考教材：")
        for tb in textbooks:
            print(f"  《{tb['name']}》{tb['publisher']}第{tb['edition']}版")

    if args.textbook_driven:
        run_textbook_driven_plan(args, courses, textbooks, title_prefix)
        return

    # 生成主题
    theme_map = None
    if args.no_ai_theme:
        print("\n已关闭 AI 主题生成，使用规则兜底生成试卷主题。")
    else:
        print(f"\n正在调用 Claude API 生成试卷主题：{args.ai_model}")
        theme_map = generate_ai_theme_map(courses, args.ai_model)
        print(f"AI 主题生成完成：{len(theme_map)} 个考点")

    topics = generate_topics(courses, args.qt, args.diff, theme_map)
    topic_count = sum(1 for t in topics if t["type"] == "topic")
    important_count = sum(1 for t in topics if t.get("level") == "极重要")
    print(f"\n共生成 {topic_count} 练（其中极重要拆分为两练的有 {important_count} 个）")

    # 构建表头信息
    title = f"{title_prefix}《一课一练》考点规划表 v1"
    config_line = f"题型：{args.qt} | 难度：{args.diff}"
    tb_str = "、".join(f"《{tb['name']}》{tb['publisher']}第{tb['edition']}版" for tb in textbooks)
    info_line = f"参考教材：{tb_str}" if tb_str else "参考教材：待填写"

    # 输出路径
    output_path = args.output
    if not output_path:
        root_dir = BASE_DIR / "04_生成输出" / "考点规划表"
        province, category = _split_province_category(title_prefix)
        if province and category:
            base_dir = root_dir / _safe_path_part(province) / _safe_path_part(category)
        else:
            base_dir = root_dir / "未分类"
        os.makedirs(base_dir, exist_ok=True)
        safe_name = _safe_path_part(title_prefix.replace("（", "(").replace("）", ")"))
        output_path = str(base_dir / f"{safe_name}_一课一练考点规划表.xlsx")
        # 避免覆盖已有文件
        if os.path.exists(output_path):
            i = 2
            while True:
                candidate = str(base_dir / f"{safe_name}_一课一练考点规划表_v{i}.xlsx")
                if not os.path.exists(candidate):
                    output_path = candidate
                    break
                i += 1

    # 生成 xlsx
    data_rows = write_planning_xlsx(output_path, title, config_line, info_line, topics, textbooks)
    print(f"\n✓ 规划表已生成: {output_path}")
    print(f"  共 {data_rows} 行数据")

    prepare_planning_assets(
        title_prefix,
        args.qt,
        total_questions=args.total_questions,
        style_mode=args.style_mode,
        type_config_mode=args.type_config_mode,
        textbooks=textbooks,
        refresh_type_config=args.refresh_type_config,
    )

    # 打印重要性统计
    levels = {}
    for t in topics:
        if t["type"] == "topic":
            lv = t["level"]
            levels[lv] = levels.get(lv, 0) + 1
    print(f"\n重要性分布：")
    for lv in ("极重要", "重要", "标准"):
        if lv in levels:
            print(f"  {lv}: {levels[lv]} 练")


if __name__ == "__main__":
    main()
