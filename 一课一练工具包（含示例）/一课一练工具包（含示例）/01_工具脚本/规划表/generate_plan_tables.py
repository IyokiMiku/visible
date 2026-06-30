# -*- coding: utf-8 -*-
"""生成重庆市机械加工类考点规划表"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
import re
from pathlib import Path

# ========== 数据定义 ==========
courses_data = [
    {
        'name': '机械基础',
        'publisher': '高教',
        'edition': '3',
        'course_num': 1,
        'sections': [
            {
                'title': '绪论',
                'section_num': 0,
                'points': [
                    ('认识机器的组成及各组成部分的作用', '机器组成', '标准'),
                    ('掌握机械、机器、机构、构件、零件的区别和联系', '机械层级关系', '极重要'),
                    ('能识别机械的类型', '机械类型识别', '重要'),
                ]
            },
            {
                'title': '1．力系与平衡',
                'section_num': 1,
                'points': [
                    ('理解力的概念及基本性质', '力的概念', '重要'),
                    ('了解物体的受力分析方法', '受力分析', '标准'),
                ]
            },
            {
                'title': '2．强度与刚度',
                'section_num': 2,
                'points': [
                    ('理解直杆基本变形的特点', '直杆变形', '重要'),
                    ('了解低碳钢拉伸变形的过程', '低碳钢拉伸', '标准'),
                    ('了解提高构件强度、刚度和稳定性的措施', '构件性能提升', '标准'),
                ]
            },
            {
                'title': '3．材料与选用',
                'section_num': 3,
                'points': [
                    ('了解金属材料的性能', '金属性能', '标准'),
                    ('了解碳素钢、合金钢和铸铁的分类、牌号、性能和应用', '钢铁材料分类', '标准'),
                    ('了解钢的热处理目的、分类和应用', '热处理概述', '标准'),
                    ('了解有色金属材料和非金属材料在工程中的应用', '有色金属应用', '标准'),
                    ('掌握钢的热处理工艺', '热处理工艺', '极重要'),
                    ('了解零件的失效形式', '零件失效', '标准'),
                    ('掌握选用材料的基本原则和方法', '材料选用原则', '极重要'),
                ]
            },
            {
                'title': '4．误差与公差',
                'section_num': 4,
                'points': [
                    ('理解互换性的概念', '互换性', '重要'),
                    ('掌握尺寸精度有关的专业术语', '尺寸精度术语', '极重要'),
                    ('认识机械零件的几何公差项目及符号', '几何公差', '标准'),
                    ('能判断孔和轴配合的三种类型', '配合类型判断', '重要'),
                    ('会计算零件的尺寸、偏差、公差及配合精度', '尺寸公差计算', '重要'),
                ]
            },
            {
                'title': '5．连接与紧固',
                'section_num': 5,
                'points': [
                    ('了解机械连接的类型', '机械连接类型', '标准'),
                    ('了解键、销连接的类型、特点和应用', '键销连接', '标准'),
                    ('熟悉螺纹及螺纹连接的类型、特点和应用', '螺纹连接', '重要'),
                    ('了解螺纹连接的拧紧与防松', '螺纹防松', '标准'),
                    ('理解联轴器、离合器的功用、类型、特点和应用', '联轴离合器', '重要'),
                ]
            },
            {
                'title': '6．机构与运动',
                'section_num': 6,
                'points': [
                    ('能识别常见运动形式的表示符号', '运动符号', '重要'),
                    ('熟悉平面四杆机构的基本类型、特点和应用，能判定铰链四杆机构的类型', '四杆机构', '重要'),
                    ('认识凸轮机构的组成、分类和应用', '凸轮机构', '标准'),
                    ('认识其他机构的特点和应用', '其他机构', '标准'),
                ]
            },
            {
                'title': '7．传动与维护',
                'section_num': 7,
                'points': [
                    ('能识别带传动的类型和特点', '带传动类型', '重要'),
                    ('熟悉普通V带的型号', 'V带型号', '重要'),
                    ('了解带传动的失效形式，会进行带传动的安装与维护', '带传动维护', '标准'),
                    ('会进行带传动的相关计算', '带传动计算', '重要'),
                    ('了解链传动的特点、类型和应用', '链传动', '标准'),
                    ('熟悉齿轮传动的类型、特点和应用', '齿轮传动', '重要'),
                    ('掌握标准直齿圆柱齿轮各部分尺寸及传动比的计算', '齿轮计算', '极重要'),
                    ('了解齿轮的加工方法及失效形式', '齿轮加工失效', '标准'),
                    ('了解蜗杆传动的特点、类型和应用', '蜗杆传动', '标准'),
                    ('能进行定轴轮系传动比的计算和旋转零件运动方向的判别', '轮系计算', '重要'),
                ]
            },
            {
                'title': '8．轴与轴承',
                'section_num': 8,
                'points': [
                    ('了解轴的结构、分类、材料和应用', '轴的结构', '标准'),
                    ('了解滑动轴承和滚动轴承的结构、分类及常见的失效形式', '轴承分类', '标准'),
                ]
            },
            {
                'title': '9．润滑与密封',
                'section_num': 9,
                'points': [
                    ('了解机械中常见的摩擦与磨损的类型', '摩擦磨损', '标准'),
                    ('了解润滑的目的及状态', '润滑目的', '标准'),
                    ('了解机械润滑剂的种类、性能及选用', '润滑剂选用', '标准'),
                    ('了解机械常用的润滑方法', '润滑方法', '标准'),
                    ('熟悉润滑管理的"五定"', '润滑五定', '重要'),
                    ('了解常用密封装置的分类、特点和应用', '密封装置', '标准'),
                    ('了解机械环保与安全防护常识', '环保安全', '标准'),
                ]
            },
            {
                'title': '10．液压与气动',
                'section_num': 10,
                'points': [
                    ('了解液压传动的组成及特点', '液压传动', '标准'),
                    ('了解气压传动的组成及特点', '气压传动', '标准'),
                ]
            },
        ]
    },
    {
        'name': '机械制图',
        'publisher': '高教',
        'edition': '5',
        'course_num': 2,
        'sections': [
            {
                'title': '1．制图基本知识与技能',
                'section_num': 1,
                'points': [
                    ('熟悉图纸幅面及格式的规定', '图纸幅面', '重要'),
                    ('掌握比例的规定及应用', '比例应用', '极重要'),
                    ('了解字体的规定及应用', '字体规定', '标准'),
                    ('掌握常用图线的主要用途和画法', '图线画法', '极重要'),
                    ('能应用标注尺寸的基本规则进行常用尺寸的标注与识读', '尺寸标注', '重要'),
                    ('掌握线段和圆的等分方法', '等分方法', '极重要'),
                    ('熟悉斜度和锥度的画法及其在图样上的标注', '斜度锥度', '重要'),
                    ('掌握圆弧连接的作图原理和方法', '圆弧连接', '极重要'),
                    ('掌握简单平面图形的尺寸和线段分析方法及其作图方法', '平面图形', '极重要'),
                ]
            },
            {
                'title': '2．正投影与三视图',
                'section_num': 2,
                'points': [
                    ('理解投影法的概念、分类及三视图的形成，熟悉三视图的关系和投影规律', '投影三视图', '重要'),
                    ('掌握点、直线、面的投影特性和规律', '点线面投影', '极重要'),
                    ('熟练掌握基本几何体的绘制、尺寸标注及其表面上点的投影作图', '几何体投影', '极重要'),
                ]
            },
            {
                'title': '3．轴测图',
                'section_num': 3,
                'points': [
                    ('了解简单组合体正等轴测图和斜二轴测图的作图方法', '轴测图', '标准'),
                ]
            },
            {
                'title': '4．组合体视图',
                'section_num': 4,
                'points': [
                    ('掌握组合体的组合形式及分析方法', '组合体分析', '极重要'),
                    ('掌握组合体表面常见的截交线画法', '截交线画法', '极重要'),
                    ('掌握相贯线的画法及其近似画法', '相贯线画法', '极重要'),
                    ('掌握组合体视图的画法及尺寸标注', '组合体视图', '极重要'),
                    ('理解组合体的识读方法', '组合体识读', '重要'),
                ]
            },
            {
                'title': '5．图样表示法',
                'section_num': 5,
                'points': [
                    ('了解基本视图的形成、名称及配置关系', '基本视图', '标准'),
                    ('掌握向视图、局部视图、斜视图的画法和标注', '辅助视图', '极重要'),
                    ('掌握全剖视图、半剖视图、局部剖视图的画法和标注', '剖视图', '极重要'),
                    ('掌握移出断面图和重合断面图的画法和标注', '断面图', '极重要'),
                    ('熟悉局部放大图和常用简化画法', '简化画法', '重要'),
                ]
            },
            {
                'title': '6．常用标准件及齿轮和弹簧表示法',
                'section_num': 6,
                'points': [
                    ('掌握螺纹的画法及标注规定', '螺纹画法', '极重要'),
                    ('掌握螺纹连接的画法', '螺纹连接', '极重要'),
                    ('掌握键连接和销连接的画法和标记', '键销画法', '极重要'),
                    ('熟练掌握标准直齿圆柱齿轮基本尺寸计算及画法', '齿轮画法', '极重要'),
                    ('了解圆柱螺旋压缩弹簧各个部分名称和尺寸关系', '弹簧表示', '标准'),
                    ('熟悉常用滚动轴承的类型、代号及简化画法和标注', '轴承画法', '重要'),
                ]
            },
            {
                'title': '7．零件图',
                'section_num': 7,
                'points': [
                    ('掌握零件的视图表达方法', '零件视图', '极重要'),
                    ('掌握表面粗糙度的概念、符号及标注', '表面粗糙度', '极重要'),
                    ('掌握零件的尺寸、公差标注及技术要求', '零件标注', '极重要'),
                    ('会识读零件图', '零件图识读', '重要'),
                ]
            },
            {
                'title': '8．装配图',
                'section_num': 8,
                'points': [
                    ('了解装配图中机器或部件的结构', '装配图结构', '标准'),
                    ('了解装配图中所标注的尺寸及技术要求', '装配图标注', '标准'),
                    ('会识读简单的装配图', '装配图识读', '重要'),
                ]
            },
        ]
    },
    {
        'name': '机械加工技术',
        'publisher': '机工',
        'edition': '2',
        'course_num': 3,
        'sections': [
            {
                'title': '1．机械加工的概念',
                'section_num': 1,
                'points': [
                    ('理解机械产品生产过程、机械加工工艺过程及组成、生产纲领和生产类型', '加工过程', '重要'),
                    ('理解基准的概念及分类', '基准分类', '重要'),
                    ('了解机械加工的劳动生产率', '劳动生产率', '标准'),
                ]
            },
            {
                'title': '2．金属切削基础知识',
                'section_num': 2,
                'points': [
                    ('掌握切削运动和切削要素，会计算切削用量三要素', '切削要素', '极重要'),
                    ('掌握车刀的组成及几何角度', '车刀角度', '极重要'),
                    ('了解切削过程及其物理现象', '切削物理', '标准'),
                    ('理解切削液的作用、种类、选用和加注方法', '切削液', '重要'),
                ]
            },
            {
                'title': '3．机械加工工艺系统',
                'section_num': 3,
                'points': [
                    ('了解切削机床的分类与型号', '机床分类', '标准'),
                    ('掌握车削加工的设备特点、工艺范围和工艺特点', '车削加工', '极重要'),
                    ('理解铣削加工的设备特点、工艺范围和工艺特点', '铣削加工', '重要'),
                    ('理解钻削加工的设备特点、工艺范围和工艺特点', '钻削加工', '重要'),
                    ('了解镗削加工的设备特点、工艺范围和工艺特点', '镗削加工', '标准'),
                    ('了解磨削加工的设备特点、工艺范围和工艺特点', '磨削加工', '标准'),
                    ('了解刨削加工的设备特点、工艺范围和工艺特点', '刨削加工', '标准'),
                    ('了解常用刀具的种类和用途', '刀具种类', '标准'),
                    ('掌握金属切削刀具的材料', '刀具材料', '极重要'),
                    ('了解刀具的磨损及寿命', '刀具磨损', '标准'),
                    ('掌握车刀刃磨的方法', '车刀刃磨', '极重要'),
                    ('了解机床夹具的分类、组成和作用', '夹具组成', '标准'),
                    ('理解工件的定位原理、会分析限制自由度与加工要求的关系', '工件定位', '重要'),
                    ('理解常见定位方式与定位元件', '定位元件', '重要'),
                    ('理解工件在夹具中的夹紧', '工件夹紧', '重要'),
                    ('了解常用的夹紧机构', '夹紧机构', '标准'),
                    ('理解机械加工精度', '加工精度', '重要'),
                    ('了解影响加工精度的主要因素', '精度因素', '标准'),
                ]
            },
            {
                'title': '4．机械加工工艺规程',
                'section_num': 4,
                'points': [
                    ('理解机械加工工艺规程的内容及格式', '工艺规程', '重要'),
                    ('会进行零件的结构工艺性分析', '结构工艺性', '重要'),
                    ('能合理选择零件毛坯', '毛坯选择', '重要'),
                    ('会选择定位基准', '基准选择', '重要'),
                    ('理解拟定工艺路线', '工艺路线', '重要'),
                    ('掌握加工余量和工序尺寸的确定', '加工余量', '极重要'),
                    ('理解尺寸链的概念、会进行工艺尺寸链的计算', '尺寸链计算', '重要'),
                    ('了解机床及工艺装备的选择', '装备选择', '标准'),
                ]
            },
            {
                'title': '5．典型零件的加工',
                'section_num': 5,
                'points': [
                    ('了解轴类零件的结构特点、功用、技术要求及工艺分析', '轴类零件', '标准'),
                    ('掌握轴类零件的车削加工方法', '轴车削', '极重要'),
                    ('掌握轴类零件的磨削加工方法', '轴磨削', '极重要'),
                    ('了解轴类零件的装夹方式', '轴装夹', '标准'),
                    ('了解套类零件的结构特点、功用、技术要求及工艺分析', '套类零件', '标准'),
                    ('掌握套类零件的孔加工方法', '套孔加工', '极重要'),
                    ('了解套类零件孔的精密加工方法', '套精密加工', '标准'),
                    ('了解套类零件的装夹方式', '套装夹', '标准'),
                    ('了解箱体类零件的结构特点、功用、技术要求及工艺分析', '箱体零件', '标准'),
                    ('掌握箱体类零件的平面加工方法', '箱体平面', '极重要'),
                    ('了解箱体类零件的孔系加工方法', '箱体孔系', '标准'),
                    ('了解箱体类零件的装夹方式', '箱体装夹', '标准'),
                    ('了解圆柱齿轮的结构特点、功用、精度要求及工艺分析', '齿轮零件', '标准'),
                    ('掌握圆柱齿轮的齿形切削加工方法', '齿轮齿形', '极重要'),
                    ('了解齿轮的装夹方式', '齿轮装夹', '标准'),
                ]
            },
            {
                'title': '6．机械装配工艺基础',
                'section_num': 6,
                'points': [
                    ('了解装配的概念', '装配概念', '标准'),
                    ('了解装配工作的基本内容', '装配内容', '标准'),
                    ('了解装配的组织形式', '装配组织', '标准'),
                    ('了解装配精度', '装配精度', '标准'),
                    ('理解装配尺寸链，会进行装配尺寸链的计算', '装配尺寸链', '重要'),
                    ('了解装配方法及选择', '装配方法', '标准'),
                    ('了解典型零部件的装配', '典型装配', '标准'),
                ]
            },
            {
                'title': '7．设备的维护',
                'section_num': 7,
                'points': [
                    ('了解设备的三级保养制', '三级保养', '标准'),
                    ('了解设备的维修方式和修理类别', '维修方式', '标准'),
                ]
            },
            {
                'title': '8．先进加工技术简介',
                'section_num': 8,
                'points': [
                    ('了解精密与超精密加工的概念', '精密加工', '标准'),
                    ('了解机械零件常见的特种加工方法', '特种加工', '标准'),
                ]
            },
        ]
    },
]

# ========== 题型和难度配置 ==========
question_type = '单选5+填空3+综合2'
difficulty = '80:10:10'

# ========== 样式定义 ==========
font_title = Font(name='微软雅黑', size=12, bold=True)
font_config = Font(name='微软雅黑', size=9)
font_header = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
font_course = Font(name='微软雅黑', size=10, bold=True)
font_section = Font(name='微软雅黑', size=10, bold=True)
font_data = Font(name='微软雅黑', size=9)

fill_header = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
fill_course = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
fill_important = PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid')
fill_normal = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

col_widths = {'A': 6, 'B': 60, 'C': 18, 'D': 8, 'E': 28, 'F': 10, 'G': 6, 'H': 16}


def create_plan_table(course_info, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '考点规划表'

    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # 第1行：标题
    ws.merge_cells('A1:H1')
    ws['A1'] = '重庆市机械加工类《一课一练》考点规划表 v1'
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center
    ws.row_dimensions[1].height = 25

    # 第2行：配置
    ws.merge_cells('A2:H2')
    ws['A2'] = f'题型：{question_type} | 难度：{difficulty}'
    ws['A2'].font = font_config
    ws['A2'].alignment = align_center
    ws.row_dimensions[2].height = 18

    # 第3行：教材
    ws.merge_cells('A3:H3')
    ws['A3'] = f'参考教材：《{course_info["name"]}》{course_info["publisher"]}第{course_info["edition"]}版'
    ws['A3'].font = font_config
    ws['A3'].alignment = align_center
    ws.row_dimensions[3].height = 18

    # 第4行：空行
    ws.row_dimensions[4].height = 8

    # 第5行：列标题
    headers = ['序号', '考纲知识点', '试卷主题', '级别', '题型', '难度', '套数', '考纲标号']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[5].height = 22

    # 数据行
    current_row = 6
    seq = 1

    # 课程行
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    cell = ws.cell(row=current_row, column=1, value=f'课程{course_info["course_num"]}：{course_info["name"]}')
    cell.font = font_course
    cell.fill = fill_course
    cell.alignment = align_center
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    for section in course_info['sections']:
        # 节标题行
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        cell = ws.cell(row=current_row, column=1, value=f'  {section["title"]}')
        cell.font = font_section
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for point_idx, (point_text, theme, level) in enumerate(section['points'], 1):
            if level == '极重要':
                for suffix in ['（一）', '（二）']:
                    ws.cell(row=current_row, column=1, value=seq).alignment = align_center
                    ws.cell(row=current_row, column=2, value=re.sub(r"[。．.]+\s*$", "", str(point_text).strip())).alignment = align_left_top
                    ws.cell(row=current_row, column=3, value=theme + suffix).alignment = align_left
                    ws.cell(row=current_row, column=4, value=level).alignment = align_center
                    ws.cell(row=current_row, column=5, value=question_type).alignment = align_left
                    ws.cell(row=current_row, column=6, value=difficulty).alignment = align_center
                    ws.cell(row=current_row, column=7, value=1).alignment = align_center
                    label = f'课程{course_info["course_num"]}§{section["section_num"]}({point_idx})'
                    ws.cell(row=current_row, column=8, value=label).alignment = align_left

                    for col in range(1, 9):
                        c = ws.cell(row=current_row, column=col)
                        c.font = font_data
                        c.fill = fill_important
                        c.border = thin_border

                    ws.row_dimensions[current_row].height = 32
                    current_row += 1
                    seq += 1
            else:
                ws.cell(row=current_row, column=1, value=seq).alignment = align_center
                ws.cell(row=current_row, column=2, value=point_text).alignment = align_left_top
                ws.cell(row=current_row, column=3, value=theme).alignment = align_left
                ws.cell(row=current_row, column=4, value=level).alignment = align_center
                ws.cell(row=current_row, column=5, value=question_type).alignment = align_left
                ws.cell(row=current_row, column=6, value=difficulty).alignment = align_center
                ws.cell(row=current_row, column=7, value=1).alignment = align_center
                label = f'课程{course_info["course_num"]}§{section["section_num"]}({point_idx})'
                ws.cell(row=current_row, column=8, value=label).alignment = align_left

                for col in range(1, 9):
                    c = ws.cell(row=current_row, column=col)
                    c.font = font_data
                    c.fill = fill_normal
                    c.border = thin_border

                ws.row_dimensions[current_row].height = 32
                current_row += 1
                seq += 1

    wb.save(output_path)
    print(f'已保存：{output_path}，共 {seq - 1} 练')


def main():
    base_dir = Path(__file__).resolve().parents[2]
    output_dir = base_dir / '04_生成输出' / '考点规划表' / '重庆市' / '机械加工类'
    os.makedirs(output_dir, exist_ok=True)

    for course in courses_data:
        filename = f'重庆市机械加工类_{course["name"]}_一课一练考点规划表.xlsx'
        output_path = output_dir / filename
        # 如果存在则重命名
        if output_path.exists():
            output_path = output_path.with_name(output_path.stem + '_v2' + output_path.suffix)
        create_plan_table(course, str(output_path))

    print('\n全部生成完成！')


if __name__ == '__main__':
    main()
