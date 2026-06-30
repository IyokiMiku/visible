"""规划表解析与输出路径辅助。"""
import re
from pathlib import Path

import openpyxl

_DIGIT_TO_CN = {"1": "一", "2": "二", "3": "三", "4": "四", "5": "五",
                "6": "六", "7": "七", "8": "八", "9": "九", "10": "十"}

_CN_TO_DIGIT = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5,
                "六": 6, "七": 7, "八": 8, "九": 9, "十": 10,
                "十一": 11, "十二": 12, "十三": 13, "十四": 14, "十五": 15}

_AUTONOMOUS_REGION_MAP = {
    "内蒙古": "内蒙古自治区",
    "新疆": "新疆维吾尔自治区",
    "西藏": "西藏自治区",
    "广西": "广西壮族自治区",
    "宁夏": "宁夏回族自治区",
}


def _province_name_variants(province: str) -> list[str]:
    """返回省份名称的全称/简称候选，用于从标题中剥离省份部分。"""
    name = str(province or "").strip()
    normalized = _AUTONOMOUS_REGION_MAP.get(name, name)
    variants = []
    for item in (normalized, name):
        if item and item not in variants:
            variants.append(item)
    for short_name, full_name in _AUTONOMOUS_REGION_MAP.items():
        if normalized == full_name and short_name not in variants:
            variants.append(short_name)
    return sorted(variants, key=len, reverse=True)


def _parse_category_from_title(title: str, province: str) -> str:
    """从规划表标题中解析考类。

    先剥离已识别省份，避免“内蒙古自治区机电类”被解析成“自治区机电类”。
    """
    text = str(title or "").strip()
    for province_name in _province_name_variants(province):
        if province_name in text:
            text = text.replace(province_name, "", 1)
            break
    cat_match = re.search(r"([一-龥]+类)", text)
    return cat_match.group(1) if cat_match else ""


def _normalize_province_name(province: str) -> str:
    """规范化省级行政区名称，自治区统一使用全称。"""
    name = str(province or "").strip()
    return _AUTONOMOUS_REGION_MAP.get(name, name)

def _parse_edition(raw_edition: str) -> tuple:
    """将原始版次字符串解析为 (出版社简称, 中文版次)。

    示例：
        '高教第3版' → ('高教版', '第三版')
        '机工第2版' → ('机工版', '第二版')
        '第1版'     → ('高教版', '第一版')  # 默认出版社
        '高教版·第三版' → ('高教版', '第三版')
    """
    raw_edition = str(raw_edition or "").strip().strip("（）()")
    raw_edition = raw_edition.replace(" ", "").replace("·", "")
    pub_match = re.match(r"([一-龥]+?)(?:版)?第([\d一二三四五六七八九十]+)版", raw_edition)
    if pub_match:
        pub_short = pub_match.group(1)
        ver_num = pub_match.group(2)
    else:
        ver_match = re.search(r"第([\d一二三四五六七八九十]+)版", raw_edition)
        pub_short = "高教"
        ver_num = ver_match.group(1) if ver_match else "1"

    if not pub_short.endswith("版"):
        pub_short += "版"
    edition_cn = f"第{_DIGIT_TO_CN.get(str(ver_num), ver_num)}版"
    return pub_short, edition_cn


def _extract_textbook_details(text: str) -> list[dict]:
    """从规划表头部文本中提取参考教材信息。

    命名必须以考纲/规划表第 3 行的参考教材为准，即使本地没有对应教材 PDF。
    兼容：
      - 参考教材：《机械基础》高教第3版
      - 参考教材：《机械基础》（高教版·第三版）
      - 参考教材：《机械基础》（高教版·第3版）
    """
    text = str(text or "")
    details = []
    pattern = re.compile(
        r"《(.+?)》\s*(?:[（(]\s*)?"
        r"([一-龥]{1,12}(?:版)?(?:[·\s]*)第[\d一二三四五六七八九十]+版)"
        r"(?:\s*[）)])?"
    )
    for name, raw_edition in pattern.findall(text):
        publisher, edition_cn = _parse_edition(raw_edition)
        detail = {
            "name": name.strip(),
            "publisher": publisher,
            "edition": edition_cn,
            "display": f"{publisher}·{edition_cn}",
        }
        if detail not in details:
            details.append(detail)
    return details

def parse_planning_table(xlsx_path):
    """解析考点规划表，返回元数据和主题列表"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # 提取头部元数据
    meta = {
        "title": str(rows[0][0] or "").strip(),
        "config_line": str(rows[1][0] or "").strip(),
        "textbooks": str(rows[2][0] or "").strip(),
    }

    # 解析题型配置
    config_line = meta["config_line"]
    qt_match = re.search(r"题型[：:]\s*(.+?)\s*[|｜]", config_line)
    diff_match = re.search(r"难度[：:]\s*([\d:]+)", config_line)
    meta["question_types_str"] = qt_match.group(1) if qt_match else ""
    meta["difficulty_str"] = diff_match.group(1) if diff_match else "80:10:10"

    # 解析教材（必须以规划表/考纲参考教材为命名依据，即使本地没有对应教材 PDF）
    tb_text = meta["textbooks"]
    meta["textbook_details"] = _extract_textbook_details(tb_text)
    meta["textbook_list"] = [(d["name"], f"{d['publisher']}{d['edition']}") for d in meta["textbook_details"]]

    # 也尝试从 row1 中查找教材信息（有些表把教材放在配置行）
    if not meta["textbook_details"]:
        for check_row in rows[:5]:
            check_text = str(check_row[0] or "")
            found_details = _extract_textbook_details(check_text)
            if found_details:
                meta["textbook_details"] = found_details
                meta["textbook_list"] = [(d["name"], f"{d['publisher']}{d['edition']}") for d in found_details]
                break

    # 解析省份/自治区。标题里写“内蒙古”“新疆”“西藏”等简称时，文件名和文档标题统一规范为自治区全称。
    province_match = re.search(
        r"(内蒙古自治区|新疆维吾尔自治区|西藏自治区|广西壮族自治区|宁夏回族自治区|[一-龥]+(?:省|市|自治区)|内蒙古|新疆|西藏|广西|宁夏)",
        meta["title"],
    )
    meta["province"] = _normalize_province_name(province_match.group(1)) if province_match else "重庆市"

    # 解析类别。必须先剥离完整省份名/简称，避免自治区标题中“自治区”被误并入考类。
    meta["category"] = _parse_category_from_title(meta["title"], meta["province"])

    # 找到表头行
    header_idx = None
    for i, row in enumerate(rows):
        if row[0] == "序号":
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("找不到表头行（'序号'列）")

    # 解析主题
    topics = []
    current_course = ""
    current_unit = ""
    current_unit_num = None
    current_unit_count = 0
    current_section = ""
    current_section_count = 0
    current_theme_count = 0

    for row in rows[header_idx + 1:]:
        seq = row[0]
        knowledge = row[1]
        theme = row[2]
        level = row[3]
        q_types = row[4]
        difficulty = row[5]
        sets = row[6]
        exam_ref = row[7]

        # 课程行
        if isinstance(seq, str) and "课程" in seq:
            current_course = seq.strip()
            continue
        # 三级模板的单元行：序号列为“绪论”“第1章 xxx”等，知识点列为空。
        # 普通文本（如“二极管基础知识”）是当前单元下的节标题，而不是新单元。
        # 注意：两级模板中的“课程三：xxx”已在上方处理，不作为单元标题。
        if isinstance(seq, str) and seq.strip() and not knowledge:
            text = seq.strip()
            unit_match = re.match(r"第\s*([\d一二三四五六七八九十百]+)\s*[章节单元]\s*(.*)", text)
            if unit_match or text == "绪论":
                current_unit = text
                current_unit_count += 1
                if unit_match:
                    raw_num = unit_match.group(1)
                    current_unit_num = int(raw_num) if raw_num.isdigit() else _CN_TO_DIGIT.get(raw_num, current_unit_count)
                else:
                    current_unit_num = current_unit_count
                current_section = ""
                current_section_count = 0
                current_theme_count = 0
            else:
                current_section_count += 1
                current_section = f"{current_section_count}．{text}"
                current_theme_count = 0
            continue
        # 章节行 — 格式1: 序号列为空，知识点列含"一、xxx"
        if (seq is None or seq == "") and knowledge and knowledge.strip().startswith(("一", "二", "三", "四", "五", "六", "七", "八", "九", "十", "（", "(")):
            current_section = knowledge.strip()
            current_section_count += 1
            current_theme_count = 0
            continue
        # 章节行 — 格式2: 序号列含"(一) xxx"或"（一）xxx"
        if isinstance(seq, str) and re.match(r'\s*[（(][一二三四五六七八九十百]+[）)]\s*\S', seq):
            current_section = seq.strip()
            current_section_count += 1
            current_theme_count = 0
            continue
        # 章节行 — 格式3: 序号列为缩进文本（如"  绪论""  1．力系与平衡"），知识点列为空
        if isinstance(seq, str) and seq.startswith(" ") and not knowledge:
            current_section_count += 1
            current_section = f"{current_section_count}．{seq.strip()}"
            current_theme_count = 0
            continue
        # 空行或汇总行
        if seq is None or not isinstance(seq, (int, float)):
            continue

        sets_count = int(sets) if sets else 1
        current_theme_count += 1
        topics.append({
            "seq": int(seq),
            "knowledge": str(knowledge or "").strip(),
            "theme": str(theme or "").strip(),
            "level": str(level or "标准").strip(),
            "question_types": str(q_types or "").strip(),
            "difficulty": str(difficulty or "80:10:10").strip(),
            "sets": sets_count,
            "exam_ref": str(exam_ref or "").strip(),
            "course": current_course,
            "unit": current_unit,
            "unit_num": current_unit_num,
            "section": current_section or ("1．" if current_unit else ""),
            "theme_num": current_theme_count,
        })

    # 构建课程→教材映射（按出现顺序，课程一对应教材一，依次类推）
    course_names = []
    for row in rows[header_idx + 1:]:
        val = row[0]
        if isinstance(val, str) and "课程" in val:
            course_names.append(val.strip())
    meta["course_textbook_map"] = {}
    if meta["textbook_details"] and course_names:
        for idx, course in enumerate(course_names):
            if idx < len(meta["textbook_details"]):
                meta["course_textbook_map"][course] = meta["textbook_details"][idx]

    return meta, topics

def _clean_course_folder_name(course):
    """从规划表课程行提取用于输出目录的课程/教材名称。"""
    if not course:
        return ""
    name = str(course).strip()
    name = re.sub(r"^课程[一二三四五六七八九十\d]+\s*[：:]\s*", "", name)
    name = re.sub(r"[（(]\s*约?\d+%\s*[）)]", "", name)
    return name.strip()

def _get_topic_textbook_name(meta, topic):
    """获取当前 topic 对应的教材名；没有教材信息时回退到课程名。"""
    course_map = meta.get("course_textbook_map", {})
    if course_map and topic.get("course") in course_map:
        return course_map[topic["course"]].get("name", "")
    if meta.get("textbook_details"):
        return meta["textbook_details"][0].get("name", "")
    return _clean_course_folder_name(topic.get("course", ""))

def _get_topic_output_base(meta, topic, output_dir):
    """按 生成结果/省份 类别/课程或教材名 定位当前规划表对应输出目录。"""
    province = meta.get("province", "")
    category = meta.get("category", "")
    base = Path(output_dir)
    if province and category:
        base = base / f"{province} {category}"
    elif province:
        base = base / province

    folder_name = _get_topic_textbook_name(meta, topic)
    if folder_name:
        base = base / folder_name
    return base
